from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from datetime import datetime, timedelta, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict


thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

class Quizreport(models.TransientModel):
    _name = 'quiz.report'
    _description = 'Báo cáo kết quả kiểm tra'

    status = fields.Selection([('all', 'Tổng hợp'), ('branch', 'Chi nhánh')], string="Loại báo cáo", default="all")
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    start_date = fields.Date('Ngày báo cáo', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc')
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.onchange('status')
    def _get_company(self):
        for rec in self:
            if rec.status == 'all':
                company_domain = [('id', 'in', self.env.user.company_ids.ids)]
                company_list = self.env['res.company'].sudo().search(company_domain)
                rec.company_ids = company_list
            else:
                rec.company_ids = False

    # check ngày
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        ret_data = []
        domain = [('state', '=', 'done'), ('create_date', '>=', self.start_datetime), ('create_date', '<=', self.end_datetime), ('company_id', 'in', self.company_ids.ids),]
        quiz = self.env['op.quiz.result'].sudo().search(domain)
        stt = 1
        for rec in quiz:
            em = rec.user_id.employee_ids
            ret_data.append({
                "stt": stt,
                "nhan_vien": rec.user_id.name or None,
                "bo_phan":em.group_job.name or None,
                "bai_kiem_tra": rec.name or None,
                "chuyen_de": rec.categ_id.name or None,
                "tong_so_cau_hoi": rec.total_question or None,
                "ket_qua_dung": rec.total_correct or None,
                "ket_qua_sai": rec.total_incorrect or None,
                "diem": rec.score or None,
                "ngay_tao": (rec.create_date + timedelta(hours=7)).strftime('%d/%m/%Y %H:%M') or None,
                "ngay_hoan_thanh": rec.finish_date or None,
                "cong_ty": rec.company_id.name or None,
            })
            stt += 1
        return ret_data

    def create_report(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('openeducat_quiz.bao_cao_ket_qua_thi_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['D2'].value += self.start_date.strftime('%d-%m-%Y')
        ws['E2'].value += self.end_datetime.strftime('%d-%m-%Y')
        key_col_list = list(range(1, 13))

        key_list = [
             "stt",
            "nhan_vien",
            "bo_phan",
            "bai_kiem_tra",
            "chuyen_de",
            "tong_so_cau_hoi",
            "ket_qua_dung",
            "ket_qua_sai",
            "diem",
            "ngay_tao",
            "ngay_hoan_thanh",
            "cong_ty",
        ]
        row = 5
        center_alm = Alignment(horizontal='center', vertical='center')
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(3, col)
                beforeCell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
                cell = ws.cell(row, col)
                for col_1 in range(6,10):
                    center = ws.cell(row, col_1)
                    center.alignment = center_alm
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_ket_qua_thi.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO KẾT QUẢ THI',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
