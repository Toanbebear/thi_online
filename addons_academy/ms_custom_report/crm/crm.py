from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta

all_border_thin = borders.Border(left=borders.Side(style='thin'), right=borders.Side(style='thin'),
                                 top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))

class ReportCRM(models.TransientModel):
    _name = 'sci.report.crm'
    _description = 'Report CRM'

    team_id = fields.Many2one('crm.team',string='Location/shop')
    start_date = fields.Date('Start date')
    end_date = fields.Date('End date')

    @api.constrains('start_date','end_date')
    def check_month(self):
        for record in self:
            if record.end_date < record.start_date:
                raise ValidationError('End date can not before start date')

    def get_data_lead(self):
        leads = self.env['crm.lead'].search([('type','=','lead'),('create_date','<=',self.end_date),('create_date','>=',self.start_date)])
        start_date = self.start_date
        # end_month = self.end_date.month
        # end_of_month = date(self.start_date.year, self.start_date.month,
        #                     monthrange(self.start_date.year, self.start_date.month)[1])
        raw_data = []
        while start_date < self.end_date:
            raw_data.append(start_date.month)
            lead_of_month = leads.filtered(lambda l: l.create_date.month == start_date.month)
            all_teams = lead_of_month.mapped('team_id')
            for team in all_teams:
                leads_of_team = lead_of_month.filtered(lambda l: l.team_id == team)
                # tổng tương tác
                all_interactive = len(leads_of_team)
                number_partner = len(leads_of_team.mapped('partner_id'))
                lead = len(leads_of_team.filtered(lambda r:r.stage_id == self.env.ref('crm_base.stage_sci_lead_2')))
                booking = len(leads_of_team.mapped('opp_id'))
                customer_come = len(leads_of_team.mapped('opp_id').filtered(lambda r:r.customer_come == 'yes'))
                bk_won = len(leads_of_team.mapped('opp_id').filtered(lambda r:r.stage_id == self.env.ref('crm.stage_lead4')))

                if number_partner > 0:
                    tc_DK =(bk_won/number_partner)*100
                    bk_DK = (booking / number_partner) * 100
                else:
                    tc_DK = 'No value'
                    bk_DK = 'No value'
                if booking > 0:
                    DC_bk = (customer_come / booking) * 100
                else:
                    DC_bk ='No value'

                if all_interactive > 0:
                    bk_TTT = (booking/all_interactive)*100
                else:
                    bk_TTT = 'No value'


                tc_ttt =(bk_won/all_interactive)*100
                paid_on = sum(leads_of_team.mapped('opp_id.payment_ids.amount'))
                team_dict = {'name': team.name,
                             'lead': leads_of_team,
                             'all_interactive': all_interactive,
                             'number_partner':number_partner,
                             'lead':lead,
                             'booking':booking,
                             'customer_come':customer_come,
                             'bk_won':bk_won,
                             'paid_on':paid_on,
                             'bk_TTT':bk_TTT,
                             'bk_DK':bk_DK,
                             'DC_bk':DC_bk,
                             'tc_DK':tc_DK,
                             'tc_ttt':tc_ttt,
                             }
                raw_data.append(team_dict)
            start_date = start_date.replace(day=1) + relativedelta(months=1)
        return raw_data


    def report_crm(self):
        crm_attch = self.env['ir.attachment'].browse(self.env.ref('ms_custom_report.crm_report_attachment').id)
        decode = base64.b64decode(crm_attch.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a1'].value = 'Từ ngày: %s đến ngày: %s' % (self.start_date.strftime('%d/%m/%y'), self.end_date.strftime('%d/%m/%y'))

        data = self.get_data_lead()
        keys = ['name','all_interactive','number_partner','lead','booking','customer_come','bk_won','bk_TTT','bk_DK','DC_bk','tc_DK','tc_ttt','paid_on']
        row = 3
        merge_start = row

        for item in data:
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = all_border_thin
            if isinstance(item, int):
                ws.cell(row, 1).value = item
                if row != merge_start:
                    ws.merge_cells(start_row=merge_start,start_column=1,end_row=row-1,end_column=1)
                    ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center',vertical='center')
                    merge_start = row
            else:
                for col,key in enumerate(keys,2):
                    ws.cell(row,col).value = item[key]
                row += 1
            if item == data[-1]:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')


        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'crm report lead',
                                                              'datas_fname': 'reportlead.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})
        return {'name': 'CRM report lead',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

        # url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=crm report.xlsx" \
        #       % (attachment.id)
        # cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        # cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        # return {'name': 'CRM report lead',
        #         'type': 'ir.actions.act_url',
        #         'url': url,
        #         'target': 'self',
        #         }



    def get_data_opp(self):
        opps = self.env['crm.lead'].search([('type','=','opportunity'),('create_date','<=',self.end_date),('create_date','>=',self.start_date)])
        print(opps)
        start_date = self.start_date
        # end_month = self.end_date.month
        # end_of_month = date(self.start_date.year, self.start_date.month,
        #                     monthrange(self.start_date.year, self.start_date.month)[1])
        raw_data = []
        while start_date < self.end_date:
            raw_data.append(start_date.month)
            opps_of_month = opps.filtered(lambda l: l.create_date.month == start_date.month)
            all_teams = opps.mapped('team_id')
            for team in all_teams:
                opps_of_team = opps_of_month.filtered(lambda l: l.team_id == team)
                not_confirm = len(opps_of_team.filtered(lambda r:r.stage_id == self.env.ref('crm_base.stage_sci_not_confirm')))
                confirm = len(opps_of_team.filtered(lambda r:r.stage_id == self.env.ref('crm_base.stage_sci_confirm')))
                cancel = len(opps_of_team.filtered(lambda r:r.stage_id.name == 'Cancel'))
                outsold = len(opps_of_team.filtered(lambda r:r.stage_id.name == 'Out sold'))
                customer_come = len(opps_of_team.filtered(lambda r:r.customer_come == 'yes'))
                bk_won = len(opps_of_team.filtered(lambda r:r.stage_id == self.env.ref('crm.stage_lead4')))
                paid_on = sum(opps_of_team.mapped('payment_ids.amount'))

                if customer_come > 0:
                    won_ondoor = (bk_won/customer_come)*100
                else:
                    won_ondoor = '∞'

                team_dict = {'name': team.name,
                             'not_confirm':not_confirm,
                             'confirm':confirm,
                             'cancel':cancel,
                             'out_sold':outsold,
                             'customer_come':customer_come,
                             'bk_won':bk_won,
                             'paid_on':paid_on,
                             'won_ondoor':won_ondoor,
                             }
                raw_data.append(team_dict)
            start_date = start_date.replace(day=1) + relativedelta(months=1)
        return raw_data

    def report_crm_opp(self):
        crm_attch = self.env['ir.attachment'].browse(self.env.ref('ms_custom_report.crm_report_attachment_opp').id)
        decode = base64.b64decode(crm_attch.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a1'].value = 'Từ ngày: %s đến ngày: %s' %(
        self.start_date.strftime('%d/%m/%y'), self.end_date.strftime('%d/%m/%y'))

        data = self.get_data_opp()
        keys = ['name','not_confirm','confirm','cancel','out_sold','customer_come','bk_won','won_ondoor','paid_on']
        row = 3
        merge_start = row

        for item in data:
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = all_border_thin
            if isinstance(item, int):
                ws.cell(row, 1).value = item
                if row != merge_start:
                    ws.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
                    # Todo
                    ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    merge_start = row
            else:
                for col, key in enumerate(keys, 2):
                    ws.cell(row, col).value = item[key]
                row += 1
            if item == data[-1]:
                ws.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'crm report booking',
                                                              'datas_fname': 'reportopp.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        # url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=crm report.xlsx" \
        #       % (attachment.id)
        # cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        # cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        # return {'name': 'CRM report opp',
        #         'type': 'ir.actions.act_url',
        #         'url': url,
        #         'target': 'self',
        #         }

        return {'name': 'CRM report booking',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}