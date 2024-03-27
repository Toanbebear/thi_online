# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta

import logging

_logger = logging.getLogger(__name__)

all_border_thin = borders.Border(left=borders.Side(style='thin'), right=borders.Side(style='thin'),
                                 top=borders.Side(style='thin'), bottom=borders.Side(style='thin'))


class MonthlyInventory(models.TransientModel):
    _name = 'monthly.inventory'
    _description = 'Monthly Inventory'

    location = fields.Many2one('stock.location', 'Stock location')
    report_type = fields.Selection([('current', 'Current inventory'), ('period', 'Period Inventory'),
                                    ('product', 'Product stock cards')], 'Report type', default='current')
    products = fields.Many2many('product.product')
    product_domain = fields.Many2many('product.product', compute='_get_product_domain')
    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')

    @api.depends('location')
    def _get_product_domain(self):
        for record in self:
            record.product_domain = self.env['stock.quant'].search([('location_id', 'child_of', record.location.id)]).mapped('product_id')

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def _get_range_inventory_data(self):
        all_quant = self.env['stock.quant'].search([('location_id', 'child_of', self.location.id)])
        moves_in = self.env['stock.move'].search([('date', '>=', self.start_date),
                                                  ('state', '=', 'done'),
                                                  ('location_dest_id', 'child_of', self.location.id),
                                                  '!', ('location_id', 'child_of', self.location.id)])
        moves_out = self.env['stock.move'].search([('date', '>=', self.start_date),
                                                   ('state', '=', 'done'),
                                                   ('location_id', 'child_of', self.location.id),
                                                   '!', ('location_dest_id', 'child_of', self.location.id)])
        products = self.env['product.product'].union(all_quant.mapped('product_id'), moves_in.mapped('product_id'), moves_out.mapped('product_id'))
        categories = products.mapped('categ_id')
        product_datas = []

        def _calculate_mixed_quantity(product, product_moves):
            identical_uom_moves = product_moves.filtered(lambda m: m.product_uom == product.uom_id)
            res = sum(identical_uom_moves.mapped('product_uom_qty'))
            if identical_uom_moves != product_moves:
                different_uom_moves = [move for move in product_moves if move not in identical_uom_moves]
                for move in different_uom_moves:
                    res += move.product_uom._compute_quantity(move.product_uom_qty, product.uom_id)
            return res

        for category in categories:
            product_datas.append(category.name)
            for product in products.filtered(lambda p: p.categ_id == category):
                product_data = {'name': product.name, 'code': product.code, 'uom': product.uom_id.name,
                                'begin': 0.0, 'begin_value': 0.0,
                                'in': 0.0, 'in_value': 0.0,
                                'out': 0.0, 'out_value': 0.0,
                                'end': 0.0, 'end_value': 0.0}
                product_datas.append(product_data)
                # calculate moves_in quantity in the period
                product_moves_in = moves_in.filtered(lambda m: m.product_id == product and m.date.date() <= self.end_date)
                product_data['in'] = _calculate_mixed_quantity(product, product_moves_in)
                product_data['in_value'] = product_data['in'] * product.standard_price
                # calculate moves_out quantity in the period
                product_moves_out = moves_out.filtered(lambda m: m.product_id == product and m.date.date() <= self.end_date)
                product_data['out'] = _calculate_mixed_quantity(product, product_moves_out)
                product_data['out_value'] = product_data['out'] * product.standard_price
                # calculate stock quantity at the end of period: current stock + moves_out quantity after end date
                # - moves_in quantity after end date
                product_data['end'] = sum(all_quant.filtered(lambda q: q.product_id == product).mapped('quantity'))
                if self.end_date < fields.date.today():
                    later_product_moves_in = moves_in.filtered(lambda m: m.product_id == product and m.date.date() > self.end_date)
                    later_moves_in_qty = _calculate_mixed_quantity(product, later_product_moves_in)
                    later_product_moves_out = moves_out.filtered(lambda m: m.product_id == product and m.date.date() > self.end_date)
                    later_moves_out_qty = _calculate_mixed_quantity(product, later_product_moves_out)
                    product_data['end'] += (later_moves_out_qty - later_moves_in_qty)
                product_data['end_value'] = product_data['end'] * product.standard_price
                # calculate begin quantity
                product_data['begin'] = product_data['end'] + product_data['out'] - product_data['in']
                product_data['begin_value'] = product_data['begin'] * product.standard_price
        # product_datas format: [category_1,
        #                       product_1_of_categ_1 dict with keys: 'code', 'name', 'uom', 'begin', 'begin_value', 'in', 'in_value', 'out', 'out_value', 'end', 'end_value',
        #                       product_2_of_categ_1 dict,
        #                       category_2, product_1_of_categ_2 dict, product_2_of_categ_2 dict, ...]
        return product_datas

    def _get_current_inventory_data(self):
        all_quant = self.env['stock.quant'].search([('location_id', 'child_of', self.location.id)])
        categories = all_quant.mapped('product_id.categ_id')
        product_datas = []
        for category in categories:
            product_datas.append(category.name)
            lots_track = {}
            category_quant = all_quant.filtered(lambda q: q.product_id.categ_id == category)
            for quant in category_quant:
                lots_track_name = str(quant.product_id.id) + '-' + str(quant.lot_id.id or 0)
                if not lots_track.get(lots_track_name, False):
                    product_data = {'code': quant.product_id.code, 'name': quant.product_id.name, 'uom': quant.product_uom_id.name,
                                    'exp_date': quant.lot_id.name or " ", 'qty': quant.quantity}
                    product_datas.append(product_data)
                    lots_track[lots_track_name] = len(product_datas) - 1
                else:
                    product_datas[lots_track[lots_track_name]]['qty'] += quant.quantity
        return product_datas

    def _get_stock_cards_data(self):
        all_quant = self.env['stock.quant'].search([('location_id', 'child_of', self.location.id), ('product_id', 'in', self.products.ids)])
        moves_in = self.env['stock.move.line'].search([('date', '>=', self.start_date),
                                                       ('state', '=', 'done'),
                                                       ('product_id', 'in', self.products.ids),
                                                       ('location_dest_id', 'child_of', self.location.id),
                                                       '!', ('location_id', 'child_of', self.location.id)])
        moves_out = self.env['stock.move.line'].search([('date', '>=', self.start_date),
                                                        ('state', '=', 'done'),
                                                        ('product_id', 'in', self.products.ids),
                                                        ('location_id', 'child_of', self.location.id),
                                                        '!', ('location_dest_id', 'child_of', self.location.id)])
        products = self.products
        product_datas = {}
        for product in products:
            product_moves = (moves_in + moves_out)\
                .filtered(lambda m: m.product_id == product and m.date.date() <= self.end_date)\
                .sorted('date')
            # calculate quantity at self.end_date
            product_quant = sum(all_quant.filtered(lambda q: q.product_id == product).mapped('quantity'))
            if self.end_date < fields.date.today():
                later_product_moves = (moves_in + moves_out)\
                                        .filtered(lambda m: m.product_id == product and m.date.date() > self.end_date)
                for move in later_product_moves:
                    qty_done = move.product_uom_id._compute_quantity(move.qty_done, product.uom_id) if move.product_uom_id != move.product_id.uom_id else move.qty_done
                    if move in moves_out:
                        product_quant += qty_done
                    else:
                        product_quant -= qty_done

            product_datas[product.name] = [product.code, product.uom_id.name, product_quant]
            index = 1
            for move in product_moves:
                qty_done = move.product_uom_id._compute_quantity(move.qty_done, product.uom_id) if move.product_uom_id != move.product_id.uom_id else move.qty_done
                move_in = 0 if move in moves_out else qty_done
                move_out = 0 if move in moves_in else qty_done
                move_data = {'index': index, 'date': move.date, 'ref': move.reference, 'in': move_in, 'out': move_out}  # move_dict format here
                product_datas[product.name].append(move_data)
                index += 1

        # product_datas format: {'product_1': [code, uom, quant at end date, move_1_dict, move_2_dict...],
        #                           'product_2': [code, uom, quant at end date, move_1_dict, move_2_dict...]}
        return product_datas

    def report_inventory(self):
        inventory_attachment = self.env['ir.attachment'].browse(self.env.ref('ms_custom_report.inventory_report_attachment').id)
        decode = base64.b64decode(inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a5'].value = 'Từ ngày: %s đến ngày: %s' % (self.start_date.strftime('%d/%m/%y'), self.end_date.strftime('%d/%m/%y'))
        ws['o6'].value = self.location.display_name
        data = self._get_range_inventory_data()
        row = 10
        for item in data:
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = all_border_thin
                ws.cell(row=row, column=col).font = Font(name='Arial', size=10)
            for col in range(5, 13):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            if isinstance(item, str):
                item_code = item.split()[0][:2].upper() if len(item.split()) == 1 else item.split()[0][0].upper() + item.split()[1][0].upper()
                ws.cell(row=row, column=1).value = 'Nhóm hàng: ' + item_code + '-' + item
                ws.cell(row=row, column=1).font = Font(bold=True)
            else:
                keys = ['code', 'name', 'uom', 'begin', 'begin_value', 'in', 'in_value', 'out', 'out_value', 'end', 'end_value']
                for col, key in enumerate(keys, 2):
                    ws.cell(row=row, column=col).value = item[key]
            row += 1
        ws.cell(row=row, column=1).value, ws.cell(row=row, column=1).font = 'Tổng: ', Font(name='Arial', size=10)
        ws.cell(row=row+2, column=9).value = 'Ngày %s tháng %s năm %s ' % (fields.date.today().strftime('%d'), fields.date.today().strftime('%m'), fields.date.today().strftime('%Y'))
        ws.cell(row=row + 2, column=9).font, ws.cell(row=row + 2, column=9).alignment = Font(name='Arial', size=10), Alignment(horizontal='center')
        ws.cell(row=row+4, column=3).value, ws.cell(row=row+5, column=3).value = 'Kế toán ', '(Ký, họ tên)'
        ws.cell(row=row + 4, column=3).font = ws.cell(row=row + 5, column=3).font = Font(name='Arial', size=10)
        ws.cell(row=row + 4, column=9).value, ws.cell(row=row + 5, column=9).value = 'Người lập', '(Ký, họ tên)'
        ws.cell(row=row + 4, column=9).font = ws.cell(row=row + 5, column=9).font = Font(name='Arial', size=10)
        ws.cell(row=row + 4, column=9).alignment = ws.cell(row=row + 5, column=9).alignment = Alignment(horizontal='center')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                             'datas_fname': 'Inventory report.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})
        return {'name': 'Inventory report',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

        # url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Inventory report.xlsx" \
        #       % (attachment.id)
        # cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        # cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        # return {'name': 'Inventory report',
        #         'type': 'ir.actions.act_url',
        #         'url': url,
        #         'target': 'self',
        #         }

    def report_current_inventory(self):
        current_inventory_attachment = self.env['ir.attachment'].browse(
            self.env.ref('ms_custom_report.current_inventory_report_attachment').id)
        decode = base64.b64decode(current_inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a5'].value = 'Ngày %s tháng %s năm %s' % (fields.date.today().strftime('%d'), fields.date.today().strftime('%m'), fields.date.today().strftime('%Y'))
        ws['a6'].value = 'Kho: ' + self.location.display_name
        data = self._get_current_inventory_data()
        row = 10
        for item in data:
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = all_border_thin
                ws.cell(row=row, column=col).font = Font(name='Arial', size=10)
            if isinstance(item, str):
                item_code = item.split()[0][:2].upper() if len(item.split()) == 1 else item.split()[0][0].upper() + item.split()[1][0].upper()
                ws.cell(row=row, column=1).value = 'Nhóm hàng: ' + item_code + '-' + item
                ws.cell(row=row, column=1).font = Font(bold=True)
            else:
                keys = ['code', 'name', 'uom', 'exp_date', 'qty']
                for num, key in enumerate(keys, 2):
                    ws.cell(row=row, column=num).value = item[key]
            row += 1

        ws.cell(row=row, column=1).value, ws.cell(row=row, column=1).font = 'Tổng: ', Font(name='Arial', size=10)
        ws.cell(row=row + 4, column=3).value, ws.cell(row=row + 5, column=3).value = 'Kế toán ', '(Ký, họ tên)'
        ws.cell(row=row + 4, column=3).font = ws.cell(row=row + 5, column=3).font = Font(name='Arial', size=10)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                             'datas_fname': 'Current inventory report.xlsx',
                                                             'datas': report,
                                                             'res_model': 'temp.creation',
                                                             'public': True})
        return {'name': 'Current inventory report',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

        # url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Current inventory report.xlsx" \
        #       % (attachment.id)
        # cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        # cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        # return {'name': 'Inventory report',
        #         'type': 'ir.actions.act_url',
        #         'url': url,
        #         'target': 'self',
        #         }

    def report_stock_cards(self):
        current_inventory_attachment = self.env['ir.attachment'].browse(
            self.env.ref('ms_custom_report.stock_cards_report_attachment').id)
        decode = base64.b64decode(current_inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        data = self._get_stock_cards_data()
        for key in data.keys():
            key = key.translate({ord(i): None for i in '\/*[]:?'})
        if len(data.keys()) > 1:
            for i in range(len(data.keys()) - 1):
                wb.copy_worksheet(wb.active)
        for ws, key in zip(wb.worksheets, data.keys()):
            ws.title = key
            ws['a4'].value = "THẺ KHO\ntừ ngày %s đến ngày %s" % (self.start_date.strftime('%d/%m/%y'), self.end_date.strftime('%d/%m/%y'))
            ws['c5'].value, ws['c6'].value = self.location.display_name, self.location.name
            ws['c7'].value, ws['f7'].value = data[key][0], data[key][1]
            row = 10
            for i in range(3, len(data[key])):
                # set font and border to column from A to F (1-6)
                for j in range(1, 7):
                    ws.cell(row=row, column=j).font = Font(name='Times New Roman')
                    ws.cell(row=row, column=j).border = all_border_thin
                # set alignment for column with number value
                for k in [1, 4, 5, 6]:
                    ws.cell(row=row, column=k).alignment = Alignment(horizontal='center')
                # set value for row
                for num, sub_key in enumerate(['index', 'date', 'ref', 'in', 'out'], 1):
                    if sub_key != 'date':
                        ws.cell(row=row, column=num).value = data[key][i][sub_key]
                    else:
                        ws.cell(row=row, column=num).value = data[key][i][sub_key].strftime("%d/%m/%y %H:%M")
                row += 1
            if len(data[key]) > 3:
                ws.cell(row=row - 1, column=6).value = data[key][2]
                for i in range(2, len(data[key]) - 2):  # calculate end of each move base on the last end amount and in out qty
                    ws.cell(row=row - i, column=6).value = ws.cell(row=row - i + 1, column=6).value +\
                                                           ws.cell(row=row - i + 1, column=5).value -\
                                                           ws.cell(row=row - i + 1, column=4).value
            else:
                ws.cell(row=row, column=6).value = data[key][2]
            ws.cell(row=row, column=1).value, ws.cell(row=row, column=1).font = 'Tổng cộng:', Font(name='Times New Roman', bold=True)
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                             'datas_fname': 'Stock cars.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})
        return {'name': 'Stock cards',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

        # url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Stock cards.xlsx" \
        #       % (attachment.id)
        # cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        # cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        # return {'name': 'Stock cards',
        #         'type': 'ir.actions.act_url',
        #         'url': url,
        #         'target': 'self',
        #         }
