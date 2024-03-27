##############################################################################
#    Copyright (C) 2018 shealth (<http://scigroup.com.vn/>). All Rights Reserved
#    shealth, Hospital Management Solutions

# Odoo Proprietary License v1.0
#
# This software and associated files (the "Software") may only be used (executed,
# modified, executed after modifications) if you have purchased a valid license
# from the authors, typically via Odoo Apps, shealth.in, openerpestore.com, or if you have received a written
# agreement from the authors of the Software.
#
# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without copying
# any source code or material from the Software. You may distribute those
# modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).
#
# It is forbidden to publish, distribute, sublicense, or sell copies of the Software
# or modified copies of the Software.
#
# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.

##############################################################################
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError,ValidationError,AccessError
from lxml import etree
import json
from odoo.tools.float_utils import float_round, float_compare
from odoo.osv import expression
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.page import (
    PrintPageSetup,
    PageMargins,
    PrintOptions,
)
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
import base64
from io import BytesIO
import pytz
from num2words import num2words

import logging

_logger = logging.getLogger(__name__)

# Inherit Stock

class SHealthStockLocation(models.Model):
    _inherit = 'stock.location'

    location_institution_type = fields.Selection([('medicine', 'Tủ Thuốc'), ('supply', 'Vật tư')], string='Loại tủ')


class SHealthStockPicking(models.Model):
    _inherit = 'stock.picking'

    def _get_default_picking_type(self):
        if self.env.context.get('view_for') in ['picking_int','picking_int_vt','picking_int_return']:
            return self.env['stock.picking.type'].sudo().search([('company_id', '=', self.env.company.id), ('code', '=', 'internal')], limit=1)
        elif self.env.context.get('view_for') == 'picking_out_company':
            return self.env['stock.picking.type'].sudo().search([('code', '=', 'outgoing'), ('company_id', '=', self.env.company.id)], limit=1)

    def _get_picking_type_domain(self):
        if self.env.context.get('view_for') == 'picking_out_company':
            return [('code', '=', 'outgoing'), ('company_id', '=', self.env.company.id)]

    scheduled_date = fields.Datetime(default=lambda self: fields.Datetime.now())
    sci_date_done = fields.Datetime('Ngày hoàn thành', copy=False, help="Ngày hoàn thành nhập kho SCI", default=lambda self: fields.Datetime.now())
    hide_partner = fields.Boolean('Ẩn đối tác', compute="_compute_hide_partner")
    patient_id = fields.Many2one('sh.medical.patient', string='Bệnh nhân', help="Bệnh nhân")

    picking_type_id = fields.Many2one('stock.picking.type', default=lambda self: self._get_default_picking_type(), domain=lambda self: self._get_picking_type_domain())

    # SonZ
    def total_price_text(self, total_price):
        if total_price and total_price > 0:
            total_price = num2words(int(total_price), lang='vi_VN') + " đồng"
        elif total_price and total_price < 0:
            raise ValidationError(_('Số tiền thanh toán không hợp lệ!!!'))
        else:
            total_price = "Không đồng"
        return total_price.capitalize()
    # SonZ

    def button_validate(self):
        institution = self.env['sh.medical.health.center'].sudo().search(
            [('his_company', '=', self.company_id.id)], limit=1)

        if institution and self.location_dest_id == institution.warehouse_ids[0].lot_stock_id:
            med_loc = institution.location_medicine_stock
            supply_loc = institution.location_medicine_stock

            for move_line in self.move_ids_without_package:
                for move_line_detail in move_line.move_line_ids:
                    # check nếu là thuốc hay vật tư thì chuyển thẳng về tủ tương ứng
                    if move_line_detail.product_id.categ_id == self.env.ref('shealth_all_in_one.sh_medicines'):
                        move_line_detail.location_dest_id = med_loc.id or institution.warehouse_ids[0].lot_stock_id.id
                    elif move_line_detail.product_id.categ_id == self.env.ref('shealth_all_in_one.sh_supplies'):
                        move_line_detail.location_dest_id = supply_loc.id or institution.warehouse_ids[0].lot_stock_id.id
                    move_line_detail.qty_done = move_line_detail.product_uom_qty

        return super(SHealthStockPicking, self).button_validate()

    def action_assign(self):
        """Adjustment to bypass waiting state"""
        picks = self.filtered(lambda p: p.state == 'waiting')
        picks.mapped('move_lines').write({'move_orig_ids': [(5, 0, 0)], 'procure_method': 'make_to_stock'})
        return super(SHealthStockPicking, self).action_assign()

    def action_cancel(self):
        # KHÔNG CHO HỦY PHIẾU CHƯA CÓ SẢN PHẨM NÀO DC CHỌN
        for picking in self:
            if picking.move_ids_without_package or self.env.context.get('no_check_quant') or self.env.uid == 1:
                return super(SHealthStockPicking, self).action_cancel()
            else:
                raise UserError(_('Phiếu của bạn không có sản phẩm nào để hủy bỏ!'))

    def action_confirm(self):
        # KHÔNG CHO XÁC NHẬN PHIẾU CHƯA CÓ SẢN PHẨM NÀO DC CHỌN
        for picking in self:
            if picking.move_ids_without_package or self.env.context.get('no_check_quant') or self.env.uid == 1:
                return super(SHealthStockPicking, self).action_confirm()
            else:
                raise UserError(_('Phiếu của bạn không có sản phẩm nào để xác nhận!'))

    def _log_less_quantities_than_expected(self, moves):
        # Fix lỗi log khi không có move
        if not moves:
            return
        return super(SHealthStockPicking, self)._log_less_quantities_than_expected(moves)

    def write(self, vals):
        if self.env.context.get('separate_pick') and vals.get('origin') is False:  # dùng để tách phiếu cơ số, xem phần stock.move
            vals.pop('origin')
        res = super(SHealthStockPicking, self).write(vals)
        for record in self:
            if record.purchase_id:
                for ml in record.move_line_ids:
                    if not ml.lot_id:
                        date_done = ml.picking_id.sci_date_done if ml.picking_id.sci_date_done else ml.create_date
                        date_done += timedelta(hours=7)  # adjust to Vietnamese time GMT + 7
                        ml.lot_name = (ml.product_id.default_code or ml.product_id.name) + '-' + date_done.strftime('%Y%m%d%H%M')
        return res

    def download_excel(self):
        for record in self:
            if record.state != 'cancel':
                self.action_cancel()
        return {
            'name': 'Nhu cầu nhập hàng',
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'temp.wizard',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'target': 'inline',
            'context': {'default_template_id': self.env.ref('shealth_all_in_one.nhu_cau_nhap_hang').id},
        }

    @api.depends('picking_type_id')
    def _compute_hide_partner(self):
        for pick in self:
            if pick.picking_type_id.code == 'incoming':
                pick.hide_partner = True
            else:
                pick.hide_partner = False

    @api.onchange('location_id')
    def onchange_location_id(self):
        if self.env.context.get('view_for', True) == 'picking_int_return':
            # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
            institution = self.env['sh.medical.health.center'].sudo().search(
                [('his_company', '=', self.env.companies.ids[0])], limit=1)

            dest_location = self.env['stock.location'].sudo().search([('company_id', '=', institution.his_company.id),
                                                                      ('usage', '=', 'internal'),('name', 'not ilike', 'cấp cứu'),('name', 'not ilike', 'bán'),
                                                                      ('location_id', '=',institution.warehouse_ids[0].lot_stock_id.id),
                                                                      ('location_institution_type', '=',
                                                                       self.location_id.location_institution_type),
                                                                      ('child_ids', '=', False)], limit=1)
            self.location_dest_id = dest_location.id

    @api.onchange('picking_type_id')
    def onchange_picking_type(self):
        super(SHealthStockPicking, self).onchange_picking_type()
        if self.env.context.get('view_for', True) == 'picking_int':
            # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
            institution = self.env['sh.medical.health.center'].sudo().search(
                [('his_company', '=', self.env.companies.ids[0])], limit=1)

            if not institution:
                raise ValidationError(_('Công ty hiện tại của bạn không có Cơ sở y tế! Hãy chọn chính xác công ty!'))

            self.location_id = institution.location_medicine_stock.id
            self.location_dest_id = False
            # self.location_id = self.env.ref(
            #     'shealth_all_in_one.sh_location_medicine_prescription_knhn').id
        elif self.env.context.get('view_for', True) == 'picking_int_vt':
            # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
            institution = self.env['sh.medical.health.center'].sudo().search(
                [('his_company', '=', self.env.companies.ids[0])], limit=1)

            supply_loc = self.env['stock.location'].sudo().search(
                [('location_id', '=', institution.warehouse_ids[0].lot_stock_id.id),('location_institution_type','=','supply'),('name', 'not ilike', 'cấp cứu')], limit=1)

            self.location_id = supply_loc.id
            self.location_dest_id = False
        elif self.env.context.get('view_for', True) == 'picking_int_return':
            # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
            institution = self.env['sh.medical.health.center'].sudo().search(
                [('his_company', '=', self.env.companies.ids[0])], limit=1)

            access_location = []
            receptionist_ward = lab_ward = spa_ward = laser_ward = surgery_ward = odontology_ward = False

            if institution:
                receptionist_ward = self.env['stock.location'].sudo().search(
                    [('company_id', '=', institution.his_company.id), ('name', 'ilike', 'lễ tân')], limit=1)
                lab_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Laboratory')], limit=1)
                spa_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Spa')], limit=1)
                laser_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Laser')], limit=1)
                surgery_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Surgery')], limit=1)
                odontology_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Odontology')], limit=1)

            grp_loc_dict = {
                'shealth_all_in_one.group_sh_medical_receptionist': receptionist_ward,
                'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': lab_ward,
                'shealth_all_in_one.group_sh_medical_physician_surgery': surgery_ward,
                'shealth_all_in_one.group_sh_medical_physician_odontology': odontology_ward,
                'shealth_all_in_one.group_sh_medical_physician_spa': spa_ward,
                'shealth_all_in_one.group_sh_medical_physician_laser': laser_ward}

            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp) and loc:
                    if grp == 'shealth_all_in_one.group_sh_medical_receptionist':
                        access_location.append(loc.id)
                    else:
                        access_location.append(loc.location_id.id)

            # quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

            loc = self.env['stock.location'].sudo().browse(access_location[0])

            out_location = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('usage', '=', 'internal'),
                 ('location_id', '=', loc.id),('child_ids','=',False),('location_institution_type','in',['medicine','supply'])], limit=1)

            self.location_id = loc.id if not out_location else out_location.id
            dest_location = self.env['stock.location'].sudo().search([('company_id','=',institution.his_company.id),
                  ('usage','=','internal'),('location_id','=',institution.warehouse_ids[0].lot_stock_id.id),
                  ('location_institution_type','=',loc.location_institution_type if not out_location else out_location.location_institution_type),('child_ids','=',False)],limit=1)
            self.location_dest_id = dest_location.id

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(SHealthStockPicking, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,
                                                      submenu=submenu)

        warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.companies.ids[0])], limit=1)

        # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
        institution = self.env['sh.medical.health.center'].search([('his_company', '=', self.env.companies.ids[0])],
                                                                         limit=1)

        # print(self.env.companies.ids[0])
        # print(institution)
        access_location = []
        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            location_stock0 = institution.warehouse_ids[0].lot_stock_id.id if institution else warehouse.lot_stock_id.id # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            receptionist_ward = lab_ward = spa_ward = laser_ward = surgery_ward = odontology_ward = False

            if institution:
                receptionist_ward = self.env['stock.location'].sudo().search(
                    [('company_id', '=', institution.his_company.id), ('name', 'ilike', 'lễ tân')], limit=1)
                lab_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Laboratory')], limit=1)
                spa_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Spa')], limit=1)
                laser_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Laser')], limit=1)
                surgery_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Surgery')], limit=1)
                odontology_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Odontology')], limit=1)

            grp_loc_dict = {
                'shealth_all_in_one.group_sh_medical_receptionist': receptionist_ward,
                'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': lab_ward,
                'shealth_all_in_one.group_sh_medical_physician_surgery': surgery_ward,
                'shealth_all_in_one.group_sh_medical_physician_odontology': odontology_ward,
                'shealth_all_in_one.group_sh_medical_physician_spa': spa_ward,
                'shealth_all_in_one.group_sh_medical_physician_laser': laser_ward}

            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp) and loc:
                    if grp == 'shealth_all_in_one.group_sh_medical_receptionist':
                        access_location.append(loc.id)
                    else:
                        access_location.append(loc.location_id.id)

            # quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        doc = etree.XML(res['arch'])

        for t in doc.xpath("//" + view_type):
            # t.attrib['delete'] = 'false'
            t.attrib['duplicate'] = 'false'

        if self.env.context.get('view_for', True) == "picking_in":
            #ko cho tao phieu
            # for t in doc.xpath("//" + view_type):
            #     t.attrib['create'] = 'false'

            for node in doc.xpath("//field[@name='date_done']"):
                node.set('string', 'Ngày nhập kho')
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('string', 'Nhà cung cấp')
            for node in doc.xpath("//field[@name='origin']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='location_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='purchase_id']"):
                node.set('column_invisible', '0')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = False
                node.set("modifiers", json.dumps(modifiers))
            # for node in doc.xpath("//field[@name='picking_type_id']"):
            #     node.set('readonly', '1')
            #     modifiers = json.loads(node.get("modifiers"))
            #     modifiers['readonly'] = True
            #     node.set("modifiers", json.dumps(modifiers))
        # elif self.env.context.get('view_for', True) in ["picking_int","picking_int_return"]:
        elif self.env.context.get('view_for', True) in ["picking_int","picking_int_vt"]:
            #neu quan ly kho duoc
            # if self.env.user.has_group('shealth_all_in_one.group_sh_medical_document') \
            #         or self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            #     #cho phép sửa phiếu
            #     for t in doc.xpath("//" + view_type):
            #         t.attrib['edit'] = 'true'
            # else:
            #     # ko cho phép sửa phiếu
            #     for t in doc.xpath("//" + view_type):
            #         t.attrib['edit'] = 'false'
            for node in doc.xpath("//field[@name='picking_type_id']"):
                node.set('readonly', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['readonly'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='location_id']"):
                # node.set('readonly', 'True')
                if self.env.context.get('view_for', True) == 'picking_int':
                    node_domain = "[('id', '=', %d)]" % institution.location_medicine_stock.id
                else:
                    supply_loc = self.env['stock.location'].sudo().search(
                        [('location_id', '=', institution.warehouse_ids[0].lot_stock_id.id),
                         ('location_institution_type', '=', 'supply'),('name', 'not ilike', 'cấp cứu')], limit=1)

                    node_domain = "[('id', '=', %d)]" % supply_loc.id
                node.set('domain', node_domain)
                modifiers = json.loads(node.get("modifiers"))
                modifiers['domain'] = node_domain
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='location_dest_id']"):
                if self.env.context.get('view_for', True) == 'picking_int':
                    node_domain = "[('location_id', 'child_of', %s),('location_institution_type', '=', 'medicine')]" % str(access_location)
                else:
                    node_domain = "[('location_id', 'child_of', %s),('location_institution_type', '=', 'supply')]" % str(access_location)

                node.set('domain', node_domain)
                modifiers = json.loads(node.get("modifiers"))
                modifiers['domain'] = node_domain
                node.set("modifiers", json.dumps(modifiers))
        elif self.env.context.get('view_for', True) in ["picking_int_return"]:
            for node in doc.xpath("//field[@name='picking_type_id']"):
                node.set('readonly', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['readonly'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))

            for node in doc.xpath("//field[@name='location_id']"):
                node_domain = "[('location_id', 'child_of', %s),('location_id.usage', '=', 'internal'),('name', 'ilike', 'Tủ')]" % (
                    str(access_location))
                node.set("domain", node_domain)

            for node in doc.xpath("//field[@name='location_dest_id']"):
                node_domain_dest = "[('company_id', '=', %s), ('location_id', '=', %s), ('name', 'not ilike', 'Cấp cứu')," \
                                   "('name', 'not ilike', 'bán'),('location_id.usage', '=', 'internal'),('location_institution_type','in',['medicine','supply'])]" \
                                   % (str(institution.his_company.id), str(institution.warehouse_ids[0].lot_stock_id.id))
                node.set("domain", node_domain_dest)
                node.set("readonly", 'True')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['readonly'] = True
                node.set("modifiers", json.dumps(modifiers))

        elif self.env.context.get('view_for', True) == "picking_out_bn":
            for t in doc.xpath("//" + view_type):
                t.attrib['create'] = 'false'

            for node in doc.xpath("//field[@name='picking_type_id']"):
                node.set('readonly', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['readonly'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='patient_id']"):
                node.set('column_invisible', '0')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['invisible'] = False
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='location_dest_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))

        elif self.env.context.get('view_for', True) == "picking_out_company":
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('required', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['required'] = True
                node.set("modifiers", json.dumps(modifiers))
                node.set('string', 'Chi nhánh')
                all_companies_partners = self.env['res.company'].sudo().search([('enable_inter_company_transfer', '=', True), ('apply_transfer_type', 'in', ['all', 'incoming']),
                                                                               ('id', '!=', self.env.company.id)]).mapped('partner_id.id')
                node.set("domain", "[('id', 'in', %s)]" % all_companies_partners)
            for node in doc.xpath("//field[@name='location_id']"):
                node_domain = "[('usage', '=', 'internal'), ('company_id', '=', %s)]" % self.env.company.id
                node.set("domain", node_domain)

        elif self.env.context.get('view_for', True) == "picking_in_company":
            for t in doc.xpath("//" + view_type):
                t.attrib['create'] = 'false'

            for node in doc.xpath("//field[@name='picking_type_id']"):
                node.set('readonly', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['readonly'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='patient_id']"):
                node.set('column_invisible', '0')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = False
                node.set("modifiers", json.dumps(modifiers))
        elif self.env.context.get('view_for', True) == "picking_out_sale":
            for t in doc.xpath("//" + view_type):
                t.attrib['create'] = 'false'

            for node in doc.xpath("//field[@name='picking_type_id']"):
                node.set('readonly', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['readonly'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('column_invisible', '1')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = True
                node.set("modifiers", json.dumps(modifiers))
            for node in doc.xpath("//field[@name='patient_id']"):
                node.set('column_invisible', '0')
                modifiers = json.loads(node.get("modifiers"))
                modifiers['column_invisible'] = False
                node.set("modifiers", json.dumps(modifiers))

        res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    #overide hàm này để ghi nhận lại ngày hoàn thành
    def action_done(self):
        res = super(SHealthStockPicking, self).action_done()

        for stock_picking in self:
            if stock_picking.sci_date_done:
                for move_line in stock_picking.move_ids_without_package:
                    move_line.move_line_ids.write({'date': stock_picking.sci_date_done})  # sửa ngày hoàn thành ở stock move line
                stock_picking.move_ids_without_package.write(
                    {'date': stock_picking.sci_date_done})  # sửa ngày hoàn thành ở stock move

                stock_picking.write({'date_done': stock_picking.sci_date_done})
            else:
                stock_picking.write({'sci_date_done': stock_picking.date_done})

        return res

    def view_current_stock(self):
        #mặc định xem tủ kê đơn và cấp cứu vs tất cả các quyền
        access_location = []

        # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
        institution = self.env['sh.medical.health.center'].sudo().search([('his_company', '=', self.env.companies.ids[0])], limit=1)

        if not institution:
            raise ValidationError(_('Công ty hiện tại của bạn không có Cơ sở y tế! Hãy chọn chính xác công ty!'))

        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            current_warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.companies.ids[0])], limit=1)
            location_stock0 = current_warehouse.lot_stock_id.id  # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            receptionist_ward = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('name', 'ilike', 'lễ tân')], limit=1)
            lab_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Laboratory')], limit=1)
            spa_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Spa')], limit=1)
            laser_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Laser')], limit=1)
            surgery_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Surgery')], limit=1)
            odontology_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Odontology')], limit=1)

            grp_loc_dict = {'shealth_all_in_one.group_sh_medical_receptionist': receptionist_ward,
                            'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': lab_ward,
                            'shealth_all_in_one.group_sh_medical_physician_surgery': surgery_ward,
                            'shealth_all_in_one.group_sh_medical_physician_odontology': odontology_ward,
                            'shealth_all_in_one.group_sh_medical_physician_spa': spa_ward,
                            'shealth_all_in_one.group_sh_medical_physician_laser': laser_ward}

            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp) and loc:
                    if grp == 'shealth_all_in_one.group_sh_medical_receptionist':
                        access_location.append(loc.id)
                    else:
                        access_location.append(loc.location_id.id)

            # quyen bác sĩ chung
            # if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician'):
            #     access_location.append(institution.location_medicine_stock.id)
            #     access_location.append(institution.location_emergency_stock.id)

            #quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        domain = [('location_id', 'child_of', access_location)]
        if self.env.context.get('single_product'):
            domain += [('product_id', '=', self.env.context.get('single_product'))]

        return {
            'name': 'Tồn kho tại tủ',
            'type': 'ir.actions.act_window',
            'view_mode': 'tree',
            'domain': domain,
            'context': {'search_default_locationgroup': True, 'search_default_productgroup': True,'view_only_name':True},
            'res_model': 'stock.quant',
            'view_id': self.env.ref('shealth_all_in_one.sci_current_stock_quant_tree').id,
            'target': 'current'
        }

    def view_stock_picking_by_group(self, type):
        access_location = []

        # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
        institution = self.env['sh.medical.health.center'].sudo().search([('his_company', '=', self.env.companies.ids[0])], limit=1)

        if not institution:
            raise ValidationError(_('Công ty hiện tại của bạn không có Cơ sở y tế! Hãy chọn chính xác công ty!'))

        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            current_warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.companies.ids[0])], limit=1)
            location_stock0 = current_warehouse.lot_stock_id.id  # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            receptionist_ward = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('name', 'ilike', 'lễ tân')], limit=1)
            lab_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Laboratory')], limit=1)
            spa_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Spa')], limit=1)
            laser_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Laser')], limit=1)
            surgery_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Surgery')], limit=1)
            odontology_ward = self.env['sh.medical.health.center.ward'].sudo().search([('institution', '=', institution.id),('type', '=', 'Odontology')], limit=1)

            grp_loc_dict = {'shealth_all_in_one.group_sh_medical_receptionist': receptionist_ward,
                            'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': lab_ward,
                            'shealth_all_in_one.group_sh_medical_physician_surgery': surgery_ward,
                            'shealth_all_in_one.group_sh_medical_physician_odontology': odontology_ward,
                            'shealth_all_in_one.group_sh_medical_physician_spa': spa_ward,
                            'shealth_all_in_one.group_sh_medical_physician_laser': laser_ward}

            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp) and loc:
                    if grp == 'shealth_all_in_one.group_sh_medical_receptionist':
                        access_location.append(loc.id)
                    else:
                        access_location.append(loc.location_id.id)

            # quyen bác sĩ chung
            # if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician'):
            #     access_location.append(institution.location_medicine_stock.id)
            #     access_location.append(institution.location_emergency_stock.id)

            #quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        domain = []
        context = {}
        name = 'Phiếu kho'
        res_model = 'stock.picking'

        view_ids = [(5, 0, 0),
                        (0, 0, {'view_mode': 'tree', 'view_id': self.env.ref('stock.vpicktree')}),
                        (0, 0, {'view_mode': 'form', 'view_id': self.env.ref('stock.view_picking_form')})]

        #phiếu nhận thuốc
        if type == 'picking_int':
            name = 'Phiếu chuyển thuốc'
            domain = [('location_dest_id', 'child_of', access_location),('name', 'ilike', '/INT/'),
                      ('location_dest_id.location_institution_type', '=', 'medicine'), ('returned_ids', '=', False),
                      ('backorder_id', '=', False)]
            context = {'contact_display': 'partner_address', 'view_for': 'picking_int', 'view_only_name': True,
                       'search_default_group_location_dest_id': True}
        # phiếu nhận vật tư
        elif type == 'picking_int_vt':
            name = 'Phiếu chuyển vật tư'
            domain = [('location_dest_id', 'child_of', access_location), ('name', 'ilike', '/INT/'),
                      ('location_dest_id.location_institution_type', '=', 'supply'), ('returned_ids', '=', False),
                      ('backorder_id', '=', False)]
            context = {'contact_display': 'partner_address', 'view_for': 'picking_int_vt',
                       'view_only_name': True, 'search_default_group_location_dest_id': True}
        # phiếu xuất bệnh nhân
        elif type == 'picking_out_bn':
            name = 'Phiếu xuất bệnh nhân'
            domain = [('location_id', 'child_of', access_location), ('patient_id','!=',False),
                      ('name','ilike','/OUT/'),'!',('name','ilike','-FP')]
            context = {'contact_display': 'partner_address', 'view_for': 'picking_out_bn',
                       'view_only_name': True, 'search_default_group_location_id': True}
        # phiếu xuất bán
        elif type == 'picking_out_sale':
            name = 'Phiếu xuất bán'
            domain = [('location_id', 'child_of', access_location),('sale_id', '!=', False), ('patient_id', '=', False),
                      ('name', 'ilike', '/OUT/'), '!', ('name', 'ilike', '-FP')]
            context = {'contact_display': 'partner_address', 'view_for': 'picking_out_sale',
                   'view_only_name': True, 'search_default_group_location_id': True}
        # phiếu xuất chuyển chi nhánh
        elif type == 'picking_out_company':
            company_partners = self.env['res.company'].sudo().search([]).mapped('partner_id')
            name = 'Phiếu xuất chuyển chi nhánh'
            domain = [('location_id', 'child_of', access_location),
                      ('location_dest_id.usage', '=', 'customer'), ('partner_id', 'in', company_partners.ids),
                      ('picking_type_id.code', '=', 'outgoing'), '!', ('name', 'ilike', '-FP')]
            context = {'contact_display': 'partner_address', 'view_for': 'picking_out_company',
                       'view_only_name': True, 'search_default_group_location_id': True}
        # phiếu nhận hàng từ chi nhánh
        elif type == 'picking_in_company':
            company_partners = self.env['res.company'].sudo().search([]).mapped('partner_id')
            name = 'Phiếu nhận hàng từ chi nhánh'
            domain = [('location_dest_id', 'child_of', access_location),
                      ('location_id.usage', '=', 'supplier'), ('partner_id', 'in', company_partners.ids),
                      ('picking_type_id.code', '=', 'incoming'), '!', ('name', 'ilike', '-FP')]
            context = {'contact_display': 'partner_address', 'view_for': 'picking_in_company',
                   'view_only_name': True, 'search_default_group_location_id': True}
        # phiếu hoàn thuốc, vật tư
        elif type == 'picking_int_return':
            name = 'Phiếu hoàn thuốc/vật tư'

            root_ins_loc = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('usage', '=', 'internal'),
                 ('location_id', '=', institution.warehouse_ids[0].lot_stock_id.id), ('child_ids', '=', False),
                 ('location_institution_type', 'in', ['medicine', 'supply'])])

            domain = [('location_id', 'child_of', access_location), ('patient_id', '=', False),
                      ('location_dest_id', 'in', root_ins_loc.ids),
                      ('name', 'ilike', '/INT/'), '!', ('name', 'ilike', '-FP')]

            out_location = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('usage', '=', 'internal'),
                 ('location_id', 'in', access_location), ('child_ids', '=', False),
                 ('location_institution_type', 'in', ['medicine', 'supply'])], limit=1)

            context = {'contact_display': 'partner_address', 'view_for': 'picking_int_return',
                       'view_only_name': True, 'search_default_group_location_id': True,
                       'default_location_id': out_location.id}
        # phiếu xuất hủy
        elif type == 'picking_scrap':
            name = 'Tiêu hủy Thuốc/Vật tư'
            res_model = 'stock.scrap'

            out_location = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('usage', '=', 'inventory'),
                 ('location_id', 'in', access_location), ('child_ids', '=', False),
                 ('location_institution_type', 'in', ['medicine', 'supply'])], limit=1)

            domain = [('location_id', 'child_of', access_location), ('scrap_location_id.name', 'ilike', 'Scrap')]
            context = {'view_for': 'picking_scrap', 'view_only_name': True, 'search_default_group_location_id': True,
                       'default_location_id': out_location.id}
            view_ids = [(5, 0, 0),
                        (0, 0, {'view_mode': 'tree', 'view_id': self.env.ref('shealth_all_in_one.sci_stock_scrap_tree')}),
                        (0, 0, {'view_mode': 'form', 'view_id': self.env.ref('shealth_all_in_one.sci_stock_scrap_form')})]

        # phiếu xuất sử dụng phòng
        elif type == 'picking_scrap_room_use':
            name = 'Xuất sử dụng phòng'
            res_model = 'stock.scrap'

            out_location = self.env['stock.location'].sudo().search(
                [('company_id', '=', institution.his_company.id), ('usage', '=', 'internal'),
                 ('location_id', 'in', access_location), ('child_ids', '=', False),
                 ('location_institution_type', 'in', ['medicine', 'supply'])], limit=1)

            domain = [('location_id', 'child_of', access_location), ('scrap_location_id.name', 'ilike', 'Sử dụng phòng')]
            context = {'view_for': 'picking_scrap_room_use', 'view_only_name': True, 'search_default_group_location_id': True,
                       'default_location_id': out_location.id, 'type_stock_scrap': 'room_use'}
            view_ids = [(5, 0, 0),
                        (0, 0, {'view_mode': 'tree', 'view_id': self.env.ref('shealth_all_in_one.sci_stock_scrap_room_use_tree')}),
                        (0, 0, {'view_mode': 'form', 'view_id': self.env.ref('shealth_all_in_one.sci_stock_scrap_room_use_form')})]
        return {
            'name': name,
            'type': 'ir.actions.act_window',
            'view_mode': 'tree,form',
            'domain': domain,
            'context': context,
            'res_model': res_model,
            'view_ids': view_ids,
            'target': 'current'
        }


class SHealthProductionLot(models.Model):
    _inherit = 'stock.production.lot'

    product_init_qty = fields.Float('Initial quantity')
    lot_price = fields.Float('Lot price')
    product_qty = fields.Float('Số lượng thực tế')
    reserved_qty = fields.Float('Số lượng giữ trước', compute="_compute_quant_ids")

    @api.depends('quant_ids')
    def _compute_quant_ids(self):
        for lot in self:
            lot.reserved_qty = sum(lot.quant_ids.mapped('reserved_quantity')) if lot.quant_ids else 0


class SHealthStockMove(models.Model):
    _inherit = 'stock.move'

    quant_in_loc = fields.Float(string='Tồn tại tủ nhận', compute='_compute_quant_in_loc',default=0)

    @api.depends('product_id','picking_id.location_dest_id')
    def _compute_quant_in_loc(self):
        for record in self:
            record.quant_in_loc = self.env['stock.quant']._get_available_quantity(product_id=record.product_id,location_id=record.picking_id.location_dest_id) \
                if record.product_id and record.picking_id.location_dest_id else 0

    def _search_picking_for_assignation(self):
        """Tách phiếu mới mỗi ngày khi chạy cơ số tủ trực thay vì để move nhảy vào phiếu cũ"""
        if self.env.context.get('separate_pick'):
            self.ensure_one()
            user = self.env.user
            tz = pytz.timezone(user.tz or 'Asia/Ho_Chi_Minh')
            now = pytz.utc.localize(datetime.now()).astimezone(tz)
            picking = self.env['stock.picking'].search([
                ('origin', '=', 'Cơ số tủ trực - %s' % now.date().strftime("%d/%m/%Y")),
                ('group_id', '=', self.group_id.id),
                ('location_id', '=', self.location_id.id),
                ('location_dest_id', '=', self.location_dest_id.id),
                ('picking_type_id', '=', self.picking_type_id.id),
                ('printed', '=', False),
                ('state', 'in', ['draft', 'confirmed', 'waiting', 'partially_available', 'assigned'])], limit=1)
            return picking
        return super(SHealthStockMove, self)._search_picking_for_assignation()

    def _get_new_picking_values(self):
        """Tách phiếu mới mỗi ngày khi chạy cơ số tủ trực thay vì để move nhảy vào phiếu cũ"""
        res = super(SHealthStockMove, self)._get_new_picking_values()
        if self.env.context.get('separate_pick'):
            user = self.env.user
            tz = pytz.timezone(user.tz or 'Asia/Ho_Chi_Minh')
            now = pytz.utc.localize(datetime.now()).astimezone(tz)
            res['origin'] = 'Cơ số tủ trực - %s' % now.date().strftime("%d/%m/%Y")
        return res

    def _get_domain_product_id(self):
        if self.env.context.get('view_for', True) in ["picking_int", "picking_int_vt"]:
            if self.env.context.get('view_for', True) == 'picking_int':
                return [('categ_id', '=', self.env.ref('shealth_all_in_one.sh_medicines').id)]
            else:
                return [('categ_id', '=', self.env.ref('shealth_all_in_one.sh_supplies').id)]
        else:
            return [('type', 'in', ['product', 'consu'])]

    product_id = fields.Many2one(domain=lambda self: self._get_domain_product_id())

    @api.onchange('product_uom_qty', 'product_id')
    def onchange_product_uom_qty(self):
        if self.product_id:
            if self.product_uom_qty <= 0:
                self.product_uom_qty = 1

            if self.product_uom_qty <= 0:
                raise UserError(_("Số lượng nhập phải lớn hơn 0!"))

    @api.model
    def create(self, vals):
        if self.env.context.get('force_location'):
            vals['location_id'] = self.env.context.get('force_location')
        return super(SHealthStockMove, self).create(vals)

    def write(self, vals):
        """Dùng cho trường hợp tạo phiếu từ cơ số, phiếu chỉ dừng ở dạng nháp để sửa được sl yc"""
        if self.env.context.get('do_not_confirm') and vals.get('state') == 'confirmed':
            vals.pop('state')
        return super(SHealthStockMove, self).write(vals)

    @api.model
    def read_group(self, domain, fields, groupby, offset=0, limit=None, orderby=False, lazy=True):
        """Đọc cả các move dạng nháp (trong hàm _product_available) để tránh tạo move nháp duplicate vì cơ số đang để move dạng nháp"""
        if self.env.context.get('do_not_confirm') and domain[0] == ('state', 'in', ('waiting', 'confirmed', 'assigned', 'partially_available')):
            domain[0] = ('state', 'in', ('draft', 'waiting', 'confirmed', 'assigned', 'partially_available'))
        return super(SHealthStockMove, self).read_group(domain, fields, groupby, offset=offset, limit=limit, orderby=orderby, lazy=lazy)



class SHealthStockMoveLine(models.Model):
    _inherit = 'stock.move.line'

    quant_in_loc_out = fields.Float(string='Tồn tại tủ xuất', compute='_compute_quant_in_loc_out', default=0)

    @api.depends('product_id', 'picking_id.location_id')
    def _compute_quant_in_loc_out(self):
        for record in self:
            record.quant_in_loc_out = self.env['stock.quant']._get_available_quantity(product_id=record.product_id,
                                                                                      location_id=record.picking_id.location_id) \
                if record.product_id and record.picking_id.location_id and record.picking_id.location_id.usage == 'internal' else 0

    def _action_done(self):
        """ This method is called during a move's `action_done`. It'll actually move a quant from
        the source location to the destination location, and unreserve if needed in the source
        location.

        This method is intended to be called on all the move lines of a move. This method is not
        intended to be called when editing a `done` move (that's what the override of `write` here
        is done.
        """
        Quant = self.env['stock.quant']

        # First, we loop over all the move lines to do a preliminary check: `qty_done` should not
        # be negative and, according to the presence of a picking type or a linked inventory
        # adjustment, enforce some rules on the `lot_id` field. If `qty_done` is null, we unlink
        # the line. It is mandatory in order to free the reservation and correctly apply
        # `action_done` on the next move lines.
        ml_to_delete = self.env['stock.move.line']
        for ml in self:
            # Check here if `ml.qty_done` respects the rounding of `ml.product_uom_id`.
            uom_qty = float_round(ml.qty_done, precision_rounding=ml.product_uom_id.rounding, rounding_method='HALF-UP')
            precision_digits = self.env['decimal.precision'].precision_get('Product Unit of Measure')
            qty_done = float_round(ml.qty_done, precision_digits=precision_digits, rounding_method='HALF-UP')
            if float_compare(uom_qty, qty_done, precision_digits=precision_digits) != 0:
                raise UserError(_('The quantity done for the product "%s" doesn\'t respect the rounding precision \
                                  defined on the unit of measure "%s". Please change the quantity done or the \
                                  rounding precision of your unit of measure.') % (ml.product_id.display_name, ml.product_uom_id.name))

            qty_done_float_compared = float_compare(ml.qty_done, 0, precision_rounding=ml.product_uom_id.rounding)
            if qty_done_float_compared > 0:
                if ml.product_id.tracking != 'none':
                    picking_type_id = ml.move_id.picking_type_id
                    if picking_type_id:
                        if picking_type_id.use_create_lots:
                            # If a picking type is linked, we may have to create a production lot on
                            # the fly before assigning it to the move line if the user checked both
                            # `use_create_lots` and `use_existing_lots`.
                            if ml.lot_name and not ml.lot_id:
                                lot_price = 0
                                if ml.picking_id.purchase_id:
                                    lot_price = ml.picking_id.purchase_id.order_line.filtered(lambda o: o.product_id == ml.product_id)[0].price_subtotal
                                lot = self.env['stock.production.lot'].create(
                                    {'name': ml.lot_name, 'product_id': ml.product_id.id,
                                     'product_init_qty': ml.qty_done, 'lot_price': lot_price}
                                )
                                ml.write({'lot_id': lot.id})
                        elif not picking_type_id.use_create_lots and not picking_type_id.use_existing_lots:
                            # If the user disabled both `use_create_lots` and `use_existing_lots`
                            # checkboxes on the picking type, he's allowed to enter tracked
                            # products without a `lot_id`.
                            continue
                    elif ml.move_id.inventory_id:
                        # If an inventory adjustment is linked, the user is allowed to enter
                        # tracked products without a `lot_id`.
                        continue

                    if not ml.lot_id:
                        raise UserError(_('You need to supply a Lot/Serial number for product %s.') % ml.product_id.display_name)
            elif qty_done_float_compared < 0:
                raise UserError(_('No negative quantities allowed'))
            else:
                ml_to_delete |= ml
        ml_to_delete.unlink()

        # Now, we can actually move the quant.
        done_ml = self.env['stock.move.line']
        for ml in self - ml_to_delete:
            if ml.product_id.type == 'product':
                rounding = ml.product_uom_id.rounding

                # if this move line is force assigned, unreserve elsewhere if needed
                if not ml.location_id.should_bypass_reservation() and float_compare(ml.qty_done, ml.product_qty, precision_rounding=rounding) > 0:
                    extra_qty = ml.qty_done - ml.product_qty
                    ml._free_reservation(ml.product_id, ml.location_id, extra_qty, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id, ml_to_ignore=done_ml)
                # unreserve what's been reserved
                if not ml.location_id.should_bypass_reservation() and ml.product_id.type == 'product' and ml.product_qty:
                    try:
                        Quant._update_reserved_quantity(ml.product_id, ml.location_id, -ml.product_qty, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id, strict=True)
                    except UserError:
                        Quant._update_reserved_quantity(ml.product_id, ml.location_id, -ml.product_qty, lot_id=False, package_id=ml.package_id, owner_id=ml.owner_id, strict=True)

                # move what's been actually done
                quantity = ml.product_uom_id._compute_quantity(ml.qty_done, ml.move_id.product_id.uom_id, rounding_method='HALF-UP')
                available_qty, in_date = Quant._update_available_quantity(ml.product_id, ml.location_id, -quantity, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id)
                if available_qty < 0 and ml.lot_id:
                    # see if we can compensate the negative quants with some untracked quants
                    untracked_qty = Quant._get_available_quantity(ml.product_id, ml.location_id, lot_id=False, package_id=ml.package_id, owner_id=ml.owner_id, strict=True)
                    if untracked_qty:
                        taken_from_untracked_qty = min(untracked_qty, abs(quantity))
                        Quant._update_available_quantity(ml.product_id, ml.location_id, -taken_from_untracked_qty, lot_id=False, package_id=ml.package_id, owner_id=ml.owner_id)
                        Quant._update_available_quantity(ml.product_id, ml.location_id, taken_from_untracked_qty, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id)
                Quant._update_available_quantity(ml.product_id, ml.location_dest_id, quantity, lot_id=ml.lot_id, package_id=ml.result_package_id, owner_id=ml.owner_id, in_date=in_date)
            done_ml |= ml
        # Reset the reserved quantity as we just moved it to the destination location.
        (self - ml_to_delete).with_context(bypass_reservation_update=True).write({
            'product_uom_qty': 0.00,
            'date': fields.Datetime.now(),
        })

#BIẾN CỤC BỘ STYLE EXCEL
thin = borders.Side(style='thin')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)

# KIỂM KÊ
class SHealthInventory(models.Model):
    _inherit = "stock.inventory"

    # user_pharmacist = fields.Many2one('sh.medical.physician', string='Dược sĩ', help="Dược sĩ", domain=[('is_pharmacist','=',True)], required=False, states={'done': [('readonly', True)]})
    # user_physician = fields.Many2one('sh.medical.physician', string='Người quản lý', help="Người quản lý", domain=[('is_pharmacist','=',False)], required=False, states={'done': [('readonly', True)]})
    user_pharmacist = fields.Many2one('res.users', string='Dược sĩ', help="Dược sĩ",
                                      domain=lambda self: [('groups_id', 'in', [self.env.ref('shealth_all_in_one.group_sh_medical_stock_manager').id])])
    user_physician = fields.Many2one('res.users', string='Người quản lý', help="Người quản lý")
    user_accountant = fields.Many2one('res.users', string='Kế toán kho', help="Kế toán kho",
                                      domain=lambda self: [('groups_id','in',[self.env.ref('account.group_account_manager').id])])

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(SHealthInventory, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,
                                                               submenu=submenu)
        doc = etree.XML(res['arch'])

        for node in doc.xpath("//field[@name='category_id']"):
            node_domain = "[('id', 'in', [%d,%d])]" % (self.env.ref('shealth_all_in_one.sh_medicines').id, self.env.ref('shealth_all_in_one.sh_supplies').id)

            node.set('domain', node_domain)
            modifiers = json.loads(node.get("modifiers"))
            modifiers['domain'] = node_domain
            node.set("modifiers", json.dumps(modifiers))

        res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    def action_start(self):
        for inventory in self.filtered(lambda x: x.state not in ('done','cancel')):
            vals = {'state': 'confirm', 'date': inventory.date}
            if not inventory.line_ids:
            # if (inventory.filter != 'partial') and not inventory.line_ids:
                vals.update({'line_ids': [(0, 0, line_values) for line_values in inventory._get_inventory_lines_values()]})
            inventory.write(vals)
        return True

    def _action_done(self):
        res = super(SHealthInventory, self)._action_done()

        for stock_inventory in self:
            if stock_inventory.date:
                for move_line in stock_inventory.move_ids:
                    move_line.move_line_ids.write(
                        {'date': stock_inventory.date})  # sửa ngày hoàn thành ở stock move line
                stock_inventory.move_ids.write(
                    {'date': stock_inventory.date})  # sửa ngày hoàn thành ở stock move

        return res

    # IN PHIẾU KIỂM KÊ KHO
    #
    def print_stock_inventory(self):
        wb = Workbook()
        ws = wb.active
        col_widths = [('a', 1.57), ('b', 5), ('c', 14.43), ('d', 30.71), ('e', 6.71), ('f', 15.86), ('g', 9.86),
                      ('h', 10.14),('i', 10.14), ('j', 30.86)]
        for value in col_widths:
            ws.column_dimensions[value[0]].width = value[1]
        row = 1
        for inventory in self:
            # inventory_attachment = self.env['ir.attachment'].browse(
            #     self.env.ref('shealth_all_in_one.sci_stock_inventory_report_attachment').id)
            # decode = base64.b64decode(inventory_attachment.datas)

            main_font = Font(name='Times New Roman', size=12)
            font_11 = Font(name='Times New Roman', size=11)
            font_14 = Font(name='Times New Roman', size=14)
            font_14_bold = Font(name='Times New Roman', size=14, bold=True)
            font_12_bold = Font(name='Times New Roman', size=12, bold=True)
            alignment_center = Alignment(horizontal='center', vertical='center')

            health_center = self.env['sh.medical.health.center'].sudo().search([('his_company', '=',
                                                                                 inventory.company_id.id)], limit=1)

            physicians = self.env['sh.medical.physician'].sudo().search([
                ('institution', '=', health_center.id), ('is_pharmacist', '=', True)])

            ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = \
                inventory.company_id.name.upper() or 'CTY CP BVTM KANGNAM HÀ NỘI', font_11
            ws.cell(row=row+1, column=2).value, ws.cell(row=row+1, column=2).font = \
                inventory.company_id.street or 'Số 190 Trường Chinh, Đống Đa, Hà Nội', font_11
            ws.cell(row=row, column=10).value, ws.cell(row=row, column=10).font = 'MS: 11D/BV-01', font_11
            ws.cell(row=row + 1, column=10).value, ws.cell(row=row + 1, column=10).font = 'Số', font_11

            ws.merge_cells(start_row=row+2, start_column=2, end_row=row+2, end_column=10)
            ws.cell(row=row + 2, column=2).value, ws.cell(row=row + 2, column=2).font, ws.cell(row=row + 2, column=2).alignment = 'BIÊN BẢN KIỂM KÊ %s' % (
            inventory.date.strftime('THÁNG %m NĂM %Y')), Font(name='Times New Roman', size=18, bold=True),alignment_center

            ws.cell(row=row + 4, column=3).value, ws.cell(row=row + 4, column=3).font = 'Tổ kiểm kê gồm có', font_14_bold

            ws.cell(row=row + 5, column=2).value, ws.cell(row=row + 5, column=2).font = 1, font_14
            ws.cell(row=row + 5, column=3).value, ws.cell(row=row + 5, column=3).font = \
                self.user_pharmacist.name, font_14
            ws.merge_cells(start_row=row + 5, start_column=5, end_row=row + 5, end_column=6)
            ws.cell(row=row + 5, column=5).value, ws.cell(row=row + 5, column=5).font = 'Chức danh:', font_14
            ws.cell(row=row + 5, column=7).value, ws.cell(row=row + 5, column=7).font = \
                self.user_pharmacist.employee_ids.job_id.name or '', font_14

            ws.cell(row=row + 6, column=2).value, ws.cell(row=row + 6, column=2).font = 2, font_14
            ws.cell(row=row + 6, column=3).value, ws.cell(row=row + 6, column=3).font = self.user_physician.name, font_14
            ws.merge_cells(start_row=row + 6, start_column=5, end_row=row + 6, end_column=6)
            ws.cell(row=row + 6, column=5).value, ws.cell(row=row + 6, column=5).font = 'Chức danh:', font_14
            ws.cell(row=row + 6, column=7).value, ws.cell(row=row + 6, column=7).font = self.user_physician.employee_ids.job_id.name or '', font_14

            ws.cell(row=row + 7, column=2).value, ws.cell(row=row + 7, column=2).font = 2, font_14
            ws.cell(row=row + 7, column=3).value, ws.cell(row=row + 7,
                                                          column=3).font = self.user_accountant.name, font_14
            ws.merge_cells(start_row=row + 7, start_column=5, end_row=row + 7, end_column=6)
            ws.cell(row=row + 7, column=5).value, ws.cell(row=row + 7, column=5).font = 'Chức danh:', font_14
            ws.cell(row=row + 7, column=7).value, ws.cell(row=row + 7,
                                                          column=7).font = self.user_accountant.employee_id.job_id.name or '', font_14

            ws.merge_cells(start_row=row + 9, start_column=3, end_row=row + 9, end_column=10)
            ws.cell(row=row + 9, column=3).value, ws.cell(row=row + 9, column=3).font = 'Đã kiểm kê tại %s từ %s' % \
                             (health_center[0].name,
                              inventory.date.strftime('%H giờ %M ngày %d tháng %m năm %Y')), font_14

            ws.cell(row=row + 11, column=3).value, ws.cell(row=row + 11,
                                                          column=3).font = 'Kết quả như sau:', font_14_bold

            ws.merge_cells(start_row=row + 13, start_column=2, end_row=row + 14, end_column=2)
            ws.cell(row=row+13, column=2).value, ws.cell(row=row+13, column=2).font = 'STT', font_12_bold
            ws.cell(row=row+13, column=2).border, ws.cell(row=row+13, column=2).alignment = all_border_thin, alignment_center
            ws.cell(row=row+14, column=2).border = all_border_thin

            ws.merge_cells(start_row=row + 13, start_column=3, end_row=row + 14, end_column=3)
            ws.cell(row=row + 13, column=3).value, ws.cell(row=row + 13, column=3).font = 'Mã', font_12_bold
            ws.cell(row=row + 13, column=3).border, ws.cell(row=row + 13,
                                                            column=3).alignment = all_border_thin, alignment_center

            ws.merge_cells(start_row=row + 13, start_column=4, end_row=row + 14, end_column=4)
            ws.cell(row=row + 13, column=4).value, ws.cell(row=row + 13,
                                                           column=4).font = 'Tên thuốc, nồng độ, hàm lượng', font_12_bold
            ws.cell(row=row + 13, column=4).border, ws.cell(row=row + 13,
                                                            column=4).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 14, column=4).border = all_border_thin

            ws.merge_cells(start_row=row + 13, start_column=5, end_row=row + 14, end_column=5)
            ws.cell(row=row + 13, column=5).value, ws.cell(row=row + 13,
                                                           column=5).font = 'Đơn vị', font_12_bold
            ws.cell(row=row + 13, column=5).border, ws.cell(row=row + 13,
                                                            column=5).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 14, column=5).border = all_border_thin

            ws.merge_cells(start_row=row + 13, start_column=6, end_row=row + 14, end_column=6)
            ws.cell(row=row + 13, column=6).value, ws.cell(row=row + 13,
                                                           column=6).font = 'Hãng, nước sản xuất', font_12_bold
            ws.cell(row=row + 13, column=6).border, ws.cell(row=row + 13,
                                                            column=6).alignment = all_border_thin, Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row + 14, column=6).border = all_border_thin

            # ws.merge_cells(start_row=row + 13, start_column=7, end_row=row + 14, end_column=7)
            # ws.cell(row=row + 13, column=7).value, ws.cell(row=row + 13,
            #                                                column=7).font = 'Hạn dùng', font_12_bold
            # ws.cell(row=row + 13, column=7).border, ws.cell(row=row + 13,
            #                                                 column=7).alignment = all_border_thin, alignment_center

            ws.merge_cells(start_row=row + 13, start_column=7, end_row=row + 13, end_column=9)
            ws.cell(row=row + 13, column=7).value, ws.cell(row=row + 13,
                                                           column=7).font = 'Số lượng', font_12_bold
            ws.cell(row=row + 13, column=7).border, ws.cell(row=row + 13,
                                                            column=7).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 13, column=8).border = all_border_thin
            ws.cell(row=row + 13, column=9).border = all_border_thin

            ws.cell(row=row + 14, column=7).value, ws.cell(row=row + 14,
                                                           column=7).font = 'Sổ sách', font_12_bold
            ws.cell(row=row + 14, column=7).border, ws.cell(row=row + 14,
                                                            column=7).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 14, column=8).value, ws.cell(row=row + 14,
                                                           column=8).font = 'Thực tế', font_12_bold
            ws.cell(row=row + 14, column=8).border, ws.cell(row=row + 14,
                                                            column=8).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 14, column=9).value, ws.cell(row=row + 14,
                                                           column=9).font = 'Chênh lệch', font_12_bold
            ws.cell(row=row + 14, column=9).border, ws.cell(row=row + 14,
                                                            column=9).alignment = all_border_thin, alignment_center

            # ws.merge_cells(start_row=row + 13, start_column=10, end_row=row + 14, end_column=10)
            # ws.cell(row=row + 13, column=10).value, ws.cell(row=row + 13,
            #                                                column=10).font = 'Hỏng vỡ', font_12_bold
            # ws.cell(row=row + 13, column=10).border, ws.cell(row=row + 13,
            #                                                 column=10).alignment = all_border_thin, alignment_center
            # ws.cell(row=row + 13, column=10).border = all_border_thin

            ws.merge_cells(start_row=row + 13, start_column=10, end_row=row + 14, end_column=10)
            ws.cell(row=row + 13, column=10).value, ws.cell(row=row + 13,
                                                           column=10).font = 'Ghi chú', font_12_bold
            ws.cell(row=row + 13, column=10).border, ws.cell(row=row + 13,
                                                            column=10).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 14, column=10).border = all_border_thin

            row += 15
            for line in inventory.line_ids:
                ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = row - 15, main_font
                ws.cell(row=row, column=2).border, ws.cell(row=row, column=2).alignment = all_border_thin, alignment_center

                ws.cell(row=row, column=3).value, ws.cell(row=row, column=3).font = line.product_id.default_code, main_font
                ws.cell(row=row, column=3).border, ws.cell(row=row, column=3).alignment = all_border_thin, Alignment(
                    horizontal='left', vertical='center', wrap_text=True)

                ws.cell(row=row, column=4).value, ws.cell(row=row, column=4).font = line.product_id.name, main_font
                ws.cell(row=row, column=4).border, ws.cell(row=row, column=4).alignment = all_border_thin, Alignment(
                    horizontal='left', vertical='center', wrap_text=True)

                max_length_product = len(line.product_id.name) or 0

                ws.cell(row=row, column=5).value, ws.cell(row=row, column=5).font = line.product_uom_id.name, main_font
                ws.cell(row=row, column=5).border, ws.cell(row=row, column=5).alignment = all_border_thin, alignment_center

                # ws.cell(row=row, column=5).value, ws.cell(row=row, column=5).font = line.prod_lot_id.name, main_font
                # ws.cell(row=row, column=5).border, ws.cell(row=row, column=5).alignment = all_border_thin, alignment_center

                max_length_lot = len(line.prod_lot_id) or 0

                medicine_details = self.env['sh.medical.medicines'].sudo().search([('default_code', '=',
                                                                                    line.product_id.default_code)], limit=1)
                ws.cell(row=row, column=6).value, ws.cell(row=row, column=6).font =\
                    '', main_font
                # ws.cell(row=row, column=6).value, ws.cell(row=row, column=6).font = \
                #     medicine_details[0].origin.name or '', main_font
                ws.cell(row=row, column=6).border, ws.cell(row=row, column=6).alignment = \
                    all_border_thin, alignment_center
                #
                # ws.cell(row=row, column=6).value, ws.cell(row=row, column=6).font = \
                #     line.prod_lot_id.removal_date.strftime('%d/%m/%Y') if line.prod_lot_id.removal_date else '', main_font
                # ws.cell(row=row, column=6).border, ws.cell(row=row, column=6).alignment = all_border_thin, alignment_center

                ws.cell(row=row, column=7).value, ws.cell(row=row, column=7).font = line.theoretical_qty, main_font
                ws.cell(row=row, column=7).border, ws.cell(row=row, column=7).alignment = all_border_thin, alignment_center
                ws.cell(row=row, column=8).value, ws.cell(row=row, column=8).font = line.product_qty, main_font
                ws.cell(row=row, column=8).border, ws.cell(row=row, column=8).alignment = all_border_thin, alignment_center
                ws.cell(row=row, column=9).value, ws.cell(row=row, column=9).font = line.difference_qty, main_font
                ws.cell(row=row, column=9).border, ws.cell(row=row, column=9).alignment = all_border_thin, alignment_center
                ws.cell(row=row, column=10).value, ws.cell(row=row, column=10).font = line.note or '', main_font
                ws.cell(row=row, column=10).border, ws.cell(row=row, column=10).alignment = all_border_thin, alignment_center
                # ws.cell(row=row, column=10).border = all_border_thin
                # ws.cell(row=row, column=9).border = all_border_thin
                row += 1
                max_length = max_length_product if max_length_product > max_length_lot else max_length_lot
                adjusted_height = (max_length + 2) * 1.2
                if adjusted_height > 0:
                    ws.row_dimensions[row].height = adjusted_height

            # Đề xuất
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = 'Ý kiến đề xuất:', font_14_bold
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
            ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=11)
            ws.cell(row=row+1, column=2).alignment, ws.cell(row=row+1, column=2).font = Alignment(
                horizontal='left', wrap_text=True), main_font

            #Ký
            ws.merge_cells(start_row=row+3, start_column=2, end_row=row+3, end_column=4)
            ws.merge_cells(start_row=row+3, start_column=8, end_row=row+3, end_column=9)
            ws.cell(row=row+3, column=2).value, ws.cell(row=row+3, column=2).font = 'THÀNH VIÊN HĐ KIỂM KÊ', font_14_bold
            ws.cell(row=row + 3, column=6).value, ws.cell(row=row + 3, column=6).font = 'THƯ KÝ', font_14_bold
            ws.cell(row=row + 3, column=8).value, ws.cell(row=row + 3, column=8).font = 'CHỦ TỊCH HĐ KIỂM KÊ', font_14_bold
            ws.cell(row=row+3, column=2).alignment = alignment_center
            ws.cell(row=row+3, column=6).alignment = alignment_center
            ws.cell(row=row+3, column=8).alignment = alignment_center

            row += 7
            ws.row_breaks.append(Break(id=row))
            row += 1

        row -= 1
        # ws.sheet_view.showGridLines = False
        ws.print_area = 'A1:K%s' % str(row)
        ws.set_printer_settings(paper_size=9, orientation='landscape')
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.3, footer=0.3)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'BB_KIEM_KE.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true&filename=BB_KIEM_KE.xlsx" \
              % (attachment.id)
        cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        return {'name': 'BIÊN BẢN KIỂM KÊ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
        }

        return {'name': 'BIÊN BẢN KIỂM KÊ',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
        }

#TIÊU HỦY THUỐC HẾT HẠN
class SHealthStockScrap(models.Model):
    _inherit = 'stock.scrap'

    sci_date_done = fields.Datetime('Ngày hoàn thành', copy=False, help="Ngày ghi nhận hoàn thành SCI", states={'done': [('readonly', True)]}, default=lambda self: fields.Datetime.now())
    note = fields.Text('Lý do chi tiết', copy=False, help="Ghi nhận lý do", states={'done': [('readonly', True)]})

    room_use = fields.Many2one('sh.medical.health.center.ot',copy=False, string='Phòng sử dụng', help="Xuất cho phòng nào sử dụng", states={'done': [('readonly', True)]})
    date_expected = fields.Datetime(states={'done': [('readonly', True)]})
    origin = fields.Char(states={'done': [('readonly', True)]})
    company_id = fields.Many2one(domain="[('id', 'in', allowed_company_ids)]")

    @api.onchange('company_id')
    def _onchange_company_id(self):
        if self.company_id:
            warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.company_id.id)], limit=1)
            # self.location_id = False

            self.scrap_location_id = self.env['stock.location'].search([('name','ilike','Sử dụng phòng'),('company_id','=', self.company_id.id)]) \
                if self.env.context.get('type_stock_scrap') == 'room_use' else \
                self.env['stock.location'].search([('scrap_location', '=', True), ('company_id', '=', self.company_id.id)], limit=1)
        else:
            self.location_id = False
            self.scrap_location_id = False

    @api.onchange('product_id','location_id')
    def onchange_product_id(self):
        if self.product_id:
            self.product_uom_id = self.product_id.uom_id

            #báo lỗi khi chưa chọn tủ xuất
            if not self.location_id:
                raise ValidationError(_('Bạn phải chọn tủ xuất!'))

            #lấy tất cả các lô của sp ở địa điểm tủ xuất đã chọn và số lượng tồn > 0
            data_lot_inlocation = self.env['stock.quant'].search([('product_id', '=', self.product_id.id), ('location_id', '=', self.location_id.id), ('quantity', '>', 0)])

            #nếu có data lô thì set trường lô hàng là lô sắp hết hạn
            if data_lot_inlocation:
                self.lot_id = self.env['stock.production.lot'].sudo().search(
                    [('id', 'in', data_lot_inlocation.mapped('lot_id').ids)], limit=1,
                    order='removal_date asc, create_date asc')
            else:
                self.lot_id = False
            #nếu trường lô hàng ko dc set (khi sp có tracking) thì thông báo ko khà dụng tại tủ xuất này
            if self.product_id.tracking == 'lot' and not self.lot_id:
                raise ValidationError(_("%s không có số lượng khả dụng tại Tủ xuất này!" % self.product_id.name))
            #nếu số lượng > 0 thì kiểm tra số lượng đủ xuất ko thì thông báo
            elif self.scrap_qty > 0:
                quantity_on_hand = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                   self.location_id,
                                                                                   self.lot_id)  # check quantity trong location và lot
                if quantity_on_hand < self.scrap_qty:
                    raise ValidationError(_(
                        "Không đủ số lượng sản phẩm! \nSản phẩm ""%s"" với mã Lô hàng ""%s"" chỉ còn %d %s tại ""%s""." % (
                            self.product_id.name, self.lot_id.name, quantity_on_hand, self.product_id.uom_id.name, self.location_id.name)))
            elif self.scrap_qty <= 0:
                raise ValidationError(_("Số lượng nhập phải lớn hơn 0!"))
            else:
                raise ValidationError(_("Bạn phải chọn Tủ xuất!"))

            return {'domain': {'lot_id': [('id', 'in', data_lot_inlocation.mapped('lot_id').ids)]}}

    @api.onchange('lot_id','scrap_qty')
    def onchange_lot_id(self):
        if self.product_id and self.location_id:
            if self.scrap_qty <= 0:
                raise ValidationError(_("Số lượng nhập phải lớn hơn 0!"))

            quantity_on_hand = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                   self.location_id,
                                                                                   self.lot_id)  # check quantity trong location và lot
            if quantity_on_hand < self.scrap_qty:
                raise ValidationError(_(
                    "Không đủ số lượng sản phẩm! \nSản phẩm ""%s"" với mã Lô hàng ""%s"" chỉ còn %d %s tại ""%s""." % (
                        self.product_id.name, self.lot_id.name, quantity_on_hand, self.product_id.uom_id.name,
                        self.location_id.name)))

    # @api.onchange('scrap_qty')
    # def onchange_scrap_qty(self):
    #     if self.scrap_qty <= 0:
    #         raise ValidationError(_("Số lượng nhập phải lớn hơn 0!"))

    # chỉ hiển thị default stock khi có quyền khoa dược
    def _get_default_location_id(self):
        # company_user = self.env.user.company_id
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.companies.ids[0])], limit=1)

        # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
        institution = self.env['sh.medical.health.center'].sudo().search([('his_company', '=', self.env.companies.ids[0])],
                                                                         limit=1)
        if warehouse:
            #quyền quản lÝ kho dược
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
                return warehouse.lot_stock_id.id
            # quyen dieu duong
            elif self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                return physician_loc[0].child_ids[0].id or False
            else:
                lab_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Laboratory')], limit=1)
                spa_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Spa')], limit=1)
                laser_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Laser')], limit=1)
                surgery_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Surgery')], limit=1)
                odontology_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                    [('institution', '=', institution.id), ('type', '=', 'Odontology')], limit=1)

                #quyền bác sĩ
                grp_loc_dict = {
                    'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': lab_ward,
                    # 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
                    'shealth_all_in_one.group_sh_medical_physician_surgery': surgery_ward,
                    'shealth_all_in_one.group_sh_medical_physician_odontology': odontology_ward,
                    'shealth_all_in_one.group_sh_medical_physician_spa': spa_ward,
                    'shealth_all_in_one.group_sh_medical_physician_laser': laser_ward}

                flag = False
                for grp, loc in grp_loc_dict.items():
                    if self.env.user.has_group(grp) and loc:
                        flag = True
                        return loc.location_id.child_ids[0].id or False

                if not flag:
                    return False
        return None

    location_id = fields.Many2one('stock.location', default=_get_default_location_id)

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(SHealthStockScrap, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,
                                                               submenu=submenu)

        access_location = []
        # CƠ SỞ GẮN VỚI CTY HIỆN TẠI
        institution = self.env['sh.medical.health.center'].with_user(1).search([('his_company', '=', self.env.companies.ids[0])],
                                                                         limit=1)
        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            location_stock0 = institution.warehouse_ids[0].lot_stock_id.id  # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            lab_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                [('institution', '=', institution.id), ('type', '=', 'Laboratory')], limit=1)
            spa_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                [('institution', '=', institution.id), ('type', '=', 'Spa')], limit=1)
            laser_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                [('institution', '=', institution.id), ('type', '=', 'Laser')], limit=1)
            surgery_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                [('institution', '=', institution.id), ('type', '=', 'Surgery')], limit=1)
            odontology_ward = self.env['sh.medical.health.center.ward'].sudo().search(
                [('institution', '=', institution.id), ('type', '=', 'Odontology')], limit=1)

            grp_loc_dict = {
                'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': lab_ward,
                # 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_surgery': surgery_ward,
                'shealth_all_in_one.group_sh_medical_physician_odontology': odontology_ward,
                'shealth_all_in_one.group_sh_medical_physician_spa': spa_ward,
                'shealth_all_in_one.group_sh_medical_physician_laser': laser_ward}

            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp) and loc:
                    access_location.append(loc.location_id.id)

            # quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        doc = etree.XML(res['arch'])

        for t in doc.xpath("//" + view_type):
            t.attrib['duplicate'] = 'false'

        for node in doc.xpath("//field[@name='location_id']"):
            node_domain = "[('location_id', 'child_of', %s),('location_id.usage', '=', 'internal'),('name', 'ilike', 'Tủ')]" % (str(access_location))
            node.set("domain", node_domain)

        res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    # overide hàm này để ghi nhận lại ngày hoàn thành hủy

    def do_scrap(self):
        for scrap in self:
            move = self.env['stock.move'].create(scrap._prepare_move_values())
            # master: replace context by cancel_backorder
            move.with_context(is_scrap=True)._action_done()
            #ghi nhận lại ngày hoàn thành stock.move
            move.write({'date': scrap.sci_date_done or fields.Datetime.now()})
            # ghi nhận lại ngày hoàn thành stock.move.line
            move.move_line_ids.write({'date': scrap.sci_date_done or fields.Datetime.now()})

            scrap.write({'move_id': move.id, 'state': 'done'})
        return True

    @api.model
    def create(self, vals):#ghi nhận mã sequence khác nhau giữa xuất sd phòng và tiêu hủy
        if 'name' not in vals or vals['name'] == _('New'):
            institution = self.env['sh.medical.health.center'].search([('his_company', '=', vals['company_id'])])
            if self.env.context.get('view_for') == 'picking_scrap':
                vals['name'] = self.env['ir.sequence'].next_by_code('stock.scrap.%s'% institution.id) or _('New')
            else:
                vals['name'] = self.env['ir.sequence'].next_by_code('stock.scrap.room.use.%s'% institution.id) or _('New')
                if not vals.get('lot_id'):
                    # lấy tất cả các lô của sp ở địa điểm tủ xuất đã chọn và số lượng tồn > 0
                    data_lot_inlocation = self.env['stock.quant'].search(
                        [('product_id', '=', vals.get('product_id')), ('location_id', '=', vals.get('location_id')),
                         ('quantity', '>', 0)])

                    # nếu có data lô thì set trường lô hàng là lô sắp hết hạn
                    if data_lot_inlocation:
                        vals['lot_id'] = self.env['stock.production.lot'].sudo().search(
                            [('id', 'in', data_lot_inlocation.mapped('lot_id').ids)], limit=1,
                            order='removal_date asc, create_date asc').id


        scrap = super(SHealthStockScrap, self).create(vals)
        return scrap

# KIỂM KÊ CHI TIẾT
class SHealthInventoryLine(models.Model):
    _inherit = "stock.inventory.line"

    note = fields.Text(string='Ghi chú')

#ĐƠN MUA HÀNG NCC
class SHealthPurchaseOrder(models.Model):
    _inherit = 'purchase.order'

    @api.onchange('order_line')
    def onchange_date_order(self):
        for purchase in self:
            for line in purchase.order_line:
                line.date_planned = purchase.date_order

#ĐƠN MUA HÀNG NCC
class SHealthPurchaseOrderLine(models.Model):
    _inherit = 'purchase.order.line'

    @api.onchange('product_qty', 'product_id')
    def onchange_product_qty(self):
        if self.product_qty <= 0 and self.product_id:
            raise UserError(_("Số lượng nhập phải lớn hơn 0!"))


class wizardMultiProductPurchase(models.TransientModel):
    _name = 'wizard.multi.product.purchase'
    _description = 'Chọn nhiều sản phẩm cho đơn mua hàng'

    product_ids = fields.Many2many('product.product', 'sci_product_product_purchase_rel', 'purchase_id', 'product_id', string="Sản phẩm")
    date_order = fields.Datetime('Ngày đặt hàng', default=lambda self: fields.Datetime.now())


    def add_product(self):
        for line in self.product_ids:
            self.env['purchase.order.line'].create({
                'product_id': line.id,
                'name':line.name,
                'product_qty':1,
                'price_unit':line.standard_price,
                'order_id': self._context.get('active_id'),
                'date_planned': self.date_order,
                'product_uom': line.uom_po_id.id or line.uom_id.id
            })
        return


class SCIStockQuant(models.Model):
    _inherit = 'stock.quant'

    def _gather(self, product_id, location_id, lot_id=None, package_id=None, owner_id=None, strict=False):
        removal_strategy = self._get_removal_strategy(product_id, location_id)
        removal_strategy_order = self._get_removal_strategy_order(removal_strategy)
        domain = [
            ('product_id', '=', product_id.id),
        ]
        if not strict:
            if lot_id:
                domain = expression.AND([[('lot_id', '=', lot_id.id)], domain])
            if package_id:
                domain = expression.AND([[('package_id', '=', package_id.id)], domain])
            if owner_id:
                domain = expression.AND([[('owner_id', '=', owner_id.id)], domain])
            if self.env.context.get('exact_location'):
                domain = expression.AND([[('location_id', '=', location_id.id)], domain])
            else:
                domain = expression.AND([[('location_id', 'child_of', location_id.id)], domain])
            # print(domain)
        else:
            domain = expression.AND([[('lot_id', '=', lot_id and lot_id.id or False)], domain])
            domain = expression.AND([[('package_id', '=', package_id and package_id.id or False)], domain])
            domain = expression.AND([[('owner_id', '=', owner_id and owner_id.id or False)], domain])
            domain = expression.AND([[('location_id', '=', location_id.id)], domain])

        # Copy code of _search for special NULLS FIRST/LAST order
        self.with_user(self._uid).check_access_rights('read')
        query = self._where_calc(domain)
        self._apply_ir_rules(query, 'read')
        from_clause, where_clause, where_clause_params = query.get_sql()
        where_str = where_clause and (" WHERE %s" % where_clause) or ''
        query_str = 'SELECT "%s".id FROM ' % self._table + from_clause + where_str + " ORDER BY "+ removal_strategy_order
        self._cr.execute(query_str, where_clause_params)
        res = self._cr.fetchall()
        # No uniquify list necessary as auto_join is not applied anyways...
        return self.browse([x[0] for x in res])

    def _get_inventory_move_values(self, qty, location_id, location_dest_id, out=False):
        res = super(SCIStockQuant, self)._get_inventory_move_values(qty, location_id, location_dest_id, out)
        res['name'] = _('Kiểm kê: %s'% self.product_id.display_name)
        return res


class SCIProcurementGroup(models.Model):
    _inherit = 'procurement.group'

    @api.model
    def run_scheduler(self, use_new_cursor=False, company_id=False):
        self = self.with_context(compute_child=False, exact_location=True, separate_pick=True, do_not_confirm=True)
        return super(SCIProcurementGroup, self).run_scheduler(use_new_cursor=use_new_cursor, company_id=company_id)


class SCIStockRule(models.Model):
    _inherit = 'stock.rule'

    def _get_custom_move_fields(self):
        """Thêm điều chỉnh location_id cho stock.move trong stock.rule"""
        res = super(SCIStockRule, self)._get_custom_move_fields()
        res.append('location_id')
        return res
