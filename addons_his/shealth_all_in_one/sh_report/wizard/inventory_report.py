# -*- coding: utf-8 -*-

import base64
import time as tm
import calendar
from pytz import timezone, utc
from ast import literal_eval
from calendar import monthrange
from copy import copy
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from operator import itemgetter
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
#from odoo.tools import pycompat
from ...sh_surgery.models.mailmerge import MailMerge
from ...sh_walkins.models.sh_medical_register_for_walkin import num2words_vnm

import random

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
thick = borders.Side(style='medium')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)
thick_dotted_top_bot = borders.Border(left=thick, right=thick, top=dotted, bottom=dotted)
all_border_thick = borders.Border(left=thick, right=thick, top=thick, bottom=thick)
center_alm = Alignment(horizontal='center', vertical='center')

original_report_type = [('supply', 'Supply inventory'), ('medicine', 'Medicine inventory'),
                        ('customer', 'Customer stock out'), ('service', 'Service stock out'),
                        ('antibiotic', 'Antibiotic'), ('addictive', 'Addictive'), ('chemical', 'Chemical report'),
                        ('daily_stock_out', 'Daily stock out'), ('medicine_stock_out', 'Medicine stock out'),
                        ('simple_inventory', 'Total inventory'), ('out_dated', 'Out dated products')]


class SHInventoryReport(models.TransientModel):
    _name = 'sh.inventory.report'
    _description = 'SH Inventory report'

    def _get_report_type(self):
        report_list = [('supply', 'Kho vật tư'), ('medicine', 'Kho thuốc')]
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            report_list = [('supply', 'XNT vật tư'), ('medicine', 'XNT thuốc'),
                           ('customer', 'VTTH theo bệnh nhân'), ('service', 'VTTH theo dịch vụ'), ('customer_stock_out', 'In phiếu XK'),
                           ('antibiotic', 'Kháng sinh'), ('addictive', 'Hướng thần'), ('chemical', 'Hóa chất'),('supply_duoc', 'Vật tư Dược'),
                           ('daily_stock_out', 'Thuốc nội viện'), ('medicine_out', 'Thuốc tiêu hao'),
                           ('simple_inventory', 'XNT Dược'), ('out_dated', 'Thuốc hết hạn')]
        elif self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_subclinical')\
                and not self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_surgery')\
                and not self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_spa')\
                and not self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_odontology'):
            report_list = [('supply', 'Kho vật tư')]
        return report_list

    report_type = fields.Selection(_get_report_type,
                                   'Report type', default='supply')
    institution = fields.Many2one('sh.medical.health.center', string='Cơ sở y tế', domain=lambda self: [('his_company', 'in', self.env.companies.ids)],
                                  default=lambda self: self.env['sh.medical.health.center'].search([('his_company', '=', self.env.companies.ids[0])], limit=1))
    # using date to simplify user input, and to extract date to report
    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date', 'report_type')
    def _onchange_start_date(self):
        if self.start_date:
            if self.report_type != 'customer_stock_out':
                if self.start_date.month == fields.Date.today().month:
                    self.end_date = fields.Date.today()
                else:
                    self.end_date = date(self.start_date.year, self.start_date.month,
                                         monthrange(self.start_date.year, self.start_date.month)[1])
            else:
                self.end_date = self.start_date + timedelta(days=6)


    @api.constrains('start_date', 'end_date', 'report_type')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))
            elif record.report_type == 'customer_stock_out' and end_date - start_date > timedelta(days=6):
                raise ValidationError(
                    _("Chọn ngày kết thúc không quá 7 ngày sau ngày bắt đầu đối với Phiếu XK."))  # Todo: need translation

    def _get_inventory_data(self):

        def _calculate_mixed_quantity(product, product_moves):  # Todo: remove this later
            if not product or not product_moves:
                return 0.0
            identical_uom_moves = product_moves.filtered(lambda m: m.product_uom_id == product.uom_id)
            res = sum(identical_uom_moves.mapped('qty_done'))
            if identical_uom_moves != product_moves:
                different_uom_moves = [move for move in product_moves if move not in identical_uom_moves]
                for move in different_uom_moves:
                    res += move.product_uom_id._compute_quantity(move.product_uom_qty, product.uom_id)
            return res

        product_categ_id = {'supply': self.env.ref('shealth_all_in_one.sh_supplies').id,  # match report type
                            'medicine': self.env.ref('shealth_all_in_one.sh_medicines').id}

        # thêm điều kiện tích chọn ẩn product trong báo cáo
        product_list = self.env['product.product'].with_context(active_test=False).search([('categ_id', '=', product_categ_id.get(self.report_type, 0)), ('hide_in_report', '=', False)])  # list of supply or medicine
        false_and_return_picks = self.env['stock.picking'].search(['&', '&', ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime),
                                                                   '|', ('name', 'ilike', '-FP'), ('origin', 'ilike', '-FP')])
        # TODO: search toàn bộ location của company thay vì search lẻ nếu chậm
        # main_location = self.env['stock.location'].search([('name', '=', 'Stock'), ('company_id', '=', self.institution.his_company.id)])
        main_location = self.institution.warehouse_ids[0].lot_stock_id
        scrap_location = self.env['stock.location'].search([('name', 'ilike', 'Scrap'), ('company_id', '=', self.institution.his_company.id)])
        room_use_location = self.env['stock.location'].search([('name', 'ilike', 'Sử dụng phòng'), ('company_id', '=', self.institution.his_company.id)])
        report_location = self.institution.location_medicine_stock if self.report_type == 'medicine' \
            else self.env['stock.location'].search([('location_id', '=', main_location.id), ('company_id', '=', self.institution.his_company.id), ('location_institution_type', '=', 'supply'), ('name', 'not ilike', 'cấp cứu')], limit=1)
        reception_location = self.env['stock.location'].search([('company_id', '=', self.institution.his_company.id), ('name', 'ilike', 'lễ tân')], limit=1)
        if self.env.user.has_group('stock.group_stock_manager'):
            locations = self.env['sh.medical.health.center.ot'].search([('institution', '=', self.institution.id)]).mapped('location_%s_stock' % self.report_type) + reception_location
        else:
            locations = self.env['stock.location'].browse()
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_subclinical'):
                lab_ward = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Laboratory')], limit=1)
                locations += self.env['stock.location'].search([('location_id', '=', lab_ward.location_id.id),
                                                                ('location_institution_type', '=', self.report_type), ('company_id', '=', self.institution.his_company.id)])
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_surgery'):
                surgery_ward = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Surgery')], limit=1)
                locations += self.env['stock.location'].search([('location_id', '=', surgery_ward.location_id.id),
                                                                ('location_institution_type', '=', self.report_type), ('company_id', '=', self.institution.his_company.id)])
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_spa'):
                spa_ward = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Spa')], limit=1)
                locations += self.env['stock.location'].search([('location_id', '=', spa_ward.location_id.id),
                                                                ('location_institution_type', '=', self.report_type), ('company_id', '=', self.institution.his_company.id)])
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_laser'):
                laser_ward = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Laser')], limit=1)
                locations += self.env['stock.location'].search([('location_id', '=', laser_ward.location_id.id),
                                                                ('location_institution_type', '=', self.report_type), ('company_id', '=', self.institution.his_company.id)])
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician_odontology'):
                odon_ward = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Odontology')], limit=1)
                locations += self.env['stock.location'].search([('location_id', '=', odon_ward.location_id.id),
                                                                ('location_institution_type', '=', self.report_type), ('company_id', '=', self.institution.his_company.id)])
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_receptionist'):
                locations += reception_location

        product_datas = []
        for product in product_list:
            product_lot_list = []
            all_quant = self.env['stock.quant'].search([('location_id', 'child_of', main_location.id),
                                                        ('product_id', '=', product.id)])
            moves = self.env['stock.move.line'].search([('date', '>=', self.start_datetime),
                                                        ('state', '=', 'done'),
                                                        ('product_id', '=', product.id),
                                                        ('picking_id', 'not in', false_and_return_picks.ids),
                                                        '|', ('location_id', 'child_of', main_location.id),
                                                        ('location_dest_id', 'child_of', main_location.id)])
            for quant in all_quant:
                if quant.lot_id not in product_lot_list:
                    product_lot_list.append(quant.lot_id)
            for move in moves:
                if move.lot_id not in product_lot_list:
                    product_lot_list.append(move.lot_id)
            if product_lot_list:
                # dòng tổng hợp các line theo lot của 1 product, tắt tạm thời
                # product_group_line = [product.default_code, product.name, product.uom_id.name, 'group'] + [''] * 4 + [0] * (21 + len(locations) * 8) # group lines by product
                # product_datas.append(product_group_line)
                for lot in product_lot_list:
                    if lot:
                        lot_unit_price = lot.lot_price / lot.product_init_qty if lot.product_init_qty > 0 else 0
                    else:
                        lot_unit_price = product.standard_price
                    pl_moves = moves.filtered(lambda m: m.product_id == product and m.lot_id == lot)

                    product_lot_line = [product.default_code or '', product.name, product.uom_id.name, '', '', lot.name or '', lot.removal_date or '', lot_unit_price]
                    # all internal locations data
                    moves_in_total = pl_moves.filtered(lambda m: m.location_id.usage != 'internal' and m.location_dest_id.usage == 'internal')
                    moves_out_patient_total = pl_moves.filtered(lambda m: m.location_id.usage == 'internal' and m.location_dest_id.usage == 'customer' and ((not m.origin) or (m.origin and 'SO' not in m.origin)))
                    moves_out_sale_total = pl_moves.filtered(lambda m: m.location_id.usage == 'internal' and m.location_dest_id.usage == 'customer' and (m.origin and 'SO' in m.origin))
                    moves_out_scrap_total = pl_moves.filtered(lambda m: m.location_id.usage == 'internal' and m.location_dest_id == scrap_location)
                    moves_out_room_use_total = pl_moves.filtered(lambda m: m.location_id.usage == 'internal' and m.location_dest_id == room_use_location)
                    moves_out_supplier_total = pl_moves.filtered(lambda m: m.location_id.usage == 'internal' and m.location_dest_id.usage == 'supplier')
                    quant_total = all_quant.filtered(lambda q: q.product_id == product and q.lot_id == lot and q.location_id.usage == 'internal')
                    qty_in_total = _calculate_mixed_quantity(product, moves_in_total.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_patient_total = _calculate_mixed_quantity(product, moves_out_patient_total.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_sale_total = _calculate_mixed_quantity(product, moves_out_sale_total.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_scrap_total = _calculate_mixed_quantity(product, moves_out_scrap_total.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_room_use_total = _calculate_mixed_quantity(product, moves_out_room_use_total.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_supplier_total = _calculate_mixed_quantity(product, moves_out_supplier_total.filtered(lambda m: m.date <= self.end_datetime))
                    qty_end_total = sum(quant_total.mapped('quantity'))
                    if self.end_datetime < fields.Datetime.now():
                        later_qty_in_total = _calculate_mixed_quantity(product, moves_in_total.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_patient_total = _calculate_mixed_quantity(product, moves_out_patient_total.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_sale_total = _calculate_mixed_quantity(product, moves_out_sale_total.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_scrap_total = _calculate_mixed_quantity(product, moves_out_scrap_total.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_room_use_total = _calculate_mixed_quantity(product, moves_out_room_use_total.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_supplier_total = _calculate_mixed_quantity(product, moves_out_supplier_total.filtered(lambda m: m.date > self.end_datetime))
                        qty_end_total += (later_qty_out_patient_total + later_qty_out_sale_total + later_qty_out_scrap_total + later_qty_out_room_use_total + later_qty_out_supplier_total - later_qty_in_total)
                    qty_begin_total = qty_end_total + qty_out_patient_total + qty_out_sale_total + qty_out_scrap_total + qty_out_room_use_total + qty_out_supplier_total - qty_in_total
                    product_lot_line.extend([qty_begin_total, qty_begin_total*lot_unit_price, qty_in_total, qty_in_total*lot_unit_price, qty_out_patient_total, qty_out_sale_total, qty_out_scrap_total, qty_out_room_use_total, qty_out_supplier_total,
                                             (qty_out_patient_total + qty_out_scrap_total + qty_out_room_use_total + qty_out_supplier_total)*lot_unit_price, qty_end_total, qty_end_total*lot_unit_price])
                    # main location data
                    moves_in_main_ext = pl_moves.filtered(lambda m: m.location_dest_id == report_location and m.location_id.usage == 'supplier')
                    moves_in_main_int = pl_moves.filtered(lambda m: m.location_dest_id == report_location and m.location_id.usage == 'internal' and m.location_id != main_location)
                    moves_out_main_supply = pl_moves.filtered(lambda m: m.location_id == report_location and m.location_dest_id.usage == 'supplier')
                    moves_out_main_int = pl_moves.filtered(lambda m: m.location_id == report_location and m.location_dest_id.usage == 'internal')
                    moves_out_main_customer = pl_moves.filtered(lambda m: m.location_id == report_location and m.location_dest_id.usage == 'customer' and ((not m.origin) or (m.origin and 'SO' not in m.origin)))
                    moves_out_main_sale = pl_moves.filtered(lambda m: m.location_id == report_location and m.location_dest_id.usage == 'customer' and (m.origin and 'SO' in m.origin))
                    moves_out_main_scrap = pl_moves.filtered(lambda m: m.location_id == report_location and m.location_dest_id == scrap_location)
                    qty_in_main_ext = _calculate_mixed_quantity(product, moves_in_main_ext.filtered(lambda m: m.date <= self.end_datetime))
                    qty_in_main_int = _calculate_mixed_quantity(product, moves_in_main_int.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_main_supply = _calculate_mixed_quantity(product, moves_out_main_supply.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_main_int = _calculate_mixed_quantity(product, moves_out_main_int.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_main_customer = _calculate_mixed_quantity(product, moves_out_main_customer.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_main_sale = _calculate_mixed_quantity(product, moves_out_main_sale.filtered(lambda m: m.date <= self.end_datetime))
                    qty_out_main_scrap = _calculate_mixed_quantity(product, moves_out_main_scrap.filtered(lambda m: m.date <= self.end_datetime))
                    quant_main = quant_total.filtered(lambda q: q.location_id == report_location)
                    qty_end_main = sum(quant_main.mapped('quantity'))
                    if self.end_datetime < fields.Datetime.now():
                        later_qty_in_main_ext = _calculate_mixed_quantity(product, moves_in_main_ext.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_in_main_int = _calculate_mixed_quantity(product, moves_in_main_int.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_main_supply = _calculate_mixed_quantity(product, moves_out_main_supply.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_main_int = _calculate_mixed_quantity(product, moves_out_main_int.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_main_customer = _calculate_mixed_quantity(product, moves_out_main_customer.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_main_sale = _calculate_mixed_quantity(product, moves_out_main_sale.filtered(lambda m: m.date > self.end_datetime))
                        later_qty_out_main_scrap = _calculate_mixed_quantity(product, moves_out_main_scrap.filtered(lambda m: m.date > self.end_datetime))
                        qty_end_main += (later_qty_out_main_supply + later_qty_out_main_int + later_qty_out_main_customer + later_qty_out_main_sale + later_qty_out_main_scrap - later_qty_in_main_ext - later_qty_in_main_int)
                    qty_begin_main = qty_end_main + qty_out_main_supply + qty_out_main_int + qty_out_main_customer + qty_out_main_sale + qty_out_main_scrap - qty_in_main_ext - qty_in_main_int
                    product_lot_line.extend([qty_begin_main, qty_in_main_ext, qty_in_main_int, qty_out_main_supply, qty_out_main_int, qty_out_main_customer, qty_out_main_sale, qty_out_main_scrap, qty_end_main])
                    # specific location data
                    for location in locations:
                        moves_in_int = pl_moves.filtered(lambda m: m.location_dest_id == location and m.location_id == report_location)
                        moves_out_patient = pl_moves.filtered(lambda m: m.location_id == location and m.location_dest_id.usage == 'customer' and '-FP' not in m.reference and ((not m.origin) or (m.origin and 'SO' not in m.origin)))
                        moves_out_int = pl_moves.filtered(lambda m: m.location_id == location and m.location_dest_id == report_location)
                        moves_out_scrap = pl_moves.filtered(lambda m: m.location_id == location and m.location_dest_id == room_use_location)
                        move_out_sale = pl_moves.filtered(lambda m: m.location_id == location and m.location_dest_id.usage == 'customer' and (m.origin and 'SO' in m.origin))
                        move_in_sale = pl_moves.filtered(lambda m: m.location_dest_id == location and m.location_id.usage == 'customer' and (m.origin and 'SO' in m.origin))
                        quant_at_location = quant_total.filtered(lambda q: q.location_id == location)
                        qty_in_int = _calculate_mixed_quantity(product, moves_in_int.filtered(lambda m: m.date <= self.end_datetime))
                        qty_out_patient = _calculate_mixed_quantity(product, moves_out_patient.filtered(lambda m: m.date <= self.end_datetime))
                        qty_out_int = _calculate_mixed_quantity(product, moves_out_int.filtered(lambda m: m.date <= self.end_datetime))
                        qty_out_scrap = _calculate_mixed_quantity(product, moves_out_scrap.filtered(lambda m: m.date <= self.end_datetime))
                        qty_final_sale = _calculate_mixed_quantity(product, move_out_sale.filtered(lambda m: m.date <= self.end_datetime)) - _calculate_mixed_quantity(product, move_in_sale.filtered(lambda m: m.date <= self.end_datetime))
                        qty_end = sum(quant_at_location.mapped('quantity'))
                        if self.end_datetime < fields.Datetime.now():
                            later_qty_in_int = _calculate_mixed_quantity(product, moves_in_int.filtered(lambda m: m.date > self.end_datetime))
                            later_qty_out_patient = _calculate_mixed_quantity(product, moves_out_patient.filtered(lambda m: m.date > self.end_datetime))
                            later_qty_out_int = _calculate_mixed_quantity(product, moves_out_int.filtered(lambda m: m.date > self.end_datetime))
                            later_qty_out_scrap = _calculate_mixed_quantity(product, moves_out_scrap.filtered(lambda m: m.date > self.end_datetime))
                            later_qty_final_sale = _calculate_mixed_quantity(product, move_out_sale.filtered(lambda m: m.date > self.end_datetime)) - _calculate_mixed_quantity(product, move_in_sale.filtered(lambda m: m.date > self.end_datetime))
                            qty_end += (later_qty_out_patient + later_qty_out_int + later_qty_out_scrap + later_qty_final_sale - later_qty_in_int)
                        # calculate begin quantity
                        qty_begin = qty_end + qty_out_int + qty_out_patient + qty_out_scrap + qty_final_sale - qty_in_int
                        product_lot_line.extend([qty_begin, qty_in_int, qty_out_patient, qty_final_sale, qty_out_int, qty_out_scrap, 0, qty_end])
                    # for i in range(8, 29 + len(locations) * 8):
                    #     product_group_line[i] += product_lot_line[i]

                    product_datas.append(product_lot_line)

        product_datas.sort(key=itemgetter(1))
        # print(product_datas)
        return locations.mapped('name'), product_datas

    #HÀM LẤY DATA VTTH THEO BỆNH NHÂN/THEO DỊCH VỤ
    #HÀM CŨ LẤY THEO STOCK MOVE
    # def _get_stock_out_data(self):
    #     customer_location = self.env.ref('stock.stock_location_customers')
    #     false_pick = self.env['stock.picking'].search([('name', 'ilike', '-FP'), ('location_dest_id', 'child_of', customer_location.id), ('company_id', '=', self.institution.his_company.id),
    #                                                    ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime)])
    #     domain = [('picking_id', 'not in', false_pick.ids), ('location_dest_id', 'child_of', customer_location.id), ('company_id', '=', self.institution.his_company.id),
    #               ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime), ('state', '=', 'done')]
    #     datas = []
    #     if self.report_type == 'customer':
    #         moves = self.env['stock.move'].search(domain)
    #         partners = moves.mapped('picking_partner_id')
    #         for partner in partners:
    #             partner_lines = [partner.name]
    #             product_lines = []  # for sorting
    #             product_index = {}
    #             index = 0
    #             partner_moves = moves.filtered(lambda m: m.picking_partner_id == partner)
    #             for partner_move in partner_moves:
    #                 product = partner_move.product_id
    #                 if str(product.id) not in product_index:
    #                     product_lines.append([index+1, product.default_code, product.name, product.uom_id.name, partner_move.quantity_done, partner_move.date])
    #                     product_index[str(product.id)] = index
    #                     index += 1
    #                 else:
    #                     product_lines[product_index[str(product.id)]][4] += partner_move.quantity_done
    #             # product_lines.sort(key=itemgetter(1))
    #             partner_lines.extend(product_lines)
    #
    #             datas.append(partner_lines)
    #         # datas.sort(key=itemgetter(1))
    #
    #     elif self.report_type == 'service':
    #         domain += [('origin', 'ilike', '-[')]
    #         origins = self.env['stock.move'].search_read(domain, ['origin'])
    #         origins = [o['origin'] for o in origins]
    #         service_dict = {}
    #         for origin in origins:
    #             # origin_moves = moves.filtered(lambda m: m.origin == origin)  # filter process 3x slower with text fields
    #             origin_moves = self.env['stock.move'].search([('origin', '=', origin)])
    #             walkin, service_list_str = origin.split('-')
    #             service_list = literal_eval(service_list_str)
    #             for move in origin_moves:
    #                 product = move.product_id
    #                 p_id = product.id
    #                 p_code = product.default_code
    #                 p_name = product.name
    #                 p_uom = product.uom_id.name
    #                 qty = move.product_qty/len(service_list)
    #                 for service in service_list:
    #                     if not service_dict.get(service):
    #                         service_dict[service] = ([walkin], {})
    #                     elif walkin not in service_dict[service][0]:
    #                         service_dict[service][0].append(walkin)
    #                     if not service_dict[service][1].get(str(p_id)):
    #                         service_dict[service][1][str(p_id)] = [p_code, p_name, p_uom, qty]
    #                     else:
    #                         service_dict[service][1][str(p_id)][3] += qty
    #         for service in service_dict:
    #             service_name = self.env['sh.medical.health.center.service'].browse(int(service)).name
    #             num_service = len(service_dict[service][0])
    #             service_line = [service_name + ' - ' + str(num_service) + 'DV']
    #             index = 1
    #             for product_key in service_dict[service][1]:
    #                 service_line.append([index] + service_dict[service][1][product_key])
    #                 index += 1
    #             datas.append(service_line)
    #     return datas

    # HÀM LẤY DATA VTTH THEO BỆNH NHÂN/THEO DỊCH VỤ
    def _get_stock_out_data(self):
        domain = [('institution', '=', self.institution.id), ('date_out', '>=', self.start_datetime), ('date_out', '<=', self.end_datetime)]
        datas = []
        if self.report_type == 'customer':
            moves = self.env['sh.medical.walkin.material'].search(domain)
            patients = moves.mapped('patient')
            for patient in patients:
                product_lines = []  # for sorting
                product_index = {}
                index = 0
                patient_moves = moves.filtered(lambda m: m.patient == patient)
                for patient_move in patient_moves:
                    product = patient_move.product_id
                    if str(product.id) not in product_index:
                        product_lines.append([index+1, product.default_code, product.name, product.uom_id.name, patient_move.init_quantity, patient_move.quantity,
                                              patient_move.init_quantity - patient_move.quantity])
                        product_index[str(product.id)] = index
                        index += 1
                    else:
                        product_lines[product_index[str(product.id)]][4] += patient_move.init_quantity
                        product_lines[product_index[str(product.id)]][5] += patient_move.quantity
                        product_lines[product_index[str(product.id)]][6] = product_lines[product_index[str(product.id)]][4] - product_lines[product_index[str(product.id)]][5]

                # product_lines.sort(key=itemgetter(1))
                patient_lines = [[patient.display_name]]
                patient_lines.extend(product_lines)

                datas.append(patient_lines)
            # datas.sort(key=itemgetter(1))

        elif self.report_type == 'service':
            moves = self.env['sh.medical.walkin.material'].search(domain)
            services = moves.mapped('services')
            for service in services:
                product_lines = []  # for sorting
                product_index = {}
                index = 0
                service_moves = moves.filtered(lambda m: service in m.services)
                total_services = service_moves.mapped('walkin')

                for service_move in service_moves:
                    product = service_move.product_id
                    if str(product.id) not in product_index:
                        product_lines.append([index + 1, product.default_code, product.name, product.uom_id.name,
                                              service_move.init_quantity, service_move.quantity/len(service_move.services),
                                             service_move.init_quantity - service_move.quantity/len(service_move.services), service_move.department.name])

                        product_index[str(product.id)] = index
                        index += 1
                    else:
                        product_lines[product_index[str(product.id)]][4] += service_move.init_quantity
                        product_lines[product_index[str(product.id)]][5] += service_move.quantity/len(service_move.services)
                        product_lines[product_index[str(product.id)]][6] = product_lines[product_index[str(product.id)]][4] - product_lines[product_index[str(product.id)]][5]

                service_lines = [[service.display_name + ' - ' + str(len(total_services)) + 'DV', service_move.department.name]]
                # product_lines.sort(key=itemgetter(1))
                service_lines.extend(product_lines)
                datas.append(service_lines)

        return datas

    def report_inventory(self):
        inventory_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.inventory_report_attachment').id)
        decode = base64.b64decode(inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        report_type = {'supply': 'vật tư', 'medicine': 'thuốc'}.get(self.report_type)  # match report type
        ws['b1'].value = 'THỐNG KÊ NHẬP XUẤT TỒN %s %s' % (report_type.upper(), self.institution.name.upper())
        ws['b2'].value = 'Từ 00:00 %s đến 23:59 %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        ws['c4'].value = 'Mã ' + report_type
        ws['d4'].value = 'Tên ' + report_type
        ws['w3'].value = 'Kho %s bệnh viện' % report_type
        ws['af3'].value = 'Các tủ ' + report_type
        data = self._get_inventory_data()
        col = 32
        head_font = Font(name='Times New Roman', size=12, bold=True)
        location_titles = ['Tồn đầu', 'Nhập từ kho %s BV' % report_type, 'Xuất bệnh nhân', 'Xuất bán', 'Xuất trả kho %s BV' % report_type, 'Xuất sử dụng phòng', 'Xuất đoàn Khám sức khỏe', 'Tồn cuối']
        title_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        for location_name in data[0]:
            start_col = col
            location_cell = ws.cell(row=4, column=col)
            location_cell.value, location_cell.font, location_cell.fill = location_name, head_font, title_fill
            location_cell.border, location_cell.alignment = all_border_thin, Alignment(horizontal='center', vertical='center')
            for n, title in enumerate(location_titles):
                title_cell = ws.cell(row=5, column=col)
                title_cell.value, title_cell.font, title_cell.fill = title, head_font, title_fill
                title_cell.border, title_cell.alignment = all_border_thin, Alignment(horizontal='center', vertical='center', wrap_text=True)
                col += 1
            ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=col-1)
        ws.unmerge_cells(start_row=3, start_column=32, end_row=3, end_column=39)
        ws.merge_cells(start_row=3, start_column=32, end_row=3, end_column=(max(39, col-1)))
        ws.sheet_properties.outlinePr.summaryBelow = False

        row = 6
        index = 0
        sub_index = 0
        start_grp = 7
        line_font = Font(name='Times New Roman', size=12)
        line_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        group_font = Font(name='Times New Roman', size=12, bold=True)
        for line in data[1]:
            ws.cell(row=row, column=2).border, ws.cell(row=row, column=2).alignment = all_border_gray, Alignment(horizontal='center')
            # trường hợp dùng lô
            # if line[3] == 'group':
            #     index += 1
            #     ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = index, group_font
            #     ws.row_dimensions.group(start=start_grp, end=row-1, hidden=True)
            #     sub_index = 0
            #     start_grp = row + 1
            # else:
            #     ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = '.'.join([str(index), str(sub_index)]), line_font
            ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = index+1, line_font
            index += 1
            for n, value in enumerate(line, 3):
                cell = ws.cell(row=row, column=n)
                cell.border = all_border_gray
                if n == 6 and value == 'group':
                    continue
                cell.value = value
                cell.font = line_font if line[3] != 'group' else group_font
                if n == 5:
                    cell.alignment = Alignment(horizontal='center')
                if n < 9:
                    cell.number_format = '@'
                if n == 9:
                    cell.number_format = 'dd/mm/yyyy'
                if n > 9:
                    cell.number_format = '0.00' if (isinstance(value, float) and not value.is_integer()) else '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
                    cell.alignment = Alignment(horizontal='right')
                    if n > 22:
                        cell.fill = line_fill
            sub_index += 1
            row += 1
        # ws.row_dimensions.group(start=start_grp, end=row-1, hidden=True)  # trường hợp dùng lô

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Inventory report.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})
        return {'name': 'Inventory report',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def report_stock_out(self):
        inventory_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.stock_out_report_attachment').id)
        decode = base64.b64decode(inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        datas = self._get_stock_out_data()
        print(datas)
        if self.report_type == 'service':
            ws['b1'].value = 'THỐNG KÊ XUẤT VẬT TƯ / THUỐC THEO DỊCH VỤ'
            ws['b5'].value = 'Dịch vụ'
            ws['j5'].value, ws['j5'].border = 'Dịch vụ', all_border_thin

        ws['b2'].value = self.institution.name
        ws['b3'].value = 'Từ 00:00 %s đến 23:59 %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        line_font = Font(name='Times New Roman', size=12)
        row = 6
        end_col = 11 if self.report_type == 'service' else 10
        for group in datas:
            ws.cell(row, 2).value, ws.cell(row, 2).font = group[0][0], line_font
            if self.report_type == 'service':
                ws.cell(row, 10).value, ws.cell(row, 10).font = group[0][1], line_font
            for col in range(2, end_col):
                ws.cell(row, col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                ws.cell(row, col).border = all_border_thin
            row += 1
            for i in range(1, len(group)):
                for col in range(3, end_col):
                    ws.cell(row, col).value, ws.cell(row, col).font = group[i][col-3], line_font
                for col in range(2, end_col):
                    ws.cell(row, col).border = all_border_thin
                row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Stock out report.xlsx',
                                                             'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Inventory report',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def customer_stock_out_report(self):
        wb = Workbook()
        ws = wb.active

        #  Todo: consider using stockpick data or walkins data
        walkins = self.env['sh.medical.appointment.register.walkin'].search([('state', '=', 'Completed'),
                                                                             ('date', '>=', self.start_datetime),
                                                                             ('date', '<=', self.end_datetime),
                                                                             '!', ('material_ids', '=', False)],
                                                                            order='date')
        fmt = "%Y-%m-%d %H:%M:%S"
        now_utc = datetime.now(timezone('UTC'))
        local_tz = self.env.user.tz or 'Etc/GMT-7'
        now_timezone = now_utc.astimezone(timezone(local_tz))
        UTC_OFFSET_TIMEDELTA = datetime.strptime(now_utc.strftime(fmt), fmt) - datetime.strptime(
            now_timezone.strftime(fmt), fmt)
        year_start = datetime(self.start_date.year, 1, 1, 0, 0, 0) + UTC_OFFSET_TIMEDELTA

        walkin_index = self.env['sh.medical.appointment.register.walkin'].search_count([('state', '=', 'Completed'),
                                                                                        ('date', '>=', year_start),
                                                                                        ('date', '<', self.start_datetime),
                                                                                        '!', ('material_ids', '=', False)])

        col_widths = [('a', 7.56), ('b', 43.33), ('c', 11), ('e', 9), ('f', 9)]
        for value in col_widths:
            ws.column_dimensions[value[0]].width = value[1]

        institute = self.institution
        address1 = ', '.join([a for a in (institute.street, institute.street2) if a])
        address2 = ', '.join([a for a in (institute.city, institute.state_id.name, institute.country_id.name) if a])
        institute_header = '\n'.join([a for a in (institute.company_name, address1, address2) if a])
        ws.oddHeader.left.text = institute_header
        ws.oddHeader.right.text = '&BMẫu số: 02 - VT&B\n' \
                                  '(Ban hành theo Thông tư số 200/2014/TT-BTC\n' \
                                  'Ngày 22/12/2014 của Bộ Tài Chính)'
        # ws.print_title_rows = '12:13'
        ws.page_margins.top = 1.2
        main_font = Font(name='Times New Roman', size=12)
        alignment_center = Alignment(horizontal='center', vertical='center')
        alignment_left = Alignment(horizontal='left', vertical='center')
        row = 3

        for walkin in walkins:
            if not walkin.material_ids:
                continue
            walkin_index += 1
            walkin_index_str = '/'.join([self.start_date.strftime('%y'), str(walkin_index).zfill(4)])
            ws.row_dimensions[row].height = 21
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row, 1).value = 'PHIẾU XUẤT KHO'
            ws.cell(row, 1).font = Font(name='Times New Roman', size=16, bold=True)
            ws.cell(row, 1).alignment = alignment_center
            row += 1
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row, 2).value = walkin.date.strftime('                                   Ngày %d tháng %m năm %Y')
            ws.cell(row, 2).font = Font(name='Times New Roman', size=12, italic=True)
            ws.cell(row, 2).alignment = alignment_center
            ws.cell(row, 5).value = 'Nợ: 621'
            ws.cell(row, 5).font = Font(name='Times New Roman', size=12)
            row += 1
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row, 2).value = '                                                    Số: PXK%s' % walkin_index_str
            ws.cell(row, 2).font = main_font
            ws.cell(row, 2).alignment = alignment_left
            ws.cell(row, 5).value = 'Có: 152'
            ws.cell(row, 5).font = Font(name='Times New Roman', size=12)
            row += 2
            for i in ['- Họ và tên người nhận hàng: %s' % walkin.patient.name,
                      '- Lý do xuất kho: %s' % '; '.join(walkin.service.mapped('name')),
                      '- Xuất tại kho: Kho dược']:
                row += 1
                ws.cell(row, 1).value, ws.cell(row, 1).font = i, main_font
                ws.row_dimensions[row].height = 19.5

            row += 2
            ws.row_dimensions[row].height = 31.5
            ws.row_dimensions[row + 1].height = 21
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
            ws.cell(row, 1).value, ws.cell(row, 1).font, ws.cell(row, 1).alignment = 'STT', main_font, alignment_center
            ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
            ws.cell(row, 2).value, ws.cell(row, 2).font, ws.cell(row, 2).alignment = 'Tên sản phẩm', main_font, alignment_center
            ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=3)
            ws.cell(row, 3).value, ws.cell(row, 3).font, ws.cell(row, 3).alignment = 'Mã số', main_font, alignment_center
            ws.merge_cells(start_row=row, start_column=4, end_row=row + 1, end_column=4)
            ws.cell(row, 4).value, ws.cell(row, 4).font, ws.cell(row, 4).alignment = 'ĐVT', main_font, alignment_center
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            ws.cell(row, 5).value, ws.cell(row, 5).font, ws.cell(row, 5).alignment = 'Số lượng', main_font, alignment_center
            for i in range(1, 7):
                ws.cell(row, i).border = ws.cell(row + 1, i).border = all_border_thin
            row += 1
            ws.cell(row, 5).value, ws.cell(row, 5).font, ws.cell(row, 5).alignment = 'Chứng từ', main_font, alignment_center
            ws.cell(row, 6).value, ws.cell(row, 6).font, ws.cell(row, 6).alignment = 'Thực xuất', main_font, alignment_center

            row += 1
            # inserting data here
            index = 1
            for mats in walkin.material_ids:
                ws.cell(row, 1).value, ws.cell(row, 1).font, ws.cell(row, 1).alignment = index, main_font, alignment_center
                ws.cell(row, 1).border = ws.cell(row, 2).border = all_border_thin
                # ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                for col, value in enumerate([mats.product_id.name, mats.product_id.default_code, mats.product_id.uom_id.name, mats.quantity, mats.quantity], 2):
                    ws.cell(row, col).value, ws.cell(row, col).font = value, main_font
                    ws.cell(row, col).border = all_border_thin
                    ws.cell(row, col).alignment = alignment_center if col > 2 else alignment_left
                row += 1
                index += 1
            # end of data

            row += 3
            bold_12 = Font(name='Times New Roman', size=12, bold=True)
            italic_12 = Font(name='Times New Roman', size=12, italic=True)
            ws.row_dimensions[row].height = 23.5
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row, 1).value = '    Người lập phiếu       Người nhận hàng        Thủ kho        Kế toán trưởng          Giám đốc'
            ws.cell(row, 1).font, ws.cell(row, 1).alignment = bold_12, alignment_left
            row += 1
            ws.row_dimensions[row].height = 29.5
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row, 1).value = '        (Ký, họ tên)             (Ký, họ tên)            (Ký, họ tên)        (Ký, họ tên)      (Ký, họ tên, đóng dấu)'
            ws.cell(row, 1).alignment = alignment_left
            ws.cell(row, 1).font = italic_12

            row += 7
            ws.row_breaks.append(Break(id=row))
            row += 3

        row -= 3 if row > 3 else 1
        ws.sheet_view.showGridLines = False
        ws.print_area = 'A1:F%s' % str(row)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Stock out report.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Stock out report',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def antibiotic_report(self):
        datas = {}
        antibiotic_categ = self.env['sh.medical.medicines.category'].search([('name', '=', 'Kháng sinh')])
        antibiotic_medicines = self.env['sh.medical.medicines'].with_context(active_test=False).search([('medicine_category_id', '=', antibiotic_categ.id)])
        false_pick = self.env['stock.picking'].search([('name', 'ilike', '-FP'), ('location_dest_id', 'child_of', self.env.ref('stock.stock_location_customers').id),
                                                       ('company_id', '=', self.institution.his_company.id), ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime)])
        for medicine in antibiotic_medicines:
            if medicine.index_medicine:
                medicine_data = [medicine.name_medicine or medicine.name_use, medicine.origin.name or '', 'not',
                                 medicine.uom_id.name, 'not']
                moves = self.env['stock.move'].search([('product_id', '=', medicine.product_id.id), ('state', '=', 'done'),
                                                       ('location_dest_id', '=', self.env.ref('stock.stock_location_customers').id),
                                                       ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime),
                                                       ('picking_id', 'not in', false_pick.ids), ('company_id', '=', self.institution.his_company.id)])
                qty_done = sum(moves.mapped('quantity_done'))  # quantity done is a computed field without storing thus we cannot use read_group
                price = medicine.product_id.standard_price  # also a computed field
                medicine_data += [qty_done, price]
                if not datas.get(medicine.index_medicine):
                    datas[medicine.index_medicine] = [medicine_data]
                else:
                    datas[medicine.index_medicine].append(medicine_data)
        antibiotic_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.antibiotic_report_attachment').id)
        decode = base64.b64decode(antibiotic_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a3'].value = self.institution.name
        ws['d4'].value = '(Từ ngày %s đến ngày %s)' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))

        for row in ws.iter_rows(8, 300, 4, 4):
            for cell in row:
                if cell.value != '' and datas.get(str(cell.value)):
                    r = cell.row
                    for medicine_data in datas.get(str(cell.value)):
                        if datas.get(str(cell.value)).index(medicine_data) > 0:
                            r += 1
                            ws.insert_rows(idx=r)
                            for col in range(1, 12):
                                ws.cell(r, col).border = copy(ws.cell(r-1, col).border)
                                ws.cell(r-1, col).border = dotted_top_bot
                                ws.cell(r, col).fill = copy(ws.cell(r-1, col).fill)
                                ws.cell(r, col).font = copy(ws.cell(r-1, col).font)
                                ws.cell(r, col).alignment = copy(ws.cell(r-1, col).alignment)
                                if col in (7, 9):
                                    ws.cell(r, col).value = copy(ws.cell(r-1, col).value)
                        for col, val in enumerate(medicine_data, 5):
                            if val == 'not':
                                continue
                            else:
                                ws.cell(r, col).value = val

        for row in ws.iter_rows(258, 300, 7, 7):
            for cell in row:
                if cell.value == 'Ngày':
                    r = cell.row
                    cell.value = date.today().strftime('Ngày %d tháng %m năm %Y')
                    ws.merge_cells('B%s:C%s' % (r+1, r+1))
                    ws.cell(r+1, 2).value = 'KHOA DƯỢC'
                    ws.cell(r+1, 5).value = 'PHÒNG TÀI CHÍNH KẾ TOÁN                           GIÁM ĐỐC BỆNH VIỆN'
                    ws.cell(r+1, 2).font = ws.cell(r+1, 5).font = Font(name='Times New Roman', size=12, bold=True)
                    ws.cell(r+1, 2).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(r + 1, 5).alignment = Alignment(horizontal='left', vertical='center')
                    break

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Báo cáo kháng sinh.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo kháng sinh',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def addictive_report(self):
        datas = {}
        Prd = self.env['product.product']
        Move = self.env['stock.move']
        main_location = self.institution.warehouse_ids[0].lot_stock_id
        supply_location = self.env.ref('stock.stock_location_suppliers')
        customer_location = self.env.ref('stock.stock_location_customers')
        addictive_categ = self.env['sh.medical.medicines.category'].search([('name', 'in', ('Thuốc gây nghiện', 'Thuốc hướng thần'))])
        addictive_medicines = self.env['sh.medical.medicines'].with_context(active_test=False).search([('medicine_category_id', 'in', addictive_categ.ids)])
        false_and_return_picks = self.env['stock.picking'].search(['&', '&', ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime),
                                                                   '|', ('name', 'ilike', '-FP'), ('origin', 'ilike', '-FP')])
        # product data
        table_data = []
        moves_domain = [('product_id', 'in', addictive_medicines.mapped('product_id.id')), ('picking_id', 'not in', false_and_return_picks.ids),
                        ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime), ('state', '=', 'done'), ('company_id', '=', self.institution.his_company.id)]
        moves_in_domain = moves_domain + [('location_dest_id', 'child_of', main_location.id), '!', ('location_id', 'child_of', main_location.id)]
        moves_out_domain = moves_domain + [('location_id', 'child_of', main_location.id), '!', ('location_dest_id', 'child_of', main_location.id)]
        moves_in_res = dict((item['product_id'][0], item['product_qty']) for item in Move.read_group(moves_in_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
        moves_out_res = dict((item['product_id'][0], item['product_qty']) for item in Move.read_group(moves_out_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
        index = 1
        for medicine in addictive_medicines:
            product = medicine.product_id
            start_qty = Prd.with_context(to_date=self.start_datetime).browse(product.id).qty_available or 0.0
            in_qty = moves_in_res.get(product.id, 0.0)
            start_in_qty = start_qty + in_qty
            out_qty = moves_out_res.get(product.id, 0.0)
            end_qty = Prd.with_context(to_date=self.end_datetime).browse(product.id).qty_available or 0.0
            adj_qty = start_in_qty - out_qty - end_qty
            table_data.append({'index': index, 'name_use': medicine.name_use, 'uom': product.uom_id.name,
                               'start_qty': str(start_qty), 'in_qty': str(in_qty), 'start_in_qty': str(start_in_qty),
                               'out_qty': str(out_qty), 'adj_qty': str(adj_qty), 'end_qty': str(end_qty), 'name_medicine': medicine.name_medicine})
            index += 1
        datas['index'] = table_data

        # date and everything else
        start_date_str = self.start_date.strftime('%d/%m') if self.start_date.year == self.end_date.year else self.start_date.strftime('%d/%m năm %Y')
        end_date_str = self.end_date.strftime('%d/%m năm %Y')
        date_str = ' đến '.join([start_date_str, end_date_str])

        start_date = (self.start_date - timedelta(days=1)).strftime('%d/%m/%Y')
        month = self.start_date.strftime('%m/%Y')
        end_date = self.end_date.strftime('%d/%m/%Y')

        num_item = str(len(addictive_medicines)).zfill(2) + ' khoản '
        num_item += '(' + num2words_vnm(len(addictive_medicines)).capitalize() + ' khoản)'
        sign_date = fields.Date.today().strftime('Ngày %d tháng %m năm %Y')

        datas.update(comp_name=self.institution.his_company.name, date_str=date_str, start_date=start_date, month=month, end_date=end_date, num_item=num_item, sign_date=sign_date)

        addictive_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.addictive_report_attachment').id)
        decode = base64.b64decode(addictive_attachment.datas)
        doc = MailMerge(BytesIO(decode))
        doc.merge(**datas)
        fp = BytesIO()
        doc.write(fp)
        doc.close()
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()

        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Báo cáo chất gây nghiện, hướng thần.docx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo chất gây nghiện, hướng thần',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def chemical_report(self):
        datas = []
        Move = self.env['stock.move']
        # chemical_categ = self.env['sh.medical.medicines.category'].search([('name', '=', 'Hóa chất')])
        # chemical_supplies = self.env['sh.medical.medicines'].search([('medicine_category_id', '=', chemical_categ.id)])
        chemical_supplies = self.env['sh.medical.medicines'].with_context(active_test=False).search([('default_code', 'like', 'HC')])
        if self.report_type == 'supply_duoc':
            chemical_supplies = self.env['sh.medical.medicines'].with_context(active_test=False).search([('default_code', 'not ilike', 'HC'),('medicament_type', '=', 'Supplies')])
        else:
            chemical_supplies = self.env['sh.medical.medicines'].with_context(active_test=False).search(
                [('default_code', 'like', 'HC')])

        false_and_return_picks = self.env['stock.picking'].search(['&', '&', ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime),
                                                                   '|', ('name', 'ilike', '-FP'), ('origin', 'ilike', '-FP')])
        moves_domain = [('product_id', 'in', chemical_supplies.mapped('product_id.id')), ('picking_id', 'not in', false_and_return_picks.ids),
                        ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime), ('state', '=', 'done'), ('company_id', '=', self.institution.his_company.id)]
        main_location = self.institution.warehouse_ids[0].lot_stock_id
        main_location_supply = self.env['stock.location'].search([('name', 'ilike', 'Tủ Vật tư'), ('company_id', '=', self.institution.his_company.id), ('location_id', '=', main_location.id)], limit=1)
        scrap_location = self.env['stock.location'].search([('name', 'ilike', 'Scrap'), ('company_id', '=', self.institution.his_company.id)], limit=1)
        # room_use_location = self.env['stock.location'].search([('name', 'ilike', 'Sử dụng phòng'), ('company_id', '=', self.institution.his_company.id)])
        lab_ward = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Laboratory')], limit=1)
        lab_test_supply_location = self.env['stock.location'].search([('location_id', '=', lab_ward.location_id.id),
                                                                      ('location_institution_type', '=', self.report_type),
                                                                      ('company_id', '=', self.institution.his_company.id)])

        other_wards = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', 'in', ['Spa', 'Laser', 'Odontology', 'Surgery'])])
        other_location = other_wards.mapped('location_id')
        scrap_domain = moves_domain + [('location_id', 'child_of', main_location.id), ('location_dest_id', '=', scrap_location.id)]
        room_use_domain = moves_domain + [('location_id', 'child_of', main_location.id), '!', ('location_dest_id', 'child_of', main_location.id)]
        scrap_res = dict((str(item['product_id'][0]), item['product_qty'])
                         for item in Move.read_group(scrap_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
        room_use_res = dict((str(item['product_id'][0]) + '-' + str(item['location_id'][0]), item['product_qty'])
                            for item in Move.read_group(room_use_domain, ['product_id', 'product_qty'], ['product_id', 'location_id'], lazy=False, orderby='id'))
        index = 1
        lab_test_num = 0
        other_num = 0
        main_num = 0
        scrap_num = 0

        for supply in chemical_supplies:
            product = supply.product_id
            supply_data = [index, product.default_code, supply.name_use, supply.uom_id.name, supply.standard_price]
            lab_test_qty = room_use_res.get(str(product.id) + '-' + str(lab_test_supply_location.id), 0.0)
            lab_test_value = lab_test_qty * supply.standard_price
            if lab_test_qty > 0:
                lab_test_num += 1

            other_location_qty = 0.0
            for location in other_location:
                other_location_qty += room_use_res.get(str(product.id) + '-' + str(location.id), 0.0)
            other_location_value = other_location_qty * supply.standard_price
            if other_location_qty > 0:
                other_num += 1

            main_location_qty = room_use_res.get(str(product.id) + '-' + str(main_location_supply.id), 0.0)
            main_location_value = main_location_qty * supply.standard_price
            if main_location_qty > 0:
                main_num += 1

            scrap_qty = scrap_res.get(str(product.id), 0.0)
            scrap_value = scrap_qty * supply.standard_price
            if scrap_qty > 0:
                scrap_num += 1

            total_qty = sum((lab_test_qty, other_location_qty, main_location_qty, scrap_qty))
            total_value = sum((lab_test_value, other_location_value, main_location_value, scrap_value))
            supply_data += [other_location_qty, other_location_value, lab_test_qty, lab_test_value,
                            main_location_qty, main_location_value, scrap_qty, scrap_value, total_qty, total_value]
            datas.append(supply_data)
            index += 1

        chemical_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.chemical_report_attachment').id)
        decode = base64.b64decode(chemical_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a2'] = self.institution.name
        ws['d2'].value = 'BÁO CÁO SỬ DỤNG HÓA CHẤT' if self.report_type == 'chemical' else 'BÁO CÁO SỬ DỤNG VẬT TƯ DƯỢC'
        ws['d3'].value = '(Từ ngày %s đến ngày %s)' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))

        row = 8
        for line in datas:
            for col, value in enumerate(line, 1):
                cell = ws.cell(row, col)
                cell.value = value
                if row > 8:
                    cell.border = thick_dotted_top_bot
                    cell.font = Font(name='Times New Roman', size=11)
                    if col in (1, 2, 4):
                        cell.alignment = center_alm
                    if col > 4:
                        cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            row += 1

        ws.cell(row, 1).border = ws.cell(row, 2).border = borders.Border(left=thick, right=thick, top=dotted, bottom=thick)
        num_list = [other_num, lab_test_num, main_num, scrap_num]
        for col in range(3, 16):
            cell = ws.cell(row, col)
            cell.border = all_border_thick
            if col == 3:
                cell.value = 'Cộng khoản'
                cell.font = Font(name='Times New Roman', size=11, italic=True)
            for sub_col, num in zip(list(range(6, 13, 2)), num_list):
                if col == sub_col:
                    cell.value = str(num) + ' khoản' if num > 0 else ''
                    cell.font = Font(name='Times New Roman', size=11, italic=True, bold=True)
        for col in list(range(6, 13, 2)):
            ws.merge_cells(start_row=row, end_row=row, start_column=col, end_column=col+1)

        row += 2
        ws.cell(row, 11).value = date.today().strftime('Ngày %d tháng %m năm %Y')
        ws.cell(row, 11).font = Font(name='Times New Roman', size=11, italic=True)

        row += 1
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=3)
        ws.cell(row, 1).value = 'NGƯỜI LẬP BÁO CÁO'
        ws.cell(row, 1).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 1).alignment = center_alm

        ws.merge_cells(start_row=row, end_row=row, start_column=4, end_column=7)
        ws.cell(row, 4).value = 'TP TÀI CHÍNH - KẾ TOÁN'
        ws.cell(row, 4).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 4).alignment = center_alm

        ws.merge_cells(start_row=row, end_row=row, start_column=8, end_column=11)
        ws.cell(row, 8).value = 'TRƯỞNG KHOA DƯỢC'
        ws.cell(row, 8).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 8).alignment = center_alm

        ws.merge_cells(start_row=row, end_row=row, start_column=12, end_column=15)
        ws.cell(row, 12).value = 'GIÁM ĐỐC'
        ws.cell(row, 12).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 12).alignment = center_alm

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Báo cáo hóa chất.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'BÁO CÁO SỬ DỤNG HÓA CHẤT' if self.report_type == 'chemical' else 'BÁO CÁO SỬ DỤNG VẬT TƯ DƯỢC',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def daily_stock_out_report(self):
        return {'name': 'Báo cáo sử dụng thuốc nội viện',
                'type': 'ir.actions.act_window',
                'res_model': 'stock.move',
                'view_mode': 'tree',
                'view_id': self.env.ref('stock.view_move_tree').id,
                'domain': [('state', '=', 'done'), ('location_dest_id', 'child_of', self.env.ref('stock.stock_location_customers').id), ('company_id', '=', self.institution.his_company.id),
                           ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime), '!', ('reference', 'ilike', '-FP')],
                'context': {'search_default_by_product': 1, 'search_default_groupby_date_2': 1, 'tz': self.env.user.tz}}

    def medicine_out_report(self):
        datas = []
        Move = self.env['stock.move']
        main_location = self.institution.warehouse_ids[0].lot_stock_id
        customer_location = self.env.ref('stock.stock_location_customers')
        scrap_location = self.env['stock.location'].search([('name', 'ilike', 'Scrap'), ('company_id', '=', self.institution.his_company.id)])
        prescription_location = self.institution.location_medicine_stock
        medicines = self.env['sh.medical.medicines'].with_context(active_test=False).search([('medicament_type','=','Medicine')])
        medicine_ids = medicines.ids
        product_ids = medicines.mapped('product_id.id')
        false_and_return_picks = self.env['stock.picking'].search(['&', '&', ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime),
                                                                   '|', ('name', 'ilike', '-FP'), ('origin', 'ilike', '-FP')])
        moves_domain = [('product_id', 'in', product_ids), ('state', '=', 'done'), ('picking_id', 'not in', false_and_return_picks.ids),
                        ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime), ('company_id', '=', self.institution.his_company.id)]
        all_patient_domain = moves_domain + [('location_dest_id', 'child_of', customer_location.id)]
        outpatient_domain = all_patient_domain + [('location_id', '=', prescription_location.id)]
        scrap_domain = moves_domain + [('location_id', 'child_of', main_location.id), ('location_dest_id', '=', scrap_location.id)]
        all_patient_res = dict((str(item['product_id'][0]), item['product_qty'])
                               for item in Move.read_group(all_patient_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
        outpatient_res = dict((str(item['product_id'][0]), item['product_qty'])
                              for item in Move.read_group(outpatient_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
        scrap_res = dict((str(item['product_id'][0]), item['product_qty'])
                         for item in Move.read_group(scrap_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))

        index = 1
        sum_dict = {7: 0, 9: 0, 13: 0, 15: 0}
        for product_id in product_ids:
            all_patient_qty = all_patient_res.get(str(product_id), 0.0)
            outpatient_qty = outpatient_res.get(str(product_id), 0.0)
            scrap_qty = scrap_res.get(str(product_id), 0.0)
            if sum((all_patient_qty, outpatient_qty, scrap_qty)) == 0:
                continue
            product = self.env['product.product'].browse(product_id)
            code = product.default_code
            name = product.name
            uom = product.uom_id.name
            price = product.standard_price
            inpatient_qty = all_patient_qty - outpatient_qty
            inpatient_value = inpatient_qty * price
            outpatient_value = outpatient_qty * price
            scrap_value = scrap_qty * price
            total_qty = inpatient_qty + outpatient_qty + scrap_qty
            total_value = inpatient_value + outpatient_value + scrap_value
            product_data = [index, code, name, uom, price, inpatient_qty, inpatient_value,
                            outpatient_qty, outpatient_value, 0.0, 0.0, scrap_qty, scrap_value, total_qty, total_value]
            sum_dict[7] += inpatient_value
            sum_dict[9] += outpatient_value
            sum_dict[13] += scrap_value
            sum_dict[15] += total_value
            datas.append(product_data)
            index += 1

        # print(datas)
        medicine_out_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.medicine_out_attachment').id)
        decode = base64.b64decode(medicine_out_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a2'].value = self.institution.name
        ws['d3'].value = '(Từ ngày %s đến ngày %s)' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))

        row = 8
        for line in datas:
            for col, value in enumerate(line, 1):
                cell = ws.cell(row, col)
                cell.value = value
                if row > 1:
                    cell.border = thick_dotted_top_bot
                    cell.font = Font(name='Times New Roman', size=11)
                    if col in (1, 2, 4):
                        cell.alignment = center_alm
                    if col > 4:
                        cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            row += 1

        for col in range(1, 16):
            cell = ws.cell(row, col)
            cell.border = borders.Border(left=thick, right=thick, top=dotted, bottom=thick)
            if col in (7, 9, 13, 15):
                cell.value = sum_dict[col]

        row += 2
        ws.cell(row, 3).value = 'Cộng khoản: %s khoản' % len(datas)
        ws.cell(row, 11).value = date.today().strftime('Ngày %d tháng %m năm %Y')
        ws.cell(row, 3).font = ws.cell(row, 11).font = Font(name='Times New Roman', size=11, italic=True)

        row += 1
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=3)
        ws.cell(row, 1).value = 'NGƯỜI LẬP BÁO CÁO'
        ws.cell(row, 1).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 1).alignment = center_alm

        ws.merge_cells(start_row=row, end_row=row, start_column=4, end_column=7)
        ws.cell(row, 4).value = 'TP TÀI CHÍNH - KẾ TOÁN'
        ws.cell(row, 4).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 4).alignment = center_alm

        ws.merge_cells(start_row=row, end_row=row, start_column=8, end_column=11)
        ws.cell(row, 8).value = 'TRƯỞNG KHOA DƯỢC'
        ws.cell(row, 8).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 8).alignment = center_alm

        ws.merge_cells(start_row=row, end_row=row, start_column=12, end_column=15)
        ws.cell(row, 12).value = 'GIÁM ĐỐC BỆNH VIỆN'
        ws.cell(row, 12).font = Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row, 12).alignment = center_alm

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Báo cáo thuốc tiêu hao.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo thuốc tiêu hao',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def simple_inventory_report(self):
        datas = []
        data_sum = ['Tổng', '', '', '', 0, 0, 0, 0, 0, 0, 0, 0]
        Prd = self.env['product.product']
        Move = self.env['stock.move']
        Quant = self.env['stock.quant']
        main_location = self.institution.warehouse_ids[0].lot_stock_id
        supply_location = self.env.ref('stock.stock_location_suppliers')
        customer_location = self.env.ref('stock.stock_location_customers')
        false_and_return_picks = self.env['stock.picking'].search(['&', '&', ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime),
                                                                   '|', ('name', 'ilike', '-FP'), ('origin', 'ilike', '-FP')])
        moves_in_domain = [('date', '>=', self.start_datetime), ('state', '=', 'done'), ('picking_id', 'not in', false_and_return_picks.ids),
                           ('company_id', '=', self.institution.his_company.id), ('location_dest_id', 'child_of', main_location.id), '!', ('location_id', 'child_of', main_location.id)]
        moves_out_domain = [('date', '>=', self.start_datetime), ('state', '=', 'done'), ('picking_id', 'not in', false_and_return_picks.ids),
                            ('company_id', '=', self.institution.his_company.id), ('location_id', 'child_of', main_location.id), '!', ('location_dest_id', 'child_of', main_location.id)]
        categs = self.env['sh.medical.medicines.category'].search([], order='type')
        index = 1
        for categ in categs:
            categ_datas = [['']*12, ['Nhóm hàng: %s' % categ.name, '', '', '', 0, 0, 0, 0, 0, 0, 0, 0]]
            meds = self.env['sh.medical.medicines'].with_context(active_test=False).search([('medicine_category_id', '=', categ.id)])
            product_ids = meds.mapped('product_id.id')
            c_moves_in_domain = moves_in_domain + [('date', '<=', self.end_datetime), ('product_id', 'in', product_ids)]
            c_moves_out_domain = moves_out_domain + [('date', '<=', self.end_datetime), ('product_id', 'in', product_ids)]
            moves_in_res = dict((item['product_id'][0], item['product_qty']) for item in
                                Move.read_group(c_moves_in_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
            moves_out_res = dict((item['product_id'][0], item['product_qty']) for item in
                                 Move.read_group(c_moves_out_domain, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
            quants_res = dict((item['product_id'][0], item['quantity']) for item in
                              Quant.read_group([('product_id', 'in', product_ids), ('location_id', 'child_of', main_location.id)],
                                               ['product_id', 'quantity'], ['product_id'], orderby='id'))
            if self.end_datetime < fields.Datetime.now():
                c_moves_in_domain_past = moves_in_domain + [('date', '>', self.end_datetime), ('product_id', 'in', product_ids)]
                c_moves_out_domain_past = moves_out_domain + [('date', '>', self.end_datetime), ('product_id', 'in', product_ids)]
                moves_in_res_past = dict((item['product_id'][0], item['product_qty']) for item in
                                         Move.read_group(c_moves_in_domain_past, ['product_id', 'product_qty'], ['product_id'], orderby='id'))
                moves_out_res_past = dict((item['product_id'][0], item['product_qty']) for item in
                                          Move.read_group(c_moves_out_domain_past, ['product_id', 'product_qty'], ['product_id'], orderby='id'))

            for med in meds:
                in_qty = moves_in_res.get(med.product_id.id, 0.0)
                out_qty = moves_out_res.get(med.product_id.id, 0.0)
                end_qty = quants_res.get(med.product_id.id, 0.0)
                if self.end_datetime < fields.Datetime.now():
                    end_qty += (moves_out_res_past.get(med.product_id.id, 0.0) - moves_in_res_past.get(med.product_id.id, 0.0))
                if end_qty == in_qty == out_qty == 0:
                    continue
                med_data = [index, med.default_code, med.name, med.uom_id.name]
                med_price = med.standard_price
                begin_qty = end_qty - in_qty + out_qty

                begin_value = begin_qty * med_price
                in_value = in_qty * med_price
                out_value = out_qty * med_price
                end_value = end_qty * med_price
                inv_data = [begin_qty, begin_value, in_qty, in_value, out_qty, out_value, end_qty, end_value]
                med_data += inv_data
                for i, j in zip(list(range(4, 12)), inv_data):
                    categ_datas[1][i] += j
                    data_sum[i] += j
                categ_datas.append(med_data)
                index += 1
            datas += categ_datas
        datas.append(data_sum)

        simple_inventory_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.simple_inventory_attachment').id)
        decode = base64.b64decode(simple_inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a1'].value = 'Tên công ty: %s' % self.institution.his_company.name
        ws['a2'].value = 'Địa chỉ: %s' % ', '.join([a for a in (self.institution.street, self.institution.street2, self.institution.city, self.institution.state_id.name) if a])
        ws['a5'].value = '(Từ ngày %s đến ngày %s)' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))

        row = 11
        for line in datas:
            # print(line)
            for col, val in enumerate(line, 1):
                cell = ws.cell(row, col)
                cell.value = val
                cell.border = all_border_thin
                if isinstance(line[0], str):
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=10)
                if col > 4:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            row += 1

        row += 1
        ws.cell(row, 9).value = date.today().strftime('Ngày %d tháng %m năm %Y')
        ws.cell(row, 9).font = Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row, 9).alignment = center_alm

        row += 2
        ws.cell(row, 3).value = 'Kế toán'
        ws.cell(row, 9).value = 'Người lập'
        ws.cell(row+1, 3).value = ws.cell(row+1, 9).value = '(Ký, họ tên)'
        ws.cell(row, 3).font = ws.cell(row, 9).font = Font(name='Times New Roman', size=10, bold=True)
        ws.cell(row+1, 3).font = ws.cell(row+1, 9).font = Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row, 3).alignment = ws.cell(row, 9).alignment = ws.cell(row+1, 3).alignment = ws.cell(row+1, 9).alignment = center_alm

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Báo cáo XNT Dược.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo XNT Dược',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def out_dated_report(self):
        datas = {}
        # Todo: write search function for field product_qty if performance is slow
        out_dated_lots = self.env['stock.production.lot'].search([('removal_date', '>=', self.start_datetime), ('removal_date', '<=', self.end_datetime), ('company_id', '=', self.institution.his_company.id)])
        table_data = []
        index = 1
        total = 0
        for lot in out_dated_lots:
            if lot.product_qty == 0:
                continue
            code = lot.product_id.default_code
            name = lot.product_id.name
            uom = lot.product_uom_id.name
            qty = lot.product_qty
            lot_name = lot.name
            exp_date = lot.removal_date.strftime('%d.%m.%Y') if lot.removal_date else ''
            price = lot.lot_price / lot.product_init_qty if lot.product_init_qty else 0
            value = qty * price
            total += value
            table_data.append({'index': index, 'code': code, 'name': name, 'uom': uom, 'qty': qty, 'lot_name': lot_name,
                               'exp_date': exp_date, 'price': format(int(price), ',d').replace(',', '.'), 'value': format(int(value), ',d').replace(',', '.')})
            index += 1

        address = ', '.join([a for a in (self.institution.street, self.institution.street2, self.institution.city, self.institution.state_id.name) if a])
        total_w = num2words_vnm(int(total))
        c_date = fields.Date.today().strftime('ngày %d/%m/%Y')
        start_date_str = self.start_date.strftime('%d/%m') if self.start_date.year == self.end_date.year else self.start_date.strftime('%d/%m năm %Y')
        end_date_str = self.end_date.strftime('%d/%m năm %Y')
        date_str = 'Từ ' + ' đến '.join([start_date_str, end_date_str])

        datas.update({'comp_name': self.institution.his_company.name.upper(), 'addr': address,
                      'date_str': date_str, 'index': table_data, 'total': format(int(total), ',d').replace(',', '.'), 'total_w': total_w, 'c_date': c_date})

        out_date_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('shealth_all_in_one.out_date_report_attachment').id)
        decode = base64.b64decode(out_date_attachment.datas)
        doc = MailMerge(BytesIO(decode))
        doc.merge(**datas)
        fp = BytesIO()
        doc.write(fp)
        doc.close()
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()

        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Báo cáo sản phẩm hết hạn.docx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo sản phẩm hết hạn',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def report(self):
        if self.report_type in ('supply', 'medicine'):
            return self.report_inventory()
        elif self.report_type in ('customer', 'service'):
            return self.report_stock_out()
        elif hasattr(self, self.report_type + '_report'):
            return getattr(self, self.report_type + '_report')()

    # BÁO CÁO ĐÁNH GIÁ CHẤT LƯỢNG DỊCH VỤ
    def _get_quality_evaluation_data(self):
        # get data phòng khám da liễu
        walkin_spa_data = self.env['sh.medical.appointment.register.walkin'].search(
            [('service_room', '=', self.env.ref('shealth_all_in_one.sh_kb_dalieu_room_knhn').id), ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime)], order='date')

        table_spa_data = [len(walkin_spa_data)]
        table_spa_item = {}
        for walkin_spa in walkin_spa_data:
            patient = walkin_spa.patient.display_name
            wait_reception = random.randint(1, 9)
            process_reception = random.randint(1, 9)
            wait_payment = random.randint(1, 9)
            process_payment = random.randint(1, 5)
            wait_exam = random.randint(1, 9)
            process_exam = random.randint(5, 9)
            total_exam = wait_exam + process_exam

            duration_service = walkin_spa.specialty_ids[0].services_length * 60 if len(walkin_spa.specialty_ids) > 0 else ''
            start_service = end_service = ''
            if duration_service != '':
                start_service = walkin_spa.specialty_ids[0].services_date.strftime('%d/%m/%Y %H:%M') if len(walkin_spa.specialty_ids) > 0 else ''
                end_service = walkin_spa.specialty_ids[0].services_end_date.strftime('%d/%m/%Y %H:%M') if len(walkin_spa.specialty_ids) > 0 else ''

            # khác ngày
            if str(walkin_spa.date.strftime('%d/%m/%Y')) not in table_spa_item:
                table_spa_item[str(walkin_spa.date.strftime('%d/%m/%Y'))] = [{'patient':patient,
                                                                              'wait_reception': wait_reception,'process_reception': process_reception,
                                                                              'wait_payment': wait_payment, 'process_payment': process_payment,
                                                                              'wait_exam': wait_exam, 'process_exam': process_exam,'total_exam': total_exam,
                                                                              'start_service': start_service,'end_service': end_service,
                                                                              'duration_service': duration_service}]
            else:
                table_spa_item[str(walkin_spa.date.strftime('%d/%m/%Y'))].append({'patient':patient,
                                                                              'wait_reception': wait_reception,'process_reception': process_reception,
                                                                              'wait_payment': wait_payment, 'process_payment': process_payment,
                                                                              'wait_exam': wait_exam, 'process_exam': process_exam,'total_exam': total_exam,
                                                                              'start_service': start_service,'end_service': end_service,
                                                                              'duration_service': duration_service})
        table_spa_data.append(table_spa_item)

        # get data phòng khám nha khoa
        walkin_rhm_data = self.env['sh.medical.appointment.register.walkin'].search(
            [('service_room', '=', self.env.ref('shealth_all_in_one.sh_kb_rhm_room_knhn').id),
             ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime)], order='date')

        table_rhm_data = [len(walkin_rhm_data)]
        table_rhm_item = {}
        for walkin_rhm in walkin_rhm_data:
            patient = walkin_rhm.patient.display_name
            wait_reception = random.randint(1, 9)
            process_reception = random.randint(1, 9)
            wait_payment = random.randint(1, 9)
            process_payment = random.randint(1, 5)
            wait_exam = random.randint(1, 9)
            process_exam = random.randint(5, 9)
            total_exam = wait_exam + process_exam

            duration_service = walkin_rhm.specialty_ids[0].services_length * 60 if len(
                walkin_rhm.specialty_ids) > 0 else ''
            start_service = end_service = ''
            if duration_service != '':
                start_service = walkin_rhm.specialty_ids[0].services_date.strftime('%d/%m/%Y %H:%M') if len(
                    walkin_rhm.specialty_ids) > 0 else ''
                end_service = walkin_rhm.specialty_ids[0].services_end_date.strftime('%d/%m/%Y %H:%M') if len(
                    walkin_rhm.specialty_ids) > 0 else ''

            # khác ngày
            if str(walkin_rhm.date.strftime('%d/%m/%Y')) not in table_rhm_item:
                table_rhm_item[str(walkin_rhm.date.strftime('%d/%m/%Y'))] = [{'patient': patient,
                                                                              'wait_reception': wait_reception,
                                                                              'process_reception': process_reception,
                                                                              'wait_payment': wait_payment,
                                                                              'process_payment': process_payment,
                                                                              'wait_exam': wait_exam,
                                                                              'process_exam': process_exam,
                                                                              'total_exam': total_exam,
                                                                              'start_service': start_service,
                                                                              'end_service': end_service,
                                                                              'duration_service': duration_service}]
            else:
                table_rhm_item[str(walkin_rhm.date.strftime('%d/%m/%Y'))].append({'patient': patient,
                                                                                  'wait_reception': wait_reception,
                                                                                  'process_reception': process_reception,
                                                                                  'wait_payment': wait_payment,
                                                                                  'process_payment': process_payment,
                                                                                  'wait_exam': wait_exam,
                                                                                  'process_exam': process_exam,
                                                                                  'total_exam': total_exam,
                                                                                  'start_service': start_service,
                                                                                  'end_service': end_service,
                                                                                  'duration_service': duration_service})

        table_rhm_data.append(table_rhm_item)

        # get data phòng khám pttm
        walkin_pttm_data = self.env['sh.medical.appointment.register.walkin'].search(
            [('service_room', '=', self.env.ref('shealth_all_in_one.sh_kb_pttt_room_knhn').id),
             ('date', '>=', self.start_datetime), ('date', '<=', self.end_datetime)], order='date')

        # print(self.start_datetime)
        # print(self.end_datetime)
        # print('===')

        table_pttm_data = [len(walkin_pttm_data)]
        table_pttm_item = {}
        for walkin_pttm in walkin_pttm_data:
            patient = walkin_pttm.patient.display_name
            wait_reception = random.randint(1, 9)
            process_reception = random.randint(1, 9)
            wait_payment = random.randint(1, 9)
            process_payment = random.randint(1, 5)
            wait_exam = random.randint(1, 14)
            process_exam = random.randint(5, 14)
            total_exam = wait_exam + process_exam

            # thực hiện dịch vụ
            duration_service = int(walkin_pttm.surgeries_ids[0].surgery_length * 60) if len(
                walkin_pttm.surgeries_ids) > 0 else ''
            start_service = end_service = ''
            if duration_service != '':
                start_service = walkin_pttm.surgeries_ids[0].surgery_date.strftime('%d/%m/%Y %H:%M') if len(
                    walkin_pttm.surgeries_ids) > 0 else ''
                end_service = walkin_pttm.surgeries_ids[0].surgery_end_date.strftime('%d/%m/%Y %H:%M') if len(
                    walkin_pttm.surgeries_ids) > 0 else ''

            # thực hiện xét nghiệm
            if walkin_pttm.lab_test_ids:
                lab_date_requested = 1.0 * calendar.timegm(
                    tm.strptime(walkin_pttm.lab_test_ids[0].date_requested.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S"))
                lab_date_analysis = 1.0 * calendar.timegm(
                    tm.strptime(walkin_pttm.lab_test_ids[0].date_analysis.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S"))
                lab_date_done = 1.0 * calendar.timegm(
                    tm.strptime(walkin_pttm.lab_test_ids[0].date_done.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S"))

                wait_labtest = (lab_date_analysis - lab_date_requested) / 60
                process_labtest = random.randint(5, 7)
                wait_result_labtest = (lab_date_done - lab_date_analysis) / 60
                total_labtest = wait_labtest + process_labtest + wait_result_labtest
            else:
                wait_labtest = process_labtest = total_labtest = wait_result_labtest = ''

            # tổng thời gian khám và xét nghiệm
            total_labtest_exam = int(0 if total_labtest == '' else total_labtest) + int(0 if total_exam == '' else total_exam)

            # thực hiện CĐHA
            if walkin_pttm.imaging_ids:
                img_date_requested = 1.0 * calendar.timegm(
                    tm.strptime(walkin_pttm.imaging_ids[0].date_requested.strftime("%Y-%m-%d %H:%M:%S"),
                                  "%Y-%m-%d %H:%M:%S"))
                img_date_analysis = 1.0 * calendar.timegm(
                    tm.strptime(walkin_pttm.imaging_ids[0].date_analysis.strftime("%Y-%m-%d %H:%M:%S"),
                                  "%Y-%m-%d %H:%M:%S"))
                img_date_done = 1.0 * calendar.timegm(
                    tm.strptime(walkin_pttm.imaging_ids[0].date_done.strftime("%Y-%m-%d %H:%M:%S"),
                                  "%Y-%m-%d %H:%M:%S"))

                wait_imaging = (img_date_analysis - img_date_requested) / 60
                process_imaging = random.randint(5, 6)
                wait_result_imaging = (img_date_done - img_date_analysis) / 60
                total_imaging = wait_imaging + process_imaging + wait_result_imaging
            else:
                total_imaging = wait_imaging = process_imaging = wait_result_imaging = ''

            # tổng thời gian khám, xét nghiệm và cđha
            total_cls_exam = int(total_labtest_exam) + int(0 if total_imaging == '' else total_imaging)

            wait_medicine = random.randint(4, 10)

            # khác ngày
            if str(walkin_pttm.date.strftime('%d/%m/%Y')) not in table_pttm_item:
                table_pttm_item[str(walkin_pttm.date.strftime('%d/%m/%Y'))] = [{'patient': patient,
                                                                              'wait_reception': wait_reception,
                                                                              'process_reception': process_reception,
                                                                              'wait_payment': wait_payment,
                                                                              'process_payment': process_payment,
                                                                              'wait_exam': wait_exam,
                                                                              'process_exam': process_exam,
                                                                              'total_exam': total_exam,
                                                                              'wait_labtest': wait_labtest,
                                                                              'process_labtest': process_labtest,
                                                                              'wait_result_labtest': wait_result_labtest,
                                                                              'wait_imaging': wait_imaging,
                                                                              'process_imaging': process_imaging,
                                                                              'wait_result_imaging': wait_result_imaging,
                                                                              'total_labtest_exam': total_labtest_exam,
                                                                              'total_cls_exam': total_cls_exam,
                                                                              'start_service': start_service,
                                                                              'end_service': end_service,
                                                                              'duration_service': duration_service,
                                                                              'wait_medicine': wait_medicine}]
            else:
                table_pttm_item[str(walkin_pttm.date.strftime('%d/%m/%Y'))].append({'patient': patient,
                                                                              'wait_reception': wait_reception,
                                                                              'process_reception': process_reception,
                                                                              'wait_payment': wait_payment,
                                                                              'process_payment': process_payment,
                                                                              'wait_exam': wait_exam,
                                                                              'process_exam': process_exam,
                                                                              'total_exam': total_exam,
                                                                              'wait_labtest': wait_labtest,
                                                                              'process_labtest': process_labtest,
                                                                              'wait_result_labtest': wait_result_labtest,
                                                                              'wait_imaging': wait_imaging,
                                                                              'process_imaging': process_imaging,
                                                                              'wait_result_imaging': wait_result_imaging,
                                                                              'total_labtest_exam': total_labtest_exam,
                                                                              'total_cls_exam': total_cls_exam,
                                                                              'start_service': start_service,
                                                                              'end_service': end_service,
                                                                              'duration_service': duration_service,
                                                                              'wait_medicine': wait_medicine})

        table_pttm_data.append(table_pttm_item)

        return [table_spa_data,table_rhm_data,table_pttm_data]
        # [[dalieu],[nha],[phauthuat],[tonghop]]
        # [[8,{'1/1/2020':[{},{},{}]},'1/1/2020':[{},{},{}]}],[8,{'1/1/2020':[{},{},{}]},'1/1/2020':[{},{},{}]}],[8,{'1/1/2020':[{},{},{}]},'1/1/2020':[{},{},{}]}]]

    def quality_evaluation_report(self):
        simple_quality_evaluation_attachment = self.env['ir.attachment'].sudo().browse(
            self.env.ref('shealth_all_in_one.quality_evaluation_report_attachment').id)
        decode = base64.b64decode(simple_quality_evaluation_attachment.datas)
        wb = load_workbook(BytesIO(decode))

        datas = self._get_quality_evaluation_data()
        line_font = Font(name='Times New Roman', size=12)
        line_avg_font = Font(name='Times New Roman', size=12, bold=True, color='1279E7')
        line_min_font = Font(name='Times New Roman', size=12, bold=True, color='41AC4D')
        line_max_font = Font(name='Times New Roman', size=12, bold=True, color='E7002A')
        key_col_list = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'L']
        key_col_pt_list = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'T','U']

        #============= TAB DA LIỄU ====================
        spa_ws = wb['Da liễu']
        spa_ws['A2'].value = 'Thời gian: Từ ngày %s đến ngày %s' % (
            self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        spa_ws['A3'].value = 'Tổng số: %d phiếu' % (datas[0][0])

        row = 6
        start_merge = 6
        for key in sorted(datas[0][1].keys()):
            spa_ws.cell(row, 1).value = key
            spa_ws.cell(row, 1).font, spa_ws.cell(row, 1).border, spa_ws.cell(row, 1).alignment = line_font, all_border_thin, \
                                                                                                  Alignment(horizontal='center', vertical='center')
            key_list = ['patient','wait_reception','process_reception','wait_payment','process_payment','wait_exam','process_exam','total_exam','start_service','end_service','duration_service']
            for line_spa_data in datas[0][1][key]:
                for col, k in enumerate(key_list,2):
                    cell = spa_ws.cell(row, col)
                    cell.value, cell.font, cell.border = line_spa_data[k], line_font, all_border_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center') if col > 2 else Alignment(horizontal='left', vertical='center')

                row += 1

            spa_ws.merge_cells(start_row=start_merge, start_column=1, end_row=row-1, end_column=1)
            start_merge = row

        # TRUNG BÌNH, MIN, MAX
        spa_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        spa_ws.cell(row, 1).value = "NHỎ NHẤT"

        spa_ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=2)
        spa_ws.cell(row+1, 1).value = "TRUNG BÌNH"

        spa_ws.merge_cells(start_row=row+2, start_column=1, end_row=row+2, end_column=2)
        spa_ws.cell(row+2, 1).value = "LỚN NHẤT"

        for i in range(1, 13):
            spa_ws.cell(row, i).border, spa_ws.cell(row, i).alignment = all_border_thin,Alignment(horizontal='center', vertical='center')
            spa_ws.cell(row+1, i).border, spa_ws.cell(row+1, i).alignment = all_border_thin,Alignment(horizontal='center', vertical='center')
            spa_ws.cell(row+2, i).border, spa_ws.cell(row+2, i).alignment = all_border_thin,Alignment(horizontal='center', vertical='center')

            spa_ws.cell(row, i).font = line_min_font
            spa_ws.cell(row+1, i).font = line_avg_font
            spa_ws.cell(row+2, i).font = line_max_font

        for col in key_col_list:
            spa_ws[col + str(start_merge)].value = '=MIN(%s6:%s%s)'% (col, col, str(start_merge-1))
            spa_ws[col + str(start_merge+1)].value = '=IF(SUM(%s6:%s%s)=0,0,AVERAGE(%s6:%s%s))'% (col, col, str(start_merge-1),col, col, str(start_merge-1))
            spa_ws[col + str(start_merge+2)].value = '=MAX(%s6:%s%s)'% (col, col, str(start_merge-1))

            spa_ws[col + str(start_merge+1)].number_format = '0.0'

        #============= TAB NHA KHOA ====================
        rhm_ws = wb['Nha khoa']
        rhm_ws['A2'].value = 'Thời gian: Từ ngày %s đến ngày %s' % (
            self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        rhm_ws['A3'].value = 'Tổng số: %d phiếu' % (datas[1][0])

        row = 6
        start_merge = 6
        for key in sorted(datas[1][1].keys()):
            rhm_ws.cell(row, 1).value = key
            rhm_ws.cell(row, 1).font, rhm_ws.cell(row, 1).border, rhm_ws.cell(row,1).alignment = line_font, all_border_thin, \
                                                                                             Alignment(
                                                                                                 horizontal='center',
                                                                                                 vertical='center')
            key_list = ['patient', 'wait_reception', 'process_reception', 'wait_payment', 'process_payment',
                        'wait_exam', 'process_exam', 'total_exam', 'start_service', 'end_service',
                        'duration_service']
            for line_rhm_data in datas[1][1][key]:
                for col, k in enumerate(key_list, 2):
                    cell = rhm_ws.cell(row, col)
                    cell.value, cell.font, cell.border = line_rhm_data[k], line_font, all_border_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center') if col > 2 else Alignment(
                        horizontal='left', vertical='center')

                row += 1

            rhm_ws.merge_cells(start_row=start_merge, start_column=1, end_row=row - 1, end_column=1)
            start_merge = row

        # TRUNG BÌNH, MIN, MAX
        rhm_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        rhm_ws.cell(row, 1).value = "NHỎ NHẤT"

        rhm_ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=2)
        rhm_ws.cell(row + 1, 1).value = "TRUNG BÌNH"

        rhm_ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=2)
        rhm_ws.cell(row + 2, 1).value = "LỚN NHẤT"

        for i in range(1, 13):
            rhm_ws.cell(row, i).border, rhm_ws.cell(row, i).alignment = all_border_thin, Alignment(
                horizontal='center', vertical='center')
            rhm_ws.cell(row + 1, i).border, rhm_ws.cell(row + 1, i).alignment = all_border_thin, Alignment(
                horizontal='center', vertical='center')
            rhm_ws.cell(row + 2, i).border, rhm_ws.cell(row + 2, i).alignment = all_border_thin, Alignment(
                horizontal='center', vertical='center')

            rhm_ws.cell(row, i).font = line_min_font
            rhm_ws.cell(row + 1, i).font = line_avg_font
            rhm_ws.cell(row + 2, i).font = line_max_font

        for col in key_col_list:
            rhm_ws[col + str(start_merge)].value = '=MIN(%s6:%s%s)' % (col, col, str(start_merge - 1))
            rhm_ws[col + str(start_merge + 1)].value = '=IF(SUM(%s6:%s%s)=0,0,AVERAGE(%s6:%s%s))' % (
            col, col, str(start_merge - 1), col, col, str(start_merge - 1))
            rhm_ws[col + str(start_merge + 2)].value = '=MAX(%s6:%s%s)' % (col, col, str(start_merge - 1))

            rhm_ws[col + str(start_merge + 1)].number_format = '0.0'

        # ============= TAB PTTM ====================
        pt_ws = wb['Thẩm mỹ']
        pt_ws['A2'].value = 'Thời gian: Từ ngày %s đến ngày %s' % (
            self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        pt_ws['A3'].value = 'Tổng số: %d phiếu' % (datas[2][0])

        row = 6
        start_merge = 6
        for key in sorted(datas[2][1].keys()):
            pt_ws.cell(row, 1).value = key
            pt_ws.cell(row, 1).font, pt_ws.cell(row, 1).border, pt_ws.cell(row,1).alignment = line_font, all_border_thin, \
                                                       Alignment(horizontal='center', vertical='center')
            key_list = ['patient', 'wait_reception', 'process_reception', 'wait_payment', 'process_payment',
                        'wait_exam', 'process_exam', 'total_exam', 'wait_labtest', 'process_labtest',
                        'wait_result_labtest', 'wait_imaging', 'process_imaging', 'wait_result_imaging',
                        'total_labtest_exam', 'total_cls_exam', 'start_service', 'end_service',
                        'duration_service','wait_medicine']
            for line_pt_data in datas[2][1][key]:
                for col, k in enumerate(key_list, 2):
                    cell = pt_ws.cell(row, col)
                    cell.value, cell.font, cell.border = line_pt_data[k], line_font, all_border_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center') if col > 2 else Alignment(
                        horizontal='left', vertical='center')

                row += 1

            pt_ws.merge_cells(start_row=start_merge, start_column=1, end_row=row - 1, end_column=1)
            start_merge = row

        # TRUNG BÌNH, MIN, MAX
        pt_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        pt_ws.cell(row, 1).value = "NHỎ NHẤT"

        pt_ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=2)
        pt_ws.cell(row + 1, 1).value = "TRUNG BÌNH"

        pt_ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=2)
        pt_ws.cell(row + 2, 1).value = "LỚN NHẤT"

        for i in range(1, 22):
            pt_ws.cell(row, i).border, pt_ws.cell(row, i).alignment = all_border_thin, Alignment(
                horizontal='center', vertical='center')
            pt_ws.cell(row + 1, i).border, pt_ws.cell(row + 1, i).alignment = all_border_thin, Alignment(
                horizontal='center', vertical='center')
            pt_ws.cell(row + 2, i).border, pt_ws.cell(row + 2, i).alignment = all_border_thin, Alignment(
                horizontal='center', vertical='center')

            pt_ws.cell(row, i).font = line_min_font
            pt_ws.cell(row + 1, i).font = line_avg_font
            pt_ws.cell(row + 2, i).font = line_max_font

        for col in key_col_pt_list:
            pt_ws[col + str(start_merge)].value = '=MIN(%s6:%s%s)' % (col, col, str(start_merge - 1))
            pt_ws[col + str(start_merge + 1)].value = '=IF(SUM(%s6:%s%s)=0,0,AVERAGE(%s6:%s%s))' % (
                col, col, str(start_merge - 1), col, col, str(start_merge - 1))
            pt_ws[col + str(start_merge + 2)].value = '=MAX(%s6:%s%s)' % (col, col, str(start_merge - 1))

            pt_ws[col + str(start_merge + 1)].number_format = '0.0'

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'DANH_GIA_CHAT_LUONG_DICH_VU.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'ĐÁNH GIÁ CHẤT LƯỢNG DỊCH VỤ',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}














