from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

import logging

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class DailySalesReport(models.TransientModel):
    _name = 'daily.sales.report'
    _description = 'Sales report'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def report_daily_sales(self):
        ret_data = []
        # tổng tiền booking
        total_amount_total_booking_inbound = 0
        total_amount_total_booking_outbound = 0

        # tổng tiền thu (VND)
        total_amount_vnd_inbound = 0
        total_amount_vnd_outbound = 0

        # tổng tiền đã nộp trước
        total_amount_paid_inbound = 0
        total_amount_paid_outbound = 0

        # tổng tiền còn lại
        total_amount_remain_inbound = 0
        total_amount_remain_outbound = 0

        # lấy data từ payment
        domain = [('payment_date', '>=', self.start_datetime), ('payment_date', '<=', self.end_datetime),
                  ('state', '!=', 'draft'), ('state', '!=', 'cancelled'), ('payment_type', '!=', 'internal')]
        list_payment = self.env['account.payment'].search(domain)

        # lấy data đơn vị tiền sử dụng
        list_currency_used = self.env['res.currency'].search([])
        total_amount_foreign_currency_inbound = {cur.name: 0 for cur in list_currency_used}
        total_amount_foreign_currency_outbound = {cur.name: 0 for cur in list_currency_used}

        for rec in list_payment:
            amount_foreign_currency = 0
            amount = 0
            payment_type = None
            # ---------------- ghi nhận kiểu thanh toán & tính tổng tiền
            if rec.payment_type:
                if rec.payment_type == 'outbound':
                    payment_type = 'Hoàn tiền'

                    # tổng tiền booking theo phiếu hoàn tiền
                    if rec.crm_id.amount_total:
                        total_amount_total_booking_outbound += rec.crm_id.amount_total

                    # tổng tiền thu(nguyên tệ) theo phiếu hoàn tiền
                    for i in list_currency_used:
                        if rec.currency_id.name == i.name:
                            total_amount_foreign_currency_outbound[i.name] += rec.amount

                    # tổng tiền đã nộp(theo booking) theo phiếu hoàn tiền
                    if rec.crm_id.amount_paid:
                        total_amount_paid_outbound += rec.crm_id.amount_paid

                    # tổng tiền còn lại(theo booking) theo phiếu hoàn tiền
                    if rec.crm_id.amount_remain:
                        total_amount_remain_outbound += rec.crm_id.amount_remain

                elif rec.payment_type == 'inbound':
                    payment_type = 'Thu tiền'

                    # tổng tiền booking theo phiếu thu
                    if rec.crm_id.amount_total:
                        total_amount_total_booking_inbound += rec.crm_id.amount_total

                    # tổng tiền thu(nguyên tệ) theo phiếu thu tiền
                    for i in list_currency_used:
                        if rec.currency_id.name == i.name:
                            total_amount_foreign_currency_inbound[i.name] += rec.amount

                    # tổng tiền đã nộp(lấy theo booking) theo phiếu thu tiền:
                    if rec.crm_id.amount_paid:
                        total_amount_paid_inbound += rec.crm_id.amount_paid

                    # tổng tiền còn lại (theo booking) theo phiếu thu tiền
                    if rec.crm_id.amount_remain:
                        total_amount_remain_inbound += rec.crm_id.amount_remain

            # ------------------ kiểm tra ngoại tệ
            if rec.currency_rate_id:
                amount_foreign_currency = format(int(rec.amount), ',d').replace(',', '.')
                amount_foreign_currency = str(amount_foreign_currency) + ' ' + rec.currency_id.name
                amount = None
            if not rec.currency_rate_id:
                amount = format(int(rec.amount), ',d').replace(',', '.')
                amount = str(amount) + ' ' + rec.currency_id.name
                amount_foreign_currency = None
                if rec.payment_type == 'outbound':
                    total_amount_vnd_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_vnd_inbound += rec.amount
            # ------------------------- lưu dữ liệu
            ret_data.append({
                'name': rec.name,
                'payment _type': payment_type,
                'communication': rec.communication,
                'code_customer': rec.partner_id.code_customer,
                'name_customer': rec.partner_id.name,
                'address_customer': rec.partner_id.street or '-',
                'code_booking': rec.crm_id.name or '-',
                'amount_total': format(int(rec.crm_id.amount_total), ',d').replace(',', '.') or '-',
                'amount_foreign_currency': amount_foreign_currency or '-',
                'amount': amount or '-',
                'amount_paid': format(int(rec.crm_id.amount_paid), ',d').replace(',', '.') or '-',
                'amount_remain': format(int(rec.crm_id.amount_remain), ',d').replace(',', '.') or '-',
                'company': rec.company_id.name,
                'source': rec.crm_id.source_id.name or '-',
                'user': rec.user.name,
                'note': None,
            })
        # ----------- in dữ liệu
        daily_sales_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.daily_sales_report_attachment').id)
        decode = base64.b64decode(daily_sales_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        if self.start_date == self.end_date:
            ws['A3'].value = self.end_date.strftime('Ngày %d tháng %m năm %Y')
        else:
            ws['A3'].value = 'Từ ngày: %s đến ngày: %s' % (
                self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=12)
        key_col_list = list(range(2, 18))
        key_list = [
            'name',
            'payment _type',
            'communication',
            'code_customer',
            'name_customer',
            'address_customer',
            'code_booking',
            'amount_total',
            'amount_foreign_currency',
            'amount',
            'amount_paid',
            'amount_remain',
            'company',
            'source',
            'user',
            'note',
        ]
        row = 6
        for line_data in ret_data:
            ws.cell(row, 1).value = row - 5
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1).value = 'Tổng tiền'
        ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
        for i in range(1, 9):
            ws.cell(row, i).border = all_border_thin

        # in tổng tiền booking
        ws.cell(row, 9).value = 'Phiếu thu: %s' % format(int(total_amount_total_booking_inbound), ',d').replace(',', '.')
        ws.cell(row, 9).border = all_border_thin
        ws.cell(row + 1, 9).value = 'Phiếu hoàn: %s' % format(int(total_amount_total_booking_outbound), ',d').replace(',', '.')
        ws.cell(row + 1, 9).border = all_border_thin

        # in tổng số tiền thu(nguyên tệ)
        list_test_total_amount_foreign_currency_inbound = [str(key) + ': ' + str(format(int(value), ',d').replace(',', '.')) for key, value in total_amount_foreign_currency_inbound.items()]
        list_test_total_amount_foreign_currency_inbound = ','.join(list_test_total_amount_foreign_currency_inbound)

        list_test_total_amount_foreign_currency_outbound = [str(key) + ': ' + str(format(int(value), ',d').replace(',', '.')) for key, value in total_amount_foreign_currency_outbound.items()]
        list_test_total_amount_foreign_currency_outbound = ','.join(list_test_total_amount_foreign_currency_outbound)

        ws.cell(row, 10).value = 'Phiếu thu: %s' % list_test_total_amount_foreign_currency_inbound
        ws.cell(row, 10).border = all_border_thin
        ws.cell(row + 1, 10).value = 'Phiếu hoàn: %s' % list_test_total_amount_foreign_currency_outbound
        ws.cell(row + 1, 10).border = all_border_thin



        # in tổng số tiền thu(VND)
        ws.cell(row, 11).value = 'Phiếu thu: %s' % format(int(total_amount_vnd_inbound), ',d').replace(',', '.')
        ws.cell(row, 11).border = all_border_thin
        ws.cell(row + 1, 11).value = 'Phiếu hoàn: %s' % format(int(total_amount_vnd_outbound), ',d').replace(',', '.')
        ws.cell(row + 1, 11).border = all_border_thin

        # in tổng tiền đã nộp
        ws.cell(row, 12).value = 'Phiếu thu: %s' % format(int(total_amount_paid_inbound), ',d').replace(',', '.')
        ws.cell(row, 12).border = all_border_thin
        ws.cell(row + 1, 12).value = 'Phiếu hoàn: %s' % format(int(total_amount_paid_outbound), ',d').replace(',', '.')
        ws.cell(row + 1, 12).border = all_border_thin

        # tin tổng tiền còn lại
        ws.cell(row, 13).value = 'Phiếu thu: %s' % format(int(total_amount_remain_inbound), ',d').replace(',', '.')
        ws.cell(row, 13).border = all_border_thin
        ws.cell(row + 1, 13).value = 'Phiếu hoàn: %s' % format(int(total_amount_paid_outbound), ',d').replace(',', '.')
        ws.cell(row + 1, 13).border = all_border_thin

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_so_ngay.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        return {
            'name': 'Báo cáo doanh số ngày',
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'attachment_id': attachment.id}
        }
