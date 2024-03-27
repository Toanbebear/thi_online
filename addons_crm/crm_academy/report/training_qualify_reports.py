import time as tm
import calendar
from pytz import timezone
from ast import literal_eval
from calendar import monthrange
from copy import copy
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from operator import itemgetter
from pytz import timezone, utc
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from odoo.tools import pycompat
import numpy
import base64

import random

import logging

_logger = logging.getLogger(__name__)


class InheritOpBatchLine(models.Model):
    _inherit = 'op.batch.line'

    num_lessons = fields.Integer(related='batch_id.num_lessons', store=True)
    tutor_fee_related = fields.Float(related='batch_id.tutor_fee', store=True)
    total_fee = fields.Float(string='Phí giảng viên/lớp', compute='_get_total_fee', store=True)

    @api.depends('num_lessons', 'tutor_fee_related')
    def _get_total_fee(self):
        for record in self:
            record.total_fee = record.num_lessons * record.tutor_fee_related if record.num_lessons and record.tutor_fee_related else 0


class TrainingQualityReports(models.TransientModel):
    _name = 'training.quality.reports'
    _description = 'Training Quality Reports'

    start_date = fields.Date('Date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    type_report = fields.Selection([('course', 'Course evaluation reports'), ('faculty', 'Teacher evaluation report')],
                                   string='Type report', default='course')

    def training_qualify_report(self):
        if self.type_report == 'course':
            if self.start_date and self.end_date:
                batch_ids = self.env['op.batch'].search(
                    [('internal', '=', True), ('start_date', '>=', self.start_date),
                     ('start_date', '<=', self.end_date)], order='start_date asc')
                rec_data = []
                for batch in batch_ids:
                    survey = self.env['survey.user_input_line'].read_group(
                        [('user_input_id.batch_id', 'in', batch.ids)], ['question_id', 'value_suggested_int'],
                        ['question_id'])
                    ques_0 = round(survey[0]['value_suggested_int'] / survey[0]['question_id_count'],
                                   1) if survey else False
                    ques_1 = round(survey[1]['value_suggested_int'] / survey[1]['question_id_count'],
                                   1) if survey else False
                    ques_2 = round(survey[2]['value_suggested_int'] / survey[2]['question_id_count'],
                                   1) if survey else False
                    ques_3 = round(survey[3]['value_suggested_int'] / survey[3]['question_id_count'],
                                   1) if survey else False
                    ques_4 = round(survey[4]['value_suggested_int'] / survey[4]['question_id_count'],
                                   1) if survey else False
                    ques_5 = round(survey[5]['value_suggested_int'] / survey[5]['question_id_count'],
                                   1) if survey else False
                    ques_6 = round(survey[6]['value_suggested_int'] / survey[6]['question_id_count'],
                                   1) if survey else False
                    ques_7 = round(survey[7]['value_suggested_int'] / survey[7]['question_id_count'],
                                   1) if survey else False
                    ques_8 = round(survey[8]['value_suggested_int'] / survey[8]['question_id_count'],
                                   1) if survey else False
                    ques_9 = round(survey[9]['value_suggested_int'] / survey[9]['question_id_count'],
                                   1) if survey else False
                    if ques_0 and ques_1 and ques_2 and ques_3 and ques_4 and ques_5 and ques_6 and ques_7 and ques_8 and ques_9:
                        score = numpy.mean(
                            [ques_0, ques_1, ques_2, ques_3, ques_4, ques_5, ques_6, ques_7, ques_8, ques_9])
                    else:
                        score = '-'
                    rec_data.append({
                        'faculty_name': batch.faculty_id.name,
                        'institute': batch.institute.name,
                        'faculty_job': batch.faculty_id.emp_id.job_id.name,
                        'faculty_department': batch.faculty_id.emp_id.department_id.name,
                        'batch_name': batch.name,
                        'start_date': batch.start_date.strftime("%d/%m/%Y"),
                        'num_lessons': batch.num_lessons,
                        'fee': int(batch.course_id.tutor_fee),
                        'total_fee': int(batch.num_lessons * batch.course_id.tutor_fee),
                        'ques_0': ques_0 or '-',
                        'ques_1': ques_1 or '-',
                        'ques_2': ques_2 or '-',
                        'ques_3': ques_3 or '-',
                        'ques_4': ques_4 or '-',
                        'ques_5': ques_5 or '-',
                        'ques_6': ques_6 or '-',
                        'ques_7': ques_7 or '-',
                        'ques_8': ques_8 or '-',
                        'ques_9': ques_9 or '-',
                        'score': score
                    })
                simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
                    self.env.ref('crm_academy.course_report_template').id)
                decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
                wb = load_workbook(BytesIO(decode))
                ws = wb.active
                thin = borders.Side(style='thin')
                all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
                line_font = Font(name='Times New Roman', size=12)
                key_col_list = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]
                key_list = [
                    'faculty_name',
                    'institute',
                    'faculty_job',
                    'faculty_department',
                    'batch_name',
                    'start_date',
                    'num_lessons',
                    'fee',
                    'total_fee',
                    'ques_0',
                    'ques_1',
                    'ques_2',
                    'ques_3',
                    'ques_4',
                    'ques_5',
                    'ques_6',
                    'ques_7',
                    'ques_8',
                    'ques_9',
                    'score',
                ]
                row = 7
                for line_data in rec_data:
                    ws.cell(row, 1).value = row - 6
                    for col, k in zip(key_col_list, key_list):
                        cell = ws.cell(row, col)
                        cell.value = line_data[k]
                        cell.font = line_font
                        cell.border = all_border_thin
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                fp = BytesIO()
                wb.save(fp)
                fp.seek(0)
                report = base64.encodebytes((fp.read()))
                fp.close()
                attachment = self.env['ir.attachment'].sudo().create({
                    'name': 'Báo cáo đánh giá khóa học',
                    'datas': report,
                    'res_model': 'temp.creation',
                    'public': True,
                })
                return {
                    'name': 'Báo cáo đánh giá khóa học',
                    'type': 'ir.actions.act_window',
                    'res_model': 'temp.wizard',
                    'view_mode': 'form',
                    'view_type': 'form',
                    'target': 'inline',
                    'view_id': self.env.ref('ms_templates.report_wizard').id,
                    'context': {'attachment_id': attachment.id}
                }
        else:
            faculty_ids = self.env['op.batch.line'].read_group([('batch_id.start_date', '>=', self.start_date),
                                                                ('batch_id.start_date', '<=', self.end_date),
                                                                ('batch_id.internal', '=', True)],
                                                               ['faculty_id', 'num_lessons', 'total_fee'],
                                                               ['faculty_id'])
            rec_data = []
            for dict in faculty_ids:
                faculty = self.env['op.faculty'].search([('id', '=', dict['faculty_id'][0])])
                survey = self.env['survey.user_input_line'].read_group(
                    [('user_input_id.batch_id.faculty_id', '=', faculty.id),
                     ('user_input_id.batch_id.start_date', '>=', self.start_date),
                     ('user_input_id.batch_id.start_date', '<=', self.end_date)],
                    ['question_id', 'value_suggested_int'], ['question_id'])
                ques_0 = round(survey[0]['value_suggested_int'] / survey[0]['question_id_count'],
                               1) if survey else False
                ques_1 = round(survey[1]['value_suggested_int'] / survey[1]['question_id_count'],
                               1) if survey else False
                ques_2 = round(survey[2]['value_suggested_int'] / survey[2]['question_id_count'],
                               1) if survey else False
                ques_3 = round(survey[3]['value_suggested_int'] / survey[3]['question_id_count'],
                               1) if survey else False
                ques_4 = round(survey[4]['value_suggested_int'] / survey[4]['question_id_count'],
                               1) if survey else False
                ques_5 = round(survey[5]['value_suggested_int'] / survey[5]['question_id_count'],
                               1) if survey else False
                ques_6 = round(survey[6]['value_suggested_int'] / survey[6]['question_id_count'],
                               1) if survey else False
                ques_7 = round(survey[7]['value_suggested_int'] / survey[7]['question_id_count'],
                               1) if survey else False
                ques_8 = round(survey[8]['value_suggested_int'] / survey[8]['question_id_count'],
                               1) if survey else False
                ques_9 = round(survey[9]['value_suggested_int'] / survey[9]['question_id_count'],
                               1) if survey else False
                if ques_0 and ques_1 and ques_2 and ques_3 and ques_4 and ques_5 and ques_6 and ques_7 and ques_8 and ques_9:
                    score = numpy.round(numpy.mean(
                        [ques_0, ques_1, ques_2, ques_3, ques_4, ques_5, ques_6, ques_7, ques_8, ques_9]), 2)
                else:
                    score = '-'
                rec_data.append({
                    'faculty_name': faculty.name,
                    'institute': faculty.institute.name,
                    'faculty_job': faculty.emp_id.job_id.name,
                    'faculty_department': faculty.emp_id.department_id.name,
                    'num_lessons': dict['num_lessons'],
                    'tutor_fee': dict['total_fee'] / dict['num_lessons'],
                    'total_fee': dict['total_fee'],
                    'ques_0': ques_0 or '-',
                    'ques_1': ques_1 or '-',
                    'ques_2': ques_2 or '-',
                    'ques_3': ques_3 or '-',
                    'ques_4': ques_4 or '-',
                    'ques_5': ques_5 or '-',
                    'ques_6': ques_6 or '-',
                    'ques_7': ques_7 or '-',
                    'ques_8': ques_8 or '-',
                    'ques_9': ques_9 or '-',
                    'score': score
                })
            simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
                self.env.ref('crm_academy.faculty_report_template').id)
            decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
            line_font = Font(name='Times New Roman', size=12)
            key_col_list = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]
            key_list = [
                'faculty_name',
                'institute',
                'faculty_job',
                'faculty_department',
                'num_lessons',
                'tutor_fee',
                'total_fee',
                'ques_0',
                'ques_1',
                'ques_2',
                'ques_3',
                'ques_4',
                'ques_5',
                'ques_6',
                'ques_7',
                'ques_8',
                'ques_9',
                'score',
            ]
            row = 7
            for line_data in rec_data:
                ws.cell(row, 1).value = row - 6
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col)
                    cell.value = line_data[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'Báo cáo đánh giá giảng viên',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            return {
                'name': 'Báo cáo đánh giá giảng viên',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }
