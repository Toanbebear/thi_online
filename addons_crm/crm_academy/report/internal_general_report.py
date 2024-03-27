import base64
import time as tm
import calendar
from pytz import timezone
from ast import literal_eval
from calendar import monthrange
from copy import copy
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from operator import itemgetter
from pytz import timezone, utc
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from odoo.tools import pycompat

import random

import logging

_logger = logging.getLogger(__name__)


class InternalGeneralReport(models.TransientModel):
    _name = 'internal.general.report'
    _description = 'Internal General Report'

    start_date = fields.Date('Date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')

    def internal_general_report(self):
        group_month = self.env['op.batch'].read_group(
            [('internal', '=', True), ('start_date', '>=', self.start_date), ('start_date', '<=', self.end_date)],
            ['name', 'num_lessons', 'num_students', 'learn_number'], ['start_date:month'])
        row = int(group_month[0]['start_date:month'][6]) + 5
        hn = self.env.ref('shealth_all_in_one.hv_hn_km_company').id
        hcm = self.env.ref('shealth_all_in_one.hv_hcm_km_company').id
        rec_data = []
        for line_month in group_month:
            group_company = self.env['op.batch'].read_group(line_month['__domain'],
                                                            ['name', 'num_lessons', 'num_students', 'learn_number'],
                                                            ['company_id'])
            data = {hn: {'company_id_count': 0, 'num_lessons': 0, 'num_students': 0, 'learn_number': 0},
                    hcm: {'company_id_count': 0, 'num_lessons': 0, 'num_students': 0, 'learn_number': 0}}
            for line_company in group_company:
                line_company_id = line_company['company_id'][0]
                data[line_company_id]['company_id_count'] = line_company['company_id_count']
                data[line_company_id]['num_lessons'] = line_company['num_lessons']
                data[line_company_id]['num_students'] = line_company['num_students']
                data[line_company_id]['learn_number'] = line_company['learn_number']
            rec_data.append({
                # 'month': line_month['start_date:month'],
                'batch_total': line_month['start_date_count'],
                'num_lessons_total': line_month['num_lessons'],
                'num_students_total': line_month['num_students'],
                'learn_number_total': line_month['learn_number'],
                'percent_total': str(
                    round((line_month['learn_number'] / line_month['num_students'] * 100), 2)) + '0%' if (
                        line_month['learn_number'] and line_month['learn_number']) else '0%',
                'batch_hn': data[hn]['company_id_count'],
                'num_lessons_hn': data[hn]['num_lessons'],
                'num_students_hn': data[hn]['num_students'],
                'learn_number_hn': data[hn]['learn_number'],
                'percent_hn': str(
                    round((data[hn]['learn_number'] / data[hn]['num_students'] * 100), 2)) + '0%' if (
                        data[hn]['learn_number'] and data[hn]['num_students']) else '0%',
                'num_lessons_hcm': data[hcm]['num_lessons'],
                'batch_hcm': data[hcm]['company_id_count'],
                'num_students_hcm': data[hcm]['num_students'],
                'learn_number_hcm': data[hcm]['learn_number'],
                'percent_hcm': str(
                    round((data[hcm]['learn_number'] / data[hcm]['num_students'] * 100), 2)) + '0%' if (
                        data[hcm]['learn_number'] and data[hcm]['num_students']) else '0%',
            })
        simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_academy.internal_general_report_template').id)
        decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['H3'].value = str(self.start_date.month) + '/' + str(self.start_date.year)
        ws['J3'].value = str(self.end_date.month) + '/' + str(self.end_date.year)
        ws['H3'].alignment = Alignment(horizontal='center', vertical='center')
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=12)
        # key_col_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
        key_col_list = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
        key_list = [
            # 'month',
            'batch_total',
            'num_lessons_total',
            'num_students_total',
            'learn_number_total',
            'percent_total',
            'batch_hn',
            'num_lessons_hn',
            'num_students_hn',
            'learn_number_hn',
            'percent_hn',
            'num_lessons_hcm',
            'batch_hcm',
            'num_students_hcm',
            'learn_number_hcm',
            'percent_hcm',
        ]
        for line_data in rec_data:
            # ws.cell(row, 1).value = row - (row - 1)
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo tổng hợp đào tạo nội bộ',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        return {
            'name': 'Báo cáo tổng hợp đào tạo nội bộ',
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'attachment_id': attachment.id}
        }
