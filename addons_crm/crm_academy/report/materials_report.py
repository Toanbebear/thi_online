import base64
import time as tm
import calendar
from pytz import timezone
from ast import literal_eval
from calendar import monthrange
from copy import copy
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from operator import itemgetter
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from odoo.tools import pycompat


import random

import logging

_logger = logging.getLogger(__name__)


class TeachingMaterialsReport(models.TransientModel):
    _name = 'materials.report'
    _description = 'Materials Report'

    report_type = fields.Selection([('teaching_materials', 'Teaching Materials'), ('gift_materials', 'Gift Materials')], 'Report type',
                                   default='teaching_materials')
    course = fields.Many2one('op.course', string='Course')
    batch = fields.Many2one('op.batch', string='Batch', domain="[('course', '=', course_id.id)]")

    def supplier_medicine_report(self):
        if self.report_type == 'teaching_materials':
            ret_data = []
            rec_data = []
            for record in self.batch.faculty_bom:
                rec_data.append({
                    'code_supplies': record.product.default_code or "-",
                    'name_product': record.product.name or "-",
                    'unit': record.uom_id.name or "-",
                    'amount': record.quantity  or "-",
                    'cost': record.cost or "-",
                    'cost_total': record.total_cost or "-",
                    'supplies_trainees': record.quantity / self.batch.num_students or "-",
                    'total_students': record.total_cost / self.batch.num_students or "-",

                })
            ret_data.append({
                'course_name': self.batch.course_id or "-",
                'name': self.batch.name or "-",
                'code_class': self.batch.code or "-",
                'study_time': self.batch.num_lessons or "-",
                'num_students': self.batch.num_students or "-",
            })
            simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
                self.env.ref('crm_academy.report_materials_template').id)
            decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            ws['A1'].value = 'Tên công ty: Công ty TNHH học viện y khoa thẩm mỹ quốc tế SCI'
            ws['A2'].value = 'Địa chỉ: 212 Kim Mã, Quận Ba Đình, Thành phố Hà Nội'
            # ws['A6'].value = 'BÁO CÁO VẬT TƯ TIÊU HAO THEO KHÓA HỌC'
            ws['B7'].value = self.course.name
            ws['B8'].value = self.batch.code
            ws['G8'].value = self.batch.name
            ws['B9'].value = str(self.batch.num_lessons) + (' Tiết')
            ws['B10'].value = self.batch.num_students

            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
            line_font = Font(name='Times New Roman', size=12)
            key_col_list = [ 2, 3, 4, 5, 6, 7, 8, 9]
            key_list = [
                'code_supplies',
                'name_product',
                'unit',
                'amount',
                'cost',
                'cost_total',
                'supplies_trainees',
                'total_students',
            ]
            row = 12
            for line_data in rec_data:
                ws.cell(row, 1).value = row - 11
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col)
                    cell.value = line_data[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'Báo cáo vật tư tiêu hao theo khóa học',
                # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            return {
                'name': 'Báo cáo vật tư tiêu hao theo khóa học',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }
        else:
            ret_data = []
            rec_data = []
            for record in self.batch.course_id.bom:
                for i in record.products:
                    rec_data.append({
                        'code_supplies': i.product.default_code or "-",
                        'name_product': i.product.name or "-",
                        'unit': i.product.uom_id.name or "-",
                        'amount': i.quantity * self.batch.num_students or "-",
                        'cost': i.cost or "-",
                        'cost_total': i.cost * (i.quantity * self.batch.num_students) or "-",
                        'gift_students': (i.quantity * self.batch.num_students) / self.batch.num_students or "-",
                        'total_students': i.cost * (i.quantity * self.batch.num_students) / self.batch.num_students or "-",
                    })
            ret_data.append({
                'course_name': self.batch.course_id.name or "-",
                'name_batch': self.batch.name or "-",
                'code_class': self.batch.code or "-",
                'study_time': self.batch.num_lessons or "-",
                'num_students': self.batch.num_students or "-",
            })
            simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
                self.env.ref('crm_academy.gift_materials_report_template').id)
            decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            ws['B1'].value = ' Công ty TNHH học viện y khoa thẩm mỹ quốc tế SCI'
            ws['B2'].value = ' 212 Kim Mã, Quận Ba Đình, Thành phố Hà Nội'
            # ws['A6'].value = 'BÁO CÁO VẬT TƯ TIÊU HAO THEO KHÓA HỌC'
            ws['B7'].value = self.course.name
            ws['B8'].value = self.batch.code
            ws['E8'].value = self.batch.name
            ws['B9'].value = str(self.batch.num_lessons) + (' Tiết')
            ws['B10'].value = self.batch.num_students

            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
            line_font = Font(name='Times New Roman', size=12)
            key_col_list = [2, 3, 4, 5, 6, 7, 8, 9]
            key_list = [
                'code_supplies',
                'name_product',
                'unit',
                'amount',
                'cost',
                'cost_total',
                'gift_students',
                'total_students',
            ]
            row = 12
            for line_data in rec_data:
                ws.cell(row, 1).value = row - 11
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col)
                    cell.value = line_data[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'Báo cáo đồ dùng tặng học viên / 1 khóa học',
                # 'datas_fname': 'bao_cao_vat_tu_theo_khoa_hoc.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            return {
                'name': 'Báo cáo đồ dùng tặng học viên / 1 khóa học',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
            }
