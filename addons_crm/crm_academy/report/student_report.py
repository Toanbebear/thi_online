import time as tm
import calendar
from pytz import timezone
from ast import literal_eval
from calendar import monthrange
from copy import copy
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from operator import itemgetter
from pytz import timezone, utc
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from odoo.tools import pycompat
import numpy
import base64

import random

import logging

_logger = logging.getLogger(__name__)


class StudentReport(models.TransientModel):
    _name = 'student.report'
    _description = 'Student Report'

    start_date = fields.Date('Date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    type_report = fields.Selection([('general', 'General student report'), ('detail', 'Detail student report')],
                                   string='Type report', default='general')
    department_id = fields.Many2many('hr.department', string='Department')
    job_id = fields.Many2many('hr.job', string='Job')
    company_id = fields.Many2many('res.company', string='Company')
    student_id = fields.Many2many('op.student', string='Student')

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.Date.today().month:
                self.end_date = fields.Date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.onchange('company_id', 'department_id', 'job_id')
    def _get_student(self):
        domain = [('internal', '=', 'True')]
        if self.company_id:
            domain += [('company_id', 'in', self.company_id.ids)]
        if self.department_id:
            domain += [('department_id', 'in', self.department_id.ids)]
        if self.job_id:
            domain += [('job_id', 'in', self.job_id.ids)]
        return {'domain': {'student_id': [('id', 'in', self.env['op.student'].search(domain).ids)]}}

    def student_report(self):
        if self.type_report == 'general':
            if self.student_id:
                rec_data = []
                for student in self.student_id:
                    arr = []
                    for per in student.course_detail_ids:
                        if per.percentage:
                            score = int(per.percentage.replace(".0", ""))
                            arr.append(score)
                    rec_data.append({
                        'student_name': student.name,
                        'job': student.job_id.name,
                        'department': student.department_id.name,
                        'city': student.hometown or "-",
                        'course_number': len(student.course_detail_ids) or "-",
                        'num_lessons_total': sum(
                            course.num_lessons for course in student.course_detail_ids.course_id) or "-",
                        'medium_score': numpy.mean(arr),
                    })
                simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
                    self.env.ref('crm_academy.internal_general_student_report_template').id)
                decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
                wb = load_workbook(BytesIO(decode))
                ws = wb.active
                thin = borders.Side(style='thin')
                all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
                line_font = Font(name='Times New Roman', size=12)
                key_col_list = [2, 3, 4, 5, 6, 7, 8]
                key_list = [
                    'student_name',
                    'job',
                    'department',
                    'city',
                    'course_number',
                    'num_lessons_total',
                    'medium_score',
                ]
                row = 5
                for line_data in rec_data:
                    ws.cell(row, 1).value = row - 4
                    for col, k in zip(key_col_list, key_list):
                        cell = ws.cell(row, col)
                        cell.value = line_data[k]
                        cell.font = line_font
                        cell.border = all_border_thin
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                fp = BytesIO()
                wb.save(fp)
                fp.seek(0)
                report = base64.encodebytes((fp.read()))
                fp.close()
                attachment = self.env['ir.attachment'].sudo().create({
                    'name': 'Báo cáo tổng hợp theo học viên',
                    'datas': report,
                    'res_model': 'temp.creation',
                    'public': True,
                })
                return {
                    'name': 'Báo cáo tổng hợp theo học viên',
                    'type': 'ir.actions.act_window',
                    'res_model': 'temp.wizard',
                    'view_mode': 'form',
                    'view_type': 'form',
                    'target': 'inline',
                    'view_id': self.env.ref('ms_templates.report_wizard').id,
                    'context': {'attachment_id': attachment.id}
                }
        else:
            if self.student_id:
                rec_data = []
                for student in self.student_id:
                    course = []
                    num_lessons = []
                    score = []
                    start_date = []
                    if student.course_detail_ids:
                        for rec in student.course_detail_ids:
                            print(rec.batch_id.start_date)
                            if rec.batch_id.start_date >= self.start_date and rec.batch_id.start_date <= self.end_date:
                                start_date.append(rec.batch_id.start_date.strftime("%d/%m/%Y"))
                                course.append(rec.course_id.name)
                                num_lessons.append(rec.batch_id.num_lessons)
                                score.append(rec.percentage)
                    rec_data.append({
                        'student_name': student.name,
                        'job_id': student.job_id.name,
                        'department_id': student.department_id.name,
                        'city': student.hometown or "-",
                        'course': course or "-",
                        'start_date': start_date or "-",
                        'num_lessons': num_lessons or "-",
                        'score': score or "-"
                    })
                simple_pharmacy_provider_attachment = self.env['ir.attachment'].browse(
                    self.env.ref('crm_academy.internal_detailed_student_report_template').id)
                decode = base64.b64decode(simple_pharmacy_provider_attachment.datas)
                wb = load_workbook(BytesIO(decode))
                ws = wb.active
                thin = borders.Side(style='thin')
                all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
                line_font = Font(name='Times New Roman', size=12)
                key_col_list = [2, 3, 4, 5, 6, 7, 8, 9]
                key_list = [
                    'student_name',
                    'job_id',
                    'department_id',
                    'city',
                    'course',
                    'start_date',
                    'num_lessons',
                    'score',
                ]
                row = 5
                for line_data in rec_data:
                    ws.cell(row, 1).value = row - 4
                    for col, k in zip(key_col_list, key_list):
                        list_len = []
                        if isinstance(line_data[k], list):
                            list_len.append(len(line_data[k]))
                            for sub_row in range(len(line_data[k])):
                                cell = ws.cell(row + sub_row, col)
                                cell.value = line_data[k][sub_row]
                                cell.font = line_font
                                cell.border = all_border_thin
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            list_len.append(1)
                            cell = ws.cell(row, col)
                            cell.value = line_data[k]
                            cell.font = line_font
                            cell.border = all_border_thin
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += max(list_len)
                fp = BytesIO()
                wb.save(fp)
                fp.seek(0)
                report = base64.encodebytes((fp.read()))
                fp.close()
                attachment = self.env['ir.attachment'].sudo().create({
                    'name': 'Báo cáo chi tiết theo học viên',
                    'datas': report,
                    'res_model': 'temp.creation',
                    'public': True,
                })
                return {
                    'name': 'Báo cáo chi tiết theo học viên',
                    'type': 'ir.actions.act_window',
                    'res_model': 'temp.wizard',
                    'view_mode': 'form',
                    'view_type': 'form',
                    'target': 'inline',
                    'view_id': self.env.ref('ms_templates.report_wizard').id,
                    'context': {'attachment_id': attachment.id}
                }
